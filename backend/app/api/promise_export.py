"""Excel export for the Promise-Dates page — one card, or the whole page.

Every date column carries its basis: a promise built on an inbound purchase is
an ESTIMATE (CRM has no expected-arrival date), and the "Estimated" column says
so per row. A promise that would take stock below the safety level is flagged
rather than withheld.
"""
from __future__ import annotations

import io

SECTION_TITLES = {
    "summary": "Summary",
    "classes": "Promise status",
    "items": "All items",
    "late": "Slipping past requirement",
    "nodate": "No dated supply",
    "risk": "Running out soonest",
}


def _row(r: dict) -> dict:
    return {"Item code": r.get("item_code"),
            "Item": r.get("item"),
            "Segment": r.get("segment3"),
            "Promise status": r.get("class"),
            "Quantity needed (KG)": r.get("need"),
            "Required by": r.get("required") or "",
            "Can promise from": r.get("ctp") or "",
            "Slip (days)": r.get("slip_days"),
            "Stock runs out": r.get("risk_date") or "",
            "Days to risk": r.get("days_to_risk"),
            "On hand (KG)": r.get("on_hand"),
            "Safety level (KG)": r.get("msl"),
            "Breaches safety level": "yes" if r.get("breaches_msl") else "",
            "Firm orders, company (KG)": r.get("firm_total"),
            "Incoming production (KG)": r.get("incoming"),
            "Supply sources": ", ".join(r.get("supply_sources") or []) or "none",
            "Estimated date": "yes" if r.get("estimated") else "",
            "Closing balance (KG)": r.get("closing"),
            "Lowest balance (KG)": r.get("low"),
            "My projection (KG)": r.get("my_projection"),
            "Projected company-wide (KG)": r.get("all_projection")}


def section_rows(payload: dict, rows: list[dict], section: str) -> list[dict]:
    k = payload.get("kpis") or {}
    if section == "summary":
        return [
            {"Metric": "Persona", "Value": payload.get("persona") or "—"},
            {"Metric": "Scope", "Value": "; ".join(payload.get("scope") or []) or "—"},
            {"Metric": "Cycle", "Value": f"{payload.get('jc_label')} "
                                         f"({payload.get('jc_from')} to {payload.get('jc_to')})"},
            {"Metric": "As of", "Value": payload.get("today")},
            {"Metric": "Items in view", "Value": k.get("items")},
            {"Metric": "Quantity still to convert (KG)", "Value": k.get("need")},
            {"Metric": "Items that can be promised", "Value": k.get("promised")},
            {"Metric": "Items whose promise slips", "Value": k.get("late")},
            {"Metric": "  quantity affected (KG)", "Value": k.get("slipping_qty")},
            {"Metric": "  worst slip (days)", "Value": k.get("worst_slip")},
            {"Metric": "  average slip (days)", "Value": k.get("avg_slip")},
            {"Metric": "Items with NO dated supply", "Value": k.get("no_date")},
            {"Metric": "  quantity affected (KG)", "Value": k.get("no_date_qty")},
            {"Metric": "Items running out within 14 days", "Value": k.get("running_out")},
            {"Metric": "Promises that dip below the safety level", "Value": k.get("breaching_msl")},
            {"Metric": "Items whose date rests on an ESTIMATED PO arrival",
             "Value": k.get("estimated_items")},
            {"Metric": "Forward horizon (days)", "Value": payload.get("horizon_days")},
            {"Metric": "Requirement dated to (days per half-cycle)",
             "Value": payload.get("half_cycle_days")},
            {"Metric": "Safety level source", "Value": payload.get("msl_ref") or "—"},
            {"Metric": "Production plan used", "Value": payload.get("plan_id") or "no saved plan"},
            {"Metric": "Data as of",
             "Value": str((payload.get("last_sync") or {}).get("finished_at") or "—")},
        ]
    if section == "classes":
        return [{"Promise status": b["label"], "Items": b["items"],
                 "Quantity (KG)": b["qty"]} for b in k.get("buckets") or []]
    if section == "items":
        return [_row(r) for r in rows]
    if section == "late":
        return [_row(r) for r in rows if r.get("class") == "late"]
    if section == "nodate":
        return [_row(r) for r in rows if r.get("class") == "none"]
    if section == "risk":
        rs = [r for r in rows if r.get("days_to_risk") is not None]
        rs.sort(key=lambda r: r["days_to_risk"])
        return [_row(r) for r in rs]
    return []


def _write(ws, rows: list[dict]) -> int:
    if not rows:
        ws.append(["No data in scope"])
        return 0
    cols = list(rows[0].keys())
    ws.append(cols)
    for r in rows:
        ws.append([r.get(c) for c in cols])
    for i, c in enumerate(cols, start=1):
        width = max(len(str(c)), *(len(str(r.get(c) or "")) for r in rows[:200])) + 2
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(44, max(9, width))
    ws.freeze_panes = "A2"
    return len(rows)


def build(payload: dict, rows: list[dict], section: str | None = None) -> bytes:
    import openpyxl
    from openpyxl.chart import BarChart, DoughnutChart, Reference

    wb = openpyxl.Workbook()
    if section:
        ws = wb.active
        ws.title = SECTION_TITLES.get(section, "Data")[:31]
        _write(ws, section_rows(payload, rows, section))
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    charts = wb.active
    charts.title = "Charts"
    charts["A1"] = "Promise Dates"
    charts["A1"].font = openpyxl.styles.Font(size=14, bold=True)
    charts["A2"] = f"{payload.get('persona') or ''} · {'; '.join(payload.get('scope') or [])}"
    charts["A3"] = (f"{payload.get('jc_label')} · as of {payload.get('today')}"
                    f" · plan {payload.get('plan_id') or '—'}"
                    f" · data {((payload.get('last_sync') or {}).get('finished_at') or '—')}")

    counts = {}
    for key in ("summary", "classes", "late", "nodate", "risk", "items"):
        ws = wb.create_sheet(SECTION_TITLES[key][:31])
        counts[key] = _write(ws, section_rows(payload, rows, key))

    def sheet(key):
        return wb[SECTION_TITLES[key][:31]]

    def col_of(key, header):
        hdr = [c.value for c in next(sheet(key).iter_rows(min_row=1, max_row=1))]
        return hdr.index(header) + 1 if header in hdr else 1

    placed = []
    if counts["classes"]:
        ch = DoughnutChart(holeSize=55)
        ch.title = "Items by promise status"
        ch.add_data(Reference(sheet("classes"), min_col=2, min_row=1,
                              max_row=counts["classes"] + 1), titles_from_data=True)
        ch.set_categories(Reference(sheet("classes"), min_col=1, min_row=2,
                                    max_row=counts["classes"] + 1))
        ch.height, ch.width = 8, 16
        placed.append(ch)
    if counts["late"]:
        n = min(counts["late"], 15)
        ch = BarChart()
        ch.type = "bar"
        ch.title = "Slip against the requirement (days)"
        ch.legend = None
        ch.add_data(Reference(sheet("late"), min_col=col_of("late", "Slip (days)"),
                              min_row=1, max_row=n + 1), titles_from_data=True)
        ch.set_categories(Reference(sheet("late"), min_col=col_of("late", "Item"),
                                    min_row=2, max_row=n + 1))
        ch.height, ch.width = 9, 16
        placed.append(ch)
    if counts["risk"]:
        n = min(counts["risk"], 15)
        ch = BarChart()
        ch.type = "bar"
        ch.title = "Days until the stock runs out"
        ch.legend = None
        ch.add_data(Reference(sheet("risk"), min_col=col_of("risk", "Days to risk"),
                              min_row=1, max_row=n + 1), titles_from_data=True)
        ch.set_categories(Reference(sheet("risk"), min_col=col_of("risk", "Item"),
                                    min_row=2, max_row=n + 1))
        ch.height, ch.width = 9, 16
        placed.append(ch)
    for ch, anchor in zip(placed, ["A5", "K5", "A24", "K24"]):
        charts.add_chart(ch, anchor)
    charts.column_dimensions["A"].width = 36

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


__all__ = ["build", "section_rows", "SECTION_TITLES"]
