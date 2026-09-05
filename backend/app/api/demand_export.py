"""Excel export for the Demand-Protection page — one card, or the whole page.

Mirrors dashboard_export / commit_export: native Excel charts on the first sheet
(linked to the data sheets), one sheet per table. The line sheet is UNCAPPED —
the page shows the worst 500 by unprotected quantity, the workbook carries every
projection line in scope.
"""
from __future__ import annotations

import io

SECTION_TITLES = {
    "summary": "Summary",
    "trend": "Protection by JC",
    "collectors": "By collector",
    "customers": "By customer",
    "items": "By item",
    "lines": "All projection lines",
    "unprotected": "Unprotected lines",
    "silent": "No order, no dispatch",
}


def _line_row(r: dict) -> dict:
    return {"Customer": r.get("customer"),
            "Collector": r.get("collector"),
            "MC": r.get("mc_code"),
            "Item code": r.get("item_code"),
            "Item": r.get("item"),
            "Segment": r.get("segment3"),
            "Projected (KG)": r.get("projected"),
            "Week 1": r.get("week1"),
            "Week 2": r.get("week2"),
            "Dispatched (KG)": r.get("dispatched"),
            "Open SOC (KG)": r.get("soc"),
            "Open SOC lines": r.get("soc_lines"),
            "Protected (KG)": r.get("covered"),
            "Unprotected (KG)": r.get("unprotected"),
            "Protected %": r.get("pct"),
            "Nothing raised": "yes" if r.get("silent") else "",
            "Overdue backlog (KG)": r.get("backlog"),
            "Next JC": r.get("next1"),
            "JC after": r.get("next2")}


def _group_row(r: dict, label: str, title: str) -> dict:
    out = {title: r.get(label)}
    if label == "customer":
        out["Collector"] = r.get("collector")
        out["MC"] = r.get("mc_code")
    if label == "item":
        out["Item code"] = r.get("item_code")
        out["Segment"] = r.get("segment3")
    out.update({"Projected (KG)": r.get("projected"),
                "Dispatched (KG)": r.get("dispatched"),
                "Open SOC (KG)": r.get("soc"),
                "Protected (KG)": r.get("protected"),
                "Unprotected (KG)": r.get("unprotected"),
                "Protected %": r.get("pct"),
                "Lines": r.get("lines"),
                "Lines with nothing raised": r.get("silent_lines")})
    return out


def section_rows(payload: dict, rows: list[dict], section: str) -> list[dict]:
    k = payload.get("kpis") or {}
    if section == "summary":
        return [
            {"Metric": "Persona", "Value": payload.get("persona") or "—"},
            {"Metric": "Scope", "Value": "; ".join(payload.get("scope") or []) or "—"},
            {"Metric": "Cycle", "Value": f"{payload.get('jc_label')} "
                                         f"({payload.get('jc_from')} to {payload.get('jc_to')})"},
            {"Metric": "Projected (KG)", "Value": k.get("projected")},
            {"Metric": "Protected (KG)", "Value": k.get("protected")},
            {"Metric": "Unprotected (KG)", "Value": k.get("unprotected")},
            {"Metric": "Protected %", "Value": k.get("protection_pct")},
            {"Metric": "  of which already dispatched (KG)", "Value": k.get("dispatched")},
            {"Metric": "  of which open committed orders (KG)", "Value": k.get("soc")},
            {"Metric": "Ordered above projection (KG)", "Value": k.get("over")},
            {"Metric": "Overdue backlog, not counted as cover (KG)", "Value": k.get("backlog")},
            {"Metric": "Projection lines", "Value": k.get("lines")},
            {"Metric": "  fully covered", "Value": k.get("full_lines")},
            {"Metric": "  with no order and no dispatch", "Value": k.get("silent_lines")},
            {"Metric": "  their projected qty (KG)", "Value": k.get("silent_qty")},
            {"Metric": "Customers", "Value": k.get("customers")},
            {"Metric": "Items", "Value": k.get("items")},
            {"Metric": "Dispatch available for this cycle",
             "Value": "yes" if payload.get("has_dispatch") else "no — future cycle"},
            {"Metric": "Data as of",
             "Value": str((payload.get("last_sync") or {}).get("finished_at") or "—")},
        ]
    if section == "trend":
        return [{"Cycle": t["label"], "Projected (KG)": t["projected"],
                 "Dispatched (KG)": t["dispatched"], "Open SOC (KG)": t["soc"],
                 "Protected (KG)": t["protected"], "Unprotected (KG)": t["unprotected"],
                 "Protected %": t["pct"]} for t in payload.get("trend") or []]
    if section == "collectors":
        return [_group_row(r, "collector", "Collector") for r in payload.get("by_collector") or []]
    if section == "customers":
        return [_group_row(r, "customer", "Customer") for r in payload.get("by_customer") or []]
    if section == "items":
        return [_group_row(r, "item", "Item") for r in payload.get("by_item") or []]
    if section == "lines":
        return [_line_row(r) for r in rows]
    if section == "unprotected":
        return [_line_row(r) for r in rows if (r.get("unprotected") or 0) > 0]
    if section == "silent":
        return [_line_row(r) for r in rows if r.get("silent")]
    return []


def _write(ws, rows: list[dict]) -> int:
    if not rows:
        ws.append(["No data in scope"])
        return 0
    cols = list(rows[0].keys())
    ws.append(cols)
    for r in rows:
        ws.append([r.get(c) for c in cols])
    for i, c in enumerate(cols, start=1):
        width = max(len(str(c)), *(len(str(r.get(c) or "")) for r in rows[:200])) + 2
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(44, max(9, width))
    ws.freeze_panes = "A2"
    return len(rows)


def build(payload: dict, rows: list[dict], section: str | None = None) -> bytes:
    import openpyxl
    from openpyxl.chart import BarChart, Reference

    wb = openpyxl.Workbook()
    if section:
        ws = wb.active
        ws.title = SECTION_TITLES.get(section, "Data")[:31]
        _write(ws, section_rows(payload, rows, section))
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    charts = wb.active
    charts.title = "Charts"
    charts["A1"] = "Demand Protection"
    charts["A1"].font = openpyxl.styles.Font(size=14, bold=True)
    charts["A2"] = f"{payload.get('persona') or ''} · {'; '.join(payload.get('scope') or [])}"
    charts["A3"] = (f"{payload.get('jc_label')} · {payload.get('jc_from')} to {payload.get('jc_to')}"
                    f" · data {((payload.get('last_sync') or {}).get('finished_at') or '—')}")

    counts = {}
    for key in ("summary", "trend", "collectors", "customers", "items",
                "unprotected", "silent", "lines"):
        ws = wb.create_sheet(SECTION_TITLES[key][:31])
        counts[key] = _write(ws, section_rows(payload, rows, key))

    def sheet(key):
        return wb[SECTION_TITLES[key][:31]]

    placed = []
    if counts["trend"]:
        ch = BarChart()
        ch.type = "col"
        ch.grouping = "stacked"
        ch.overlap = 100
        ch.title = "Protected vs unprotected by cycle"
        ch.add_data(Reference(sheet("trend"), min_col=5, max_col=6, min_row=1,
                              max_row=counts["trend"] + 1), titles_from_data=True)
        ch.set_categories(Reference(sheet("trend"), min_col=1, min_row=2,
                                    max_row=counts["trend"] + 1))
        ch.height, ch.width = 8, 16
        placed.append(ch)
    if counts["collectors"]:
        n = min(counts["collectors"], 15)
        ch = BarChart()
        ch.type = "bar"
        ch.title = "Unprotected projection by collector"
        ch.legend = None
        ch.add_data(Reference(sheet("collectors"), min_col=6, min_row=1,
                              max_row=n + 1), titles_from_data=True)
        ch.set_categories(Reference(sheet("collectors"), min_col=1, min_row=2, max_row=n + 1))
        ch.height, ch.width = 9, 16
        placed.append(ch)
    if counts["items"]:
        n = min(counts["items"], 15)
        ch = BarChart()
        ch.type = "bar"
        ch.title = "Unprotected projection by item"
        ch.legend = None
        ch.add_data(Reference(sheet("items"), min_col=7, min_row=1,
                              max_row=n + 1), titles_from_data=True)
        ch.set_categories(Reference(sheet("items"), min_col=1, min_row=2, max_row=n + 1))
        ch.height, ch.width = 9, 16
        placed.append(ch)
    for ch, anchor in zip(placed, ["A5", "K5", "A24", "K24"]):
        charts.add_chart(ch, anchor)
    charts.column_dimensions["A"].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


__all__ = ["build", "section_rows", "SECTION_TITLES"]
