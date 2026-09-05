"""Excel export for My Supply Position — the headline, the action list, and the
per-row derivation behind the "why is this at risk" panel.

Supply figures (other SOC, available, commit date) are the ITEM's, shared across
every customer row for that item — the same way the board shows them.
"""
from __future__ import annotations

import io

SECTION_TITLES = {
    "summary": "Summary",
    "action": "Action required",
    "critical": "Critical",
    "watch": "At risk",
    "why": "Why at risk",
    "collectors": "Exposure by collector",
    "items": "Exposure by item",
    "silent": "Nothing raised yet",
    "supply": "Supply behind the dates",
    "competing": "Competing demand",
}


def _row(r: dict) -> dict:
    return {"Risk": r.get("risk"),
            "Item code": r.get("item_code"),
            "Item": r.get("item"),
            "Customer": r.get("customer"),
            "Collector": r.get("collector"),
            "MC": r.get("mc_code"),
            "Projection (KG)": r.get("projection"),
            "SOC (KG)": r.get("soc"),
            "  of which open orders": r.get("my_soc"),
            "  of which dispatched": r.get("dispatched"),
            "Protected (KG)": r.get("protected"),
            "Unprotected (KG)": r.get("unprotected"),
            "Overdue backlog (KG)": r.get("backlog"),
            "Other SOC — item (KG)": r.get("other_soc"),
            "Other customers": r.get("other_customers"),
            "On hand — item (KG)": r.get("on_hand"),
            "ATP — item (KG)": r.get("atp"),
            "Expected shortage (KG)": r.get("shortfall"),
            "Required date": r.get("required") or "",
            "Commit date": r.get("commit_date") or "",
            "Potential delay (days)": r.get("delay_days"),
            "Promise dips below safety level": "yes" if r.get("breaches_msl") else "",
            "Incoming production (KG)": r.get("incoming"),
            "Production from": r.get("incoming_date") or "",
            "Stock runs out": r.get("risk_date") or ""}


def _why(r: dict) -> list[dict]:
    """The derivation panel, one metric per line — the same order it is shown."""
    return [
        {"Item": r.get("item"), "Customer": r.get("customer"),
         "Metric": "Your projection", "Value": r.get("projection")},
        {"Item": r.get("item"), "Customer": r.get("customer"),
         "Metric": "Your SOC", "Value": r.get("soc")},
        {"Item": r.get("item"), "Customer": r.get("customer"),
         "Metric": "Unprotected", "Value": r.get("unprotected")},
        {"Item": r.get("item"), "Customer": r.get("customer"),
         "Metric": "Other executive SOC", "Value": r.get("other_soc")},
        {"Item": r.get("item"), "Customer": r.get("customer"),
         "Metric": "Available after firm commitments", "Value": r.get("atp")},
        {"Item": r.get("item"), "Customer": r.get("customer"),
         "Metric": "Expected shortage", "Value": r.get("shortfall")},
        {"Item": r.get("item"), "Customer": r.get("customer"),
         "Metric": "Required date", "Value": r.get("required") or "—"},
        {"Item": r.get("item"), "Customer": r.get("customer"),
         "Metric": "Expected commitment date", "Value": r.get("commit_date") or "no dated supply"},
        {"Item": r.get("item"), "Customer": r.get("customer"),
         "Metric": "Potential delay (days)",
         "Value": r.get("delay_days") if r.get("delay_days") is not None else "—"},
    ]


def section_rows(payload: dict, rows: list[dict], section: str) -> list[dict]:
    k = payload.get("kpis") or {}
    if section == "summary":
        return [
            {"Metric": "Persona", "Value": payload.get("persona") or "—"},
            {"Metric": "Scope", "Value": "; ".join(payload.get("scope") or []) or "—"},
            {"Metric": "Cycle", "Value": f"{payload.get('jc_label')} "
                                         f"({payload.get('jc_from')} to {payload.get('jc_to')})"},
            {"Metric": "As of", "Value": payload.get("today")},
            {"Metric": "Projection — total demand (KG)", "Value": k.get("projection")},
            {"Metric": "SOC — firm demand (KG)", "Value": k.get("soc")},
            {"Metric": "Protected — covered quantity (KG)", "Value": k.get("protected")},
            {"Metric": "Overdue backlog — owed but committed before this cycle (KG)",
             "Value": k.get("backlog")},
            {"Metric": "  lines carrying it", "Value": k.get("backlog_lines")},
            {"Metric": "Your open order book — lines", "Value": (k.get("book") or {}).get("lines")},
            {"Metric": "  committed inside this cycle (KG)",
             "Value": (k.get("book") or {}).get("in_cycle")},
            {"Metric": "  committed before it — overdue (KG)",
             "Value": (k.get("book") or {}).get("overdue")},
            {"Metric": "  committed after it (KG)", "Value": (k.get("book") or {}).get("later")},
            {"Metric": "At risk — exposure (KG)", "Value": k.get("at_risk")},
            {"Metric": "Critical — shortage (KG)", "Value": k.get("critical")},
            {"Metric": "Protected %", "Value": k.get("protection_pct")},
            {"Metric": "Customer-item lines", "Value": k.get("lines")},
            {"Metric": "Lines requiring action", "Value": k.get("action_lines")},
            {"Metric": "  with no promisable date", "Value": k.get("no_date")},
            {"Metric": "  worst delay (days)", "Value": k.get("worst_delay")},
            {"Metric": "  promises that dip below the safety level", "Value": k.get("below_msl")},
            {"Metric": "Customers", "Value": k.get("customers")},
            {"Metric": "Items", "Value": k.get("items")},
            {"Metric": "Data as of",
             "Value": str((payload.get("last_sync") or {}).get("finished_at") or "—")},
        ]
    if section == "action":
        return [_row(r) for r in rows]
    if section == "critical":
        return [_row(r) for r in rows if r.get("risk") == "critical"]
    if section == "watch":
        return [_row(r) for r in rows if r.get("risk") == "watch"]
    if section in ("collectors", "items"):
        label, title = (("collector", "Collector") if section == "collectors"
                        else ("item", "Item"))
        out = []
        for r in payload.get("by_collector" if section == "collectors" else "by_item") or []:
            row = {title: r.get(label)}
            if section == "items":
                row["Item code"] = r.get("item_code")
                row["Segment"] = r.get("segment3")
            row.update({"Projected (KG)": r.get("projected"),
                        "Dispatched (KG)": r.get("dispatched"),
                        "Open SOC (KG)": r.get("soc"),
                        "Protected (KG)": r.get("protected"),
                        "Unprotected (KG)": r.get("unprotected"),
                        "Protected %": r.get("pct"),
                        "Lines": r.get("lines"),
                        "Lines with nothing raised": r.get("silent_lines")})
            out.append(row)
        return out
    if section == "silent":
        return [_row(r) for r in rows if r.get("silent")]
    if section == "supply":
        return [{"Item code": r.get("item_code"), "Item": r.get("item"),
                 "Segment": r.get("segment3"),
                 "On hand (KG)": r.get("on_hand"),
                 "Safety level (KG)": r.get("msl"),
                 "Firm orders — others (KG)": r.get("firm_others"),
                 "My unprotected (KG)": r.get("my_unprotected"),
                 "Incoming production (KG)": r.get("incoming"),
                 "Production from": r.get("incoming_date") or "",
                 "ATP (KG)": r.get("atp"),
                 "Exposure (KG)": r.get("exposure"),
                 "Stock runs out": r.get("risk_date") or "",
                 "Days to risk": r.get("days_to_risk"),
                 "Dated supply from": ", ".join(r.get("sources") or []) or "none",
                 "Date is estimated": "yes" if r.get("estimated") else ""}
                for r in payload.get("items") or []]
    if section == "competing":
        out = []
        for r in payload.get("competing_by_collector") or []:
            out.append({"Grouped by": "Collector", "Name": r.get("collector"),
                        "Committed balance (KG)": r.get("balance"),
                        "Order lines": r.get("lines"), "Customers": r.get("customers"),
                        "Items": r.get("items")})
        for r in payload.get("competing_by_mc") or []:
            out.append({"Grouped by": "Market circle", "Name": r.get("mc_code"),
                        "Committed balance (KG)": r.get("balance"),
                        "Order lines": r.get("lines"), "Customers": r.get("customers"),
                        "Items": r.get("items")})
        return out
    if section == "why":
        out = []
        for r in rows[:60]:
            out.extend(_why(r))
        return out
    return []


def _write(ws, rows: list[dict]) -> int:
    if not rows:
        ws.append(["Nothing requires action in this scope"])
        return 0
    cols = list(rows[0].keys())
    ws.append(cols)
    for r in rows:
        ws.append([r.get(c) for c in cols])
    for i, c in enumerate(cols, start=1):
        width = max(len(str(c)), *(len(str(r.get(c) or "")) for r in rows[:200])) + 2
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(44, max(9, width))
    ws.freeze_panes = "A2"
    return len(rows)


def build(payload: dict, rows: list[dict], section: str | None = None) -> bytes:
    import openpyxl
    from openpyxl.chart import BarChart, Reference

    wb = openpyxl.Workbook()
    if section:
        ws = wb.active
        ws.title = SECTION_TITLES.get(section, "Data")[:31]
        _write(ws, section_rows(payload, rows, section))
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    charts = wb.active
    charts.title = "Headline"
    k = payload.get("kpis") or {}
    charts["A1"] = f"My Supply Position — {payload.get('jc_label') or ''}"
    charts["A1"].font = openpyxl.styles.Font(size=14, bold=True)
    charts["A2"] = f"{payload.get('persona') or ''} · {'; '.join(payload.get('scope') or [])}"
    charts["A3"] = f"As of {payload.get('today')}"
    headline = [("Projection", "Total demand", k.get("projection")),
                ("SOC", "Firm demand", k.get("soc")),
                ("Protected", "Covered qty", k.get("protected")),
                ("At risk", "Exposure", k.get("at_risk")),
                ("Critical", "Shortage", k.get("critical"))]
    charts["A5"], charts["B5"], charts["C5"] = "Measure", "Means", "KG"
    for i, (a, b, c) in enumerate(headline, start=6):
        charts[f"A{i}"], charts[f"B{i}"], charts[f"C{i}"] = a, b, c
    charts.column_dimensions["A"].width = 18
    charts.column_dimensions["B"].width = 20
    charts.column_dimensions["C"].width = 16

    counts = {}
    for key in ("summary", "action", "critical", "watch", "why"):
        ws = wb.create_sheet(SECTION_TITLES[key][:31])
        counts[key] = _write(ws, section_rows(payload, rows, key))

    ch = BarChart()
    ch.type = "col"
    ch.title = "Where the projection stands"
    ch.legend = None
    ch.add_data(Reference(charts, min_col=3, min_row=5, max_row=10), titles_from_data=True)
    ch.set_categories(Reference(charts, min_col=1, min_row=6, max_row=10))
    ch.height, ch.width = 8, 16
    charts.add_chart(ch, "E5")

    if counts["action"]:
        sh = wb[SECTION_TITLES["action"][:31]]
        hdr = [c.value for c in next(sh.iter_rows(min_row=1, max_row=1))]
        n = min(counts["action"], 15)
        ch2 = BarChart()
        ch2.type = "bar"
        ch2.title = "Biggest expected shortages"
        ch2.legend = None
        ch2.add_data(Reference(sh, min_col=hdr.index("Expected shortage (KG)") + 1,
                               min_row=1, max_row=n + 1), titles_from_data=True)
        ch2.set_categories(Reference(sh, min_col=hdr.index("Item") + 1,
                                     min_row=2, max_row=n + 1))
        ch2.height, ch2.width = 9, 16
        charts.add_chart(ch2, "E24")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


__all__ = ["build", "section_rows", "SECTION_TITLES"]
