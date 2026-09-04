"""Excel export for My Dashboard — one card, or the whole page.

Two shapes, both built from exactly the payload the page renders (so a download
always matches what the user is looking at, scope included):

  * one section  -> a single sheet with that card's table
  * whole page   -> a "Charts" sheet of native Excel charts, then one sheet per
                    table behind them

The charts are real Excel chart objects rather than pictures: they stay linked
to the data sheets, so the reader can resize them, change the range or re-plot.
(Embedding the on-screen ECharts images would need Pillow, which isn't
installed, and would hand the reader a flat bitmap.)
"""
from __future__ import annotations

import io

# One entry per exportable card: sheet title + the rows/columns it writes.
# `key` matches the `section` query parameter the download buttons send.
SECTION_TITLES = {
    "summary": "Summary",
    "collector": "By collector",
    "segment": "By segment",
    "jc_trend": "Projection by JC",
    "status": "Projection status",
    "items": "Items by status",
    "groups": "Item groups",
    "compare": "Projection vs sales",
    "pipeline": "Projection pipeline",
    "missing": "Missing projections",
}

_FLAG_LABEL = {"ontrack": "On-track", "over": "Over-projected",
               "under": "Under-projected", "none": "No projection",
               "new": "New (no sales yet)"}


def _num(v):
    return 0 if v is None else v


def _cube_by(payload: dict, field: str) -> list[dict]:
    agg: dict = {}
    for r in payload.get("cube") or []:
        k = r.get(field) or "—"
        a = agg.setdefault(k, {"qty": 0.0, "value": 0.0})
        a["qty"] += r.get("qty") or 0
        a["value"] += r.get("value") or 0
    rows = [{field.title(): k, "Dispatched (KG)": round(v["qty"]),
             "Dispatch value": round(v["value"])} for k, v in agg.items()]
    return sorted(rows, key=lambda d: -d["Dispatched (KG)"])


def section_rows(payload: dict, section: str) -> list[dict]:
    """The table behind one card, as a list of row dicts (already display-ready)."""
    p = payload.get("projection") or {}
    k = payload.get("kpis") or {}

    if section == "summary":
        rows = [
            {"Metric": "Persona", "Value": payload.get("persona") or "—"},
            {"Metric": "Scope", "Value": "; ".join(payload.get("scope") or []) or "—"},
            {"Metric": "Dispatched (KG, 13 JCs)", "Value": _num(k.get("qty"))},
            {"Metric": "Dispatch value (13 JCs)", "Value": _num(k.get("value"))},
            {"Metric": "Customers served", "Value": _num(k.get("customers"))},
            {"Metric": "Items shipped", "Value": _num(k.get("items"))},
        ]
        if p:
            rows += [
                {"Metric": "Planning cycle", "Value": f"JC{p.get('jc')} {p.get('acc_year')}"},
                {"Metric": "Projection basis",
                 "Value": "your collectors" if p.get("basis") == "collector" else "per item"},
                {"Metric": "Accuracy on projected items (%)", "Value": p.get("overall_accuracy_proj")},
                {"Metric": "Accuracy incl. unprojected items (%)", "Value": p.get("overall_accuracy")},
                {"Metric": "Sales volume projected (%)", "Value": p.get("coverage_pct")},
                {"Metric": f"Items projected (JC{p.get('jc')})", "Value": p.get("items_projected")},
                {"Metric": "Items selling", "Value": p.get("items_selling")},
                {"Metric": "Items with no projection", "Value": p.get("missing_total")},
                {"Metric": "Unprojected volume (KG/JC)", "Value": p.get("missing_kg")},
            ]
        sync = (payload.get("last_sync") or {}).get("finished_at")
        rows.append({"Metric": "Data as of", "Value": str(sync or "—")})
        return rows

    if section == "collector":
        return _cube_by(payload, "collector")
    if section == "segment":
        return _cube_by(payload, "segment")

    if section == "jc_trend":
        return [{"Job cycle": t.get("label"),
                 "From": t.get("from") or "", "To": t.get("to") or "",
                 "Projected (KG)": _num(t.get("proj")),
                 "Actual sales (KG)": t.get("actual"),
                 "Items projected": _num(t.get("items_projected")),
                 "Items sold": t.get("items_sold"),
                 "Accuracy on projected (%)": t.get("accuracy_proj"),
                 "Accuracy all items (%)": t.get("accuracy"),
                 "Volume projected (%)": t.get("coverage_pct"),
                 "Status": "completed" if t.get("done") else "planning cycle"}
                for t in (p.get("jc_trend") or [])]

    if section == "status":
        return [{"Status": _FLAG_LABEL.get(s.get("flag"), s.get("flag")),
                 "Items": _num(s.get("items")),
                 "3-JC avg sales (KG)": _num(s.get("kg"))}
                for s in (p.get("summary") or [])]

    if section == "items":
        out = []
        for flag, rows in (p.get("items_by_flag") or {}).items():
            for r in rows:
                out.append({"Status": _FLAG_LABEL.get(flag, flag),
                            "Item code": r.get("code") or "", "Item": r.get("name"),
                            "3-JC avg sales (KG)": _num(r.get("avg3")),
                            "Projection (KG)": _num(r.get("proj"))})
        return sorted(out, key=lambda d: (d["Status"], -d["3-JC avg sales (KG)"]))

    if section == "groups":
        out = []
        for level, label in (("segment3", "Segment 3"), ("segment2", "Segment 2")):
            for g in (p.get("by_group") or {}).get(level) or []:
                out.append({"Level": label, "Item group": g.get("name"),
                            "Projected (KG)": _num(g.get("proj")),
                            "3-JC avg sales (KG)": _num(g.get("avg3")),
                            "Sales with a projection (KG)": _num(g.get("covered_kg")),
                            "Sales with none (KG)": _num(g.get("uncovered_kg")),
                            "Items": _num(g.get("items")),
                            "No projection": _num(g.get("missing")),
                            "Accuracy on projected (%)": g.get("accuracy_proj")})
        return out

    if section == "compare":
        return [{"Item code": c.get("code") or "", "Item": c.get("name"),
                 "3-JC avg sales (KG)": _num(c.get("avg3")),
                 "Projection (KG)": _num(c.get("proj")),
                 "Projected %": (round(c["proj"] / c["avg3"] * 100, 1)
                                 if c.get("avg3") else None),
                 "Status": _FLAG_LABEL.get(c.get("flag"), c.get("flag"))}
                for c in (p.get("compare") or [])]

    if section == "pipeline":
        jc = p.get("jc")
        return [{"Item code": r.get("code") or "", "Item": r.get("name"),
                 "3-JC avg sales (KG)": _num(r.get("avg3")),
                 f"Current JC{jc} (KG)": _num(r.get("proj")),
                 "Next JC (KG)": _num(r.get("next1")),
                 "JC after next (KG)": _num(r.get("next2")),
                 "Status": _FLAG_LABEL.get(r.get("flag"), r.get("flag"))}
                for r in (p.get("pipeline_rows") or [])]

    if section == "missing":
        total = p.get("missing_kg") or 0
        return [{"#": i + 1, "Item code": m.get("code") or "", "Item": m.get("name"),
                 "3-JC avg sales (KG)": _num(m.get("avg3")),
                 "Share of gap (%)": (round(m["avg3"] / total * 100, 1) if total else None)}
                for i, m in enumerate(p.get("missing_all") or [])]

    return []


def _write(ws, rows: list[dict]) -> int:
    """Header + rows, with sane column widths. Returns the row count."""
    if not rows:
        ws.append(["No data in scope"])
        return 0
    cols = list(rows[0].keys())
    ws.append(cols)
    for r in rows:
        ws.append([r.get(c) for c in cols])
    for i, c in enumerate(cols, start=1):
        width = max(len(str(c)), *(len(str(r.get(c) or "")) for r in rows[:200])) + 2
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(46, max(10, width))
    ws.freeze_panes = "A2"
    return len(rows)


def _bar(ws, title, n, cat_col, val_col, horizontal=False):
    from openpyxl.chart import BarChart, Reference
    ch = BarChart()
    ch.type = "bar" if horizontal else "col"
    ch.title = title
    ch.legend = None
    ch.add_data(Reference(ws, min_col=val_col, min_row=1, max_row=n + 1), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=cat_col, min_row=2, max_row=n + 1))
    ch.height, ch.width = 8, 16
    return ch


def _pie(ws, title, n, cat_col, val_col):
    from openpyxl.chart import DoughnutChart, Reference
    ch = DoughnutChart(holeSize=55)
    ch.title = title
    ch.add_data(Reference(ws, min_col=val_col, min_row=1, max_row=n + 1), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=cat_col, min_row=2, max_row=n + 1))
    ch.height, ch.width = 8, 16
    return ch


def _line(ws, title, n, cat_col, val_cols):
    from openpyxl.chart import LineChart, Reference
    ch = LineChart()
    ch.title = title
    for c in val_cols:
        ch.add_data(Reference(ws, min_col=c, min_row=1, max_row=n + 1), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=cat_col, min_row=2, max_row=n + 1))
    ch.height, ch.width = 8, 16
    ch.y_axis.scaling.min = 0
    return ch


def build(payload: dict, section: str | None = None) -> bytes:
    """One section, or the whole dashboard (charts sheet + every table)."""
    import openpyxl

    wb = openpyxl.Workbook()

    if section:
        ws = wb.active
        ws.title = SECTION_TITLES.get(section, "Data")[:31]
        _write(ws, section_rows(payload, section))
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    charts = wb.active
    charts.title = "Charts"
    charts["A1"] = "My Dashboard"
    charts["A1"].font = openpyxl.styles.Font(size=14, bold=True)
    charts["A2"] = f"{payload.get('persona') or ''} · {'; '.join(payload.get('scope') or [])}"
    charts["A3"] = f"Data as of {(payload.get('last_sync') or {}).get('finished_at') or '—'}"

    counts = {}
    for key in ("summary", "collector", "segment", "jc_trend", "status",
                "items", "groups", "compare", "pipeline", "missing"):
        ws = wb.create_sheet(SECTION_TITLES[key][:31])
        counts[key] = _write(ws, section_rows(payload, key))

    def sheet(key):
        return wb[SECTION_TITLES[key][:31]]

    # charts are laid out two per row on the Charts sheet
    placed, anchors = [], ["A5", "K5", "A24", "K24", "A43", "K43", "A62", "K62"]
    if counts["collector"]:
        placed.append(_bar(sheet("collector"), "Dispatched KG by collector",
                           min(counts["collector"], 15), 1, 2, horizontal=True))
    if counts["segment"]:
        placed.append(_pie(sheet("segment"), "Product mix by segment",
                           min(counts["segment"], 10), 1, 2))
    if counts["jc_trend"]:
        n = counts["jc_trend"]
        placed.append(_bar(sheet("jc_trend"), "Projected KG by JC", n, 1, 4))
        placed.append(_line(sheet("jc_trend"), "Projection accuracy by JC (%)", n, 1, [8, 10]))
        placed.append(_bar(sheet("jc_trend"), "Items projected by JC", n, 1, 6))
    if counts["status"]:
        placed.append(_pie(sheet("status"), "Projection status (items)", counts["status"], 1, 2))
    if counts["groups"]:
        placed.append(_bar(sheet("groups"), "Projected KG by item group",
                           min(counts["groups"], 14), 2, 3, horizontal=True))
    if counts["missing"]:
        placed.append(_bar(sheet("missing"), "Biggest unprojected items (KG/JC)",
                           min(counts["missing"], 15), 3, 4, horizontal=True))

    for ch, anchor in zip(placed, anchors):
        charts.add_chart(ch, anchor)
    charts.column_dimensions["A"].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


__all__ = ["build", "section_rows", "SECTION_TITLES"]
