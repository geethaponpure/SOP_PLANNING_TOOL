"""Excel export for the Supply-Competition page — one card, or the whole page.

Mirrors demand_export / commit_export: native Excel charts on the first sheet,
one sheet per table, item sheets UNCAPPED.

The holder sheets carry the same permission rule as the page: a persona sees
competing demand rolled up by collector and market circle, never the names of
customers outside its own scope.
"""
from __future__ import annotations

import io

SECTION_TITLES = {
    "summary": "Summary",
    "risk": "Supply risk",
    "items": "All items",
    "exposed": "Exposed items",
    "collectors": "Competing by collector",
    "mc": "Competing by market circle",
}


def _item_row(r: dict) -> dict:
    return {"Item code": r.get("item_code"),
            "Item": r.get("item"),
            "Segment": r.get("segment3"),
            "Risk": r.get("risk"),
            "My projection (KG)": r.get("my_projection"),
            "My firm orders (KG)": r.get("my_firm"),
            "My unprotected (KG)": r.get("my_unprotected"),
            "My customers": r.get("my_customers"),
            "Projected company-wide (KG)": r.get("all_projection"),
            "Unfirm company-wide (KG)": r.get("all_unprotected"),
            "Customers projecting": r.get("all_customers"),
            "Collectors projecting": r.get("all_collectors"),
            "On hand (KG)": r.get("on_hand"),
            "Safety level MSL (KG)": r.get("msl"),
            "Firm orders, company (KG)": r.get("firm_total"),
            "Firm orders, others (KG)": r.get("firm_others"),
            "Other customers": r.get("other_customers"),
            "Other order lines": r.get("other_lines"),
            "Incoming production (KG)": r.get("incoming"),
            "Production available from": r.get("incoming_date") or "",
            "ATP company (KG)": r.get("atp"),
            "ATP left for me (KG)": r.get("atp_for_me"),
            "Exposure (KG)": r.get("exposure"),
            "Overdue 90+ excluded (KG)": r.get("stale")}


def section_rows(payload: dict, rows: list[dict], section: str) -> list[dict]:
    k = payload.get("kpis") or {}
    if section == "summary":
        return [
            {"Metric": "Persona", "Value": payload.get("persona") or "—"},
            {"Metric": "Scope", "Value": "; ".join(payload.get("scope") or []) or "—"},
            {"Metric": "Cycle", "Value": f"{payload.get('jc_label')} "
                                         f"({payload.get('jc_from')} to {payload.get('jc_to')})"},
            {"Metric": "Items in view", "Value": k.get("items")},
            {"Metric": "Items with exposure", "Value": k.get("exposed_items")},
            {"Metric": "Exposed quantity (KG)", "Value": k.get("exposure")},
            {"Metric": "My projection (KG)", "Value": k.get("my_projection")},
            {"Metric": "My firm orders (KG)", "Value": k.get("my_firm")},
            {"Metric": "My unprotected projection (KG)", "Value": k.get("my_unprotected")},
            {"Metric": "Projected company-wide (KG)", "Value": k.get("all_projection")},
            {"Metric": "Unfirm company-wide — competes for the same supply (KG)",
             "Value": k.get("all_unprotected")},
            {"Metric": "On hand at selling orgs (KG)", "Value": k.get("on_hand")},
            {"Metric": "Firm orders, whole company (KG)", "Value": k.get("firm_total")},
            {"Metric": "  of which held by others (KG)", "Value": k.get("firm_others")},
            {"Metric": "Incoming production this plan (KG)", "Value": k.get("incoming")},
            {"Metric": f"Overdue {payload.get('stale_days')}+ days, NOT counted as demand (KG)",
             "Value": k.get("stale")},
            {"Metric": "Safety level source", "Value": payload.get("msl_ref") or "no MSL snapshot"},
            {"Metric": "Production plan used", "Value": payload.get("plan_id") or "no saved plan"},
            {"Metric": "Selling orgs counted as available", "Value": payload.get("sell_orgs")},
            {"Metric": "Competing customers shown by name",
             "Value": "yes" if payload.get("show_names") else "no — rolled up by collector"},
            {"Metric": "Data as of",
             "Value": str((payload.get("last_sync") or {}).get("finished_at") or "—")},
        ]
    if section == "risk":
        return [{"Risk": b["label"], "Items": b["items"], "Quantity (KG)": b["qty"]}
                for b in k.get("buckets") or []]
    if section == "collectors":
        return [{"Collector": r.get("collector"), "Committed balance (KG)": r.get("balance"),
                 "Order lines": r.get("lines"), "Customers": r.get("customers"),
                 "Items": r.get("items")} for r in payload.get("by_collector") or []]
    if section == "mc":
        return [{"Market circle": r.get("mc_code"), "Committed balance (KG)": r.get("balance"),
                 "Order lines": r.get("lines"), "Customers": r.get("customers"),
                 "Items": r.get("items")} for r in payload.get("by_mc") or []]
    if section == "items":
        return [_item_row(r) for r in rows]
    if section == "exposed":
        return [_item_row(r) for r in rows if (r.get("exposure") or 0) > 0]
    return []


def _write(ws, rows: list[dict]) -> int:
    if not rows:
        ws.append(["No data in scope"])
        return 0
    cols = list(rows[0].keys())
    ws.append(cols)
    for r in rows:
        ws.append([r.get(c) for c in cols])
    for i, c in enumerate(cols, start=1):
        width = max(len(str(c)), *(len(str(r.get(c) or "")) for r in rows[:200])) + 2
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(44, max(9, width))
    ws.freeze_panes = "A2"
    return len(rows)


def build(payload: dict, rows: list[dict], section: str | None = None) -> bytes:
    import openpyxl
    from openpyxl.chart import BarChart, DoughnutChart, Reference

    wb = openpyxl.Workbook()
    if section:
        ws = wb.active
        ws.title = SECTION_TITLES.get(section, "Data")[:31]
        _write(ws, section_rows(payload, rows, section))
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    charts = wb.active
    charts.title = "Charts"
    charts["A1"] = "Supply Competition"
    charts["A1"].font = openpyxl.styles.Font(size=14, bold=True)
    charts["A2"] = f"{payload.get('persona') or ''} · {'; '.join(payload.get('scope') or [])}"
    charts["A3"] = (f"{payload.get('jc_label')} · MSL {payload.get('msl_ref') or '—'}"
                    f" · plan {payload.get('plan_id') or '—'}"
                    f" · data {((payload.get('last_sync') or {}).get('finished_at') or '—')}")

    counts = {}
    for key in ("summary", "risk", "collectors", "mc", "exposed", "items"):
        ws = wb.create_sheet(SECTION_TITLES[key][:31])
        counts[key] = _write(ws, section_rows(payload, rows, key))

    def sheet(key):
        return wb[SECTION_TITLES[key][:31]]

    placed = []
    if counts["risk"]:
        ch = DoughnutChart(holeSize=55)
        ch.title = "Items by supply risk"
        ch.add_data(Reference(sheet("risk"), min_col=2, min_row=1,
                              max_row=counts["risk"] + 1), titles_from_data=True)
        ch.set_categories(Reference(sheet("risk"), min_col=1, min_row=2,
                                    max_row=counts["risk"] + 1))
        ch.height, ch.width = 8, 16
        placed.append(ch)
    if counts["exposed"]:
        n = min(counts["exposed"], 15)
        ch = BarChart()
        ch.type = "bar"
        ch.title = "Most exposed items"
        ch.legend = None
        # "Exposure (KG)" — recomputed from the header so added columns cannot
        # silently repoint this chart at the wrong series
        hdr = [c.value for c in next(sheet("exposed").iter_rows(min_row=1, max_row=1))]
        col = hdr.index("Exposure (KG)") + 1
        ch.add_data(Reference(sheet("exposed"), min_col=col, min_row=1, max_row=n + 1),
                    titles_from_data=True)
        ch.set_categories(Reference(sheet("exposed"), min_col=2, min_row=2, max_row=n + 1))
        ch.height, ch.width = 9, 16
        placed.append(ch)
    if counts["collectors"]:
        n = min(counts["collectors"], 15)
        ch = BarChart()
        ch.type = "bar"
        ch.title = "Competing demand by collector"
        ch.legend = None
        ch.add_data(Reference(sheet("collectors"), min_col=2, min_row=1, max_row=n + 1),
                    titles_from_data=True)
        ch.set_categories(Reference(sheet("collectors"), min_col=1, min_row=2, max_row=n + 1))
        ch.height, ch.width = 9, 16
        placed.append(ch)
    for ch, anchor in zip(placed, ["A5", "K5", "A24", "K24"]):
        charts.add_chart(ch, anchor)
    charts.column_dimensions["A"].width = 34

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


__all__ = ["build", "section_rows", "SECTION_TITLES"]
