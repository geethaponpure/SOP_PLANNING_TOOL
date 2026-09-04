"""Excel export for the Commitment-Risk page — one card, or the whole page.

Mirrors dashboard_export: native Excel charts on the first sheet (linked to the
data sheets), one sheet per table. Line sheets are UNCAPPED — the page shows the
first 500 by urgency, the workbook carries everything in scope.
"""
from __future__ import annotations

import io

from .commit import BUCKETS

SECTION_TITLES = {
    "summary": "Summary",
    "buckets": "Risk buckets",
    "timeline": "Commitment timeline",
    "lines": "All lines",
    "rush": "Rush - due 48h",
    "emergency": "Emergency - overdue",
    "pushed": "Pushed commitments",
    "reasons": "Push reasons",
}

_BL = dict(BUCKETS)


def _line_row(r: dict) -> dict:
    days = r.get("days")
    return {"Order": r.get("order_ref") or r.get("order_no"),
            "Order date": r.get("soc_date"),
            "Customer": r.get("customer_name"),
            "Collector": r.get("collector"), "MC": r.get("mc_code"),
            "Item code": r.get("item_code"), "Item": r.get("item_name"),
            "Org": r.get("inv_org"),
            "Balance (KG)": r.get("balance"),
            "Committed (original)": r.get("sched_date"),
            "Committed (current)": r.get("resched_date"),
            "Days late" if days is not None and days < 0 else "Days to due":
                (abs(days) if days is not None and days < 0 else days),
            "Risk": _BL.get(r.get("bucket"), r.get("bucket")),
            "Pushed": "yes" if r.get("pushed") else "",
            "Reschedule reason": r.get("resched_reason") or "",
            "Confirm status": r.get("confirm_status") or "",
            "Plan supply date": r.get("supply_date") or "",
            "Supply risk": "YES" if r.get("supply_risk") else ""}


def _norm_line(r: dict) -> dict:
    """_line_row with a stable column set (Days column unified)."""
    d = _line_row(r)
    days = r.get("days")
    d.pop("Days late", None)
    d.pop("Days to due", None)
    out = {}
    for k, v in d.items():
        out[k] = v
        if k == "Committed (current)":
            out["Days vs commitment"] = days
    return out


def section_rows(payload: dict, rows: list[dict], section: str) -> list[dict]:
    k = payload.get("kpis") or {}
    if section == "summary":
        return [
            {"Metric": "Persona", "Value": payload.get("persona") or "—"},
            {"Metric": "Scope", "Value": "; ".join(payload.get("scope") or []) or "—"},
            {"Metric": "As of", "Value": payload.get("today")},
            {"Metric": "Open committed lines", "Value": k.get("lines")},
            {"Metric": "Balance to dispatch (KG)", "Value": k.get("kg")},
            {"Metric": "Overdue lines", "Value": k.get("overdue_lines")},
            {"Metric": "Overdue balance (KG)", "Value": k.get("overdue_kg")},
            {"Metric": "Rush lines (due ≤48h)", "Value": k.get("rush_lines")},
            {"Metric": "Rush balance (KG)", "Value": k.get("rush_kg")},
            {"Metric": "Lines pushed past original commitment", "Value": k.get("pushed_lines")},
            {"Metric": "Urgent lines with a supply risk", "Value": k.get("supply_risk_lines")},
            {"Metric": "Supply check against JC plan", "Value": payload.get("supply_plan") or "no saved plan"},
            {"Metric": "Data as of", "Value": str((payload.get("last_sync") or {}).get("finished_at") or "—")},
        ]
    if section == "buckets":
        return [{"Risk": b["label"], "Lines": b["lines"], "Balance (KG)": b["kg"]}
                for b in payload.get("buckets") or []]
    if section == "timeline":
        return [{"Day": t["label"], "Lines due": t["lines"], "Committed (KG)": t["kg"]}
                for t in payload.get("timeline") or []]
    if section == "reasons":
        return [{"Reschedule reason": x["reason"], "Lines": x["lines"], "Balance (KG)": x["kg"]}
                for x in payload.get("reasons") or []]
    if section == "lines":
        return [_norm_line(r) for r in rows]
    if section == "rush":
        return [_norm_line(r) for r in rows if r.get("bucket") in ("today", "d2")]
    if section == "emergency":
        return [_norm_line(r) for r in rows if r.get("bucket") in ("overdue7", "overdue")]
    if section == "pushed":
        return [_norm_line(r) for r in rows if r.get("pushed")]
    return []


def _write(ws, rows: list[dict]) -> int:
    if not rows:
        ws.append(["No data in scope"])
        return 0
    cols = list(rows[0].keys())
    ws.append(cols)
    for r in rows:
        ws.append([r.get(c) for c in cols])
    for i, c in enumerate(cols, start=1):
        width = max(len(str(c)), *(len(str(r.get(c) or "")) for r in rows[:200])) + 2
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(44, max(9, width))
    ws.freeze_panes = "A2"
    return len(rows)


def build(payload: dict, rows: list[dict], section: str | None = None) -> bytes:
    import openpyxl
    from openpyxl.chart import BarChart, DoughnutChart, Reference

    wb = openpyxl.Workbook()
    if section:
        ws = wb.active
        ws.title = SECTION_TITLES.get(section, "Data")[:31]
        _write(ws, section_rows(payload, rows, section))
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    charts = wb.active
    charts.title = "Charts"
    charts["A1"] = "Commitment Risk"
    charts["A1"].font = openpyxl.styles.Font(size=14, bold=True)
    charts["A2"] = f"{payload.get('persona') or ''} · {'; '.join(payload.get('scope') or [])}"
    charts["A3"] = f"As of {payload.get('today')} · data {((payload.get('last_sync') or {}).get('finished_at') or '—')}"

    counts = {}
    for key in ("summary", "buckets", "timeline", "reasons",
                "rush", "emergency", "pushed", "lines"):
        ws = wb.create_sheet(SECTION_TITLES[key][:31])
        counts[key] = _write(ws, section_rows(payload, rows, key))

    def sheet(key):
        return wb[SECTION_TITLES[key][:31]]

    placed = []
    if counts["buckets"]:
        ch = DoughnutChart(holeSize=55)
        ch.title = "Open lines by risk"
        ch.add_data(Reference(sheet("buckets"), min_col=2, min_row=1,
                              max_row=counts["buckets"] + 1), titles_from_data=True)
        ch.set_categories(Reference(sheet("buckets"), min_col=1, min_row=2,
                                    max_row=counts["buckets"] + 1))
        ch.height, ch.width = 8, 16
        placed.append(ch)
    if counts["timeline"]:
        ch = BarChart()
        ch.type = "col"
        ch.title = "Committed KG by day"
        ch.legend = None
        ch.add_data(Reference(sheet("timeline"), min_col=3, min_row=1,
                              max_row=counts["timeline"] + 1), titles_from_data=True)
        ch.set_categories(Reference(sheet("timeline"), min_col=1, min_row=2,
                                    max_row=counts["timeline"] + 1))
        ch.height, ch.width = 8, 16
        placed.append(ch)
    if counts["reasons"]:
        ch = DoughnutChart(holeSize=55)
        ch.title = "Why commitments were pushed"
        ch.add_data(Reference(sheet("reasons"), min_col=2, min_row=1,
                              max_row=min(counts["reasons"], 10) + 1), titles_from_data=True)
        ch.set_categories(Reference(sheet("reasons"), min_col=1, min_row=2,
                                    max_row=min(counts["reasons"], 10) + 1))
        ch.height, ch.width = 8, 16
        placed.append(ch)
    for ch, anchor in zip(placed, ["A5", "K5", "A24", "K24"]):
        charts.add_chart(ch, anchor)
    charts.column_dimensions["A"].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


__all__ = ["build", "section_rows", "SECTION_TITLES"]
