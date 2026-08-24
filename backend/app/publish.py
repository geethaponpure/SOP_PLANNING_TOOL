"""Excel workbook builders for the live-CRM RM/purchasing pages.

Each function takes the page result dict (+ cycle label) and returns .xlsx bytes.
"""
from __future__ import annotations

import io

from .integration import planning_filter as _pf

_NAVY, _TEAL, _BAND, _RED, _GREEN, _AMBER = "1F3A5F", "2A9D8F", "E6FFFA", "FFE5E5", "E6F6EC", "FFF4DA"
_QTY = "#,##0.0"
_ACT_LABEL = {"manufacturing": "Manufacturing", "repack_relabel": "Repack/Relabel",
              "trading": "Trading/Distribution", "internal": "Internal",
              "unclassified": "Unclassified", "none": "No BOM"}


def _styles():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    thin = Side(style="thin", color="D7DEE8")
    return {
        "hdr_fill": PatternFill("solid", fgColor=_NAVY),
        "band_fill": PatternFill("solid", fgColor=_TEAL),
        "zebra": PatternFill("solid", fgColor="F6F9FC"),
        "white": Font(bold=True, color="FFFFFF", size=10),
        "title": Font(bold=True, color="FFFFFF", size=16),
        "ctr": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
    }


def _hdr_row(ws, row, ncols, st, fill=None):
    fill = fill or st["hdr_fill"]
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = st["white"]
        cell.alignment = st["ctr"]
        cell.border = st["border"]
    ws.row_dimensions[row].height = 26


def _widths(ws, widths):
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _grid(wb, st, title, headers, rows, widths, qty_cols, fill_map=None, flag_col=None, qty_fmt=None):
    from openpyxl.utils import get_column_letter
    fmt = qty_fmt or _QTY
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    for c, name in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=name)
    _hdr_row(ws, 1, len(headers), st)
    rr = 2
    for r in rows:
        for ci, v in enumerate(r, 1):
            cell = ws.cell(row=rr, column=ci, value=v)
            cell.border = st["border"]
            if ci in qty_cols:
                cell.number_format = fmt
            if flag_col and ci == flag_col and fill_map and v in fill_map:
                cell.fill = fill_map[v]
            elif rr % 2 == 0:
                cell.fill = st["zebra"]
        rr += 1
    _widths(ws, widths)
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(rr - 1, 1)}"
    return ws


def _cover(wb, st, title, subtitle, metrics):
    from openpyxl.styles import Font, PatternFill, Alignment
    cov = wb.create_sheet("Summary")
    cov.sheet_view.showGridLines = False
    cov.merge_cells("A1:F2")
    t = cov["A1"]; t.value = title
    t.font = st["title"]; t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for col in "ABCDEF":
        cov[f"{col}1"].fill = st["hdr_fill"]; cov[f"{col}2"].fill = st["hdr_fill"]
    cov["A3"] = subtitle
    cov["A3"].font = Font(italic=True, color="718096")
    for i, (label, val, color) in enumerate(metrics):
        col = 1 + (i % 3) * 2
        row = 5 + (i // 3) * 3
        cov.cell(row=row, column=col, value=label).font = Font(size=10, color="718096")
        v = cov.cell(row=row + 1, column=col, value=val); v.font = Font(bold=True, size=18, color=_NAVY)
        for rr in (row, row + 1):
            cov.cell(row=rr, column=col).fill = PatternFill("solid", fgColor=color)
            cov.cell(row=rr, column=col + 1).fill = PatternFill("solid", fgColor=color)
    _widths(cov, [22, 12, 22, 12, 22, 12])
    return cov


def _book():
    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    return wb


def _color_lead(ws, col, amber=30, red=60):
    """Colour a numeric lead-time column: amber >= amber days, red >= red days."""
    from openpyxl.styles import PatternFill
    fa = PatternFill("solid", fgColor=_AMBER)
    fr = PatternFill("solid", fgColor=_RED)
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=col)
        v = cell.value
        if isinstance(v, (int, float)):
            if v >= red:
                cell.fill = fr
            elif v >= amber:
                cell.fill = fa


def _drop_first_col(headers, widths, rows, qty_cols):
    """Drop the leading column (e.g. Product) and shift the qty-format indices."""
    return headers[1:], widths[1:], [r[1:] for r in rows], {c - 1 for c in qty_cols}


def _save(wb):
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Supply & RM planning workbook ────────────────────────────────────────────
def build_rm_planning_workbook(rp, cycle="", stock_lots=None, intransit_lots=None) -> bytes:
    from openpyxl.styles import PatternFill
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.utils import get_column_letter
    st = _styles()
    s = rp.get("summary", {})
    products = rp.get("products", [])
    red = PatternFill("solid", fgColor=_RED)
    green = PatternFill("solid", fgColor=_GREEN)
    wb = _book()
    win = rp.get("soc_window") or {}
    _minq = s.get("min_plan_qty")
    _cover(wb, st, "Supply & RM — Planning Filtration Report",
           f"Cycle {cycle} · stock: {s.get('stock_source','—')}"
           + (f" · pending-SOC {win.get('from')}→{win.get('to')}" if win else "")
           + (f" · planned only if a JC projection OR Pending SOC > {_minq:g} KG" if _minq else ""),
           [("Projected products", s.get("projected_products"), _BAND),
            ("Manufacturing", s.get("manufacturing"), _GREEN),
            ("Repack / Relabel", s.get("repack_relabel"), _BAND),
            ("Trading / Distribution", s.get("trading"), _AMBER)])

    # Consolidated RM — by ITEM DESCRIPTION (one row per material, all its codes
    # rolled up). Manufacturing / Repack / Packing are planned on separate sheets.
    _cpjc = rp.get("planning_jc") or 0
    _c_cur = f"JC{_cpjc}" if _cpjc else "Curr"
    _c_n1 = f"JC{_cpjc + 1}" if _cpjc else "Next1"
    _c_n2 = f"JC{_cpjc + 2}" if _cpjc else "Next2"

    def _plan_jcs_label(x):
        """Actionable plan: 'Available' when RM is on hand (no shortfall in the lead
        horizon); else the JC(s) to buy in, chosen by the lead+preprocessing bucket."""
        if not x.get("to_buy", x.get("net_total", 0) > 0):
            return "Available"
        m = {"current": _c_cur, "next1": _c_n1, "next2": _c_n2}
        buy = x.get("buy_jcs") or x.get("planned_jcs", [])
        return "Buy " + ", ".join(m.get(k, k) for k in buy)

    def _subs_label(x):
        return "; ".join(f"{su.get('desc') or su['code']} ({su['stock']})"
                         for su in x.get("substitutes", []))

    def _cons_sheet(title, lst):
        _grid(wb, st, title,
              ["Item Description", "#Codes", "Item Codes", "Business", "Activity", "#FG", _c_cur, _c_n1, _c_n2,
               "Total", "Stock", "Sub stk", "In-transit", f"Net {_c_cur}", f"Net {_c_n1}", f"Net {_c_n2}", "Net Total",
               "Avg lead (d)", "Lead+7 (d)", "Plan / Buy", "Trade", "Currency", "Suppliers", "Used in FGs",
               "Substitutes (name · stock)"],
              [[x["rm_desc"], x.get("code_count"), ", ".join(x.get("rm_codes", [])[:6]), x.get("business", ""),
                x.get("activity", ""), x["fg_count"],
                x["gross"]["current"], x["gross"]["next1"], x["gross"]["next2"], x["gross_total"],
                x["main_stock"], x["substitute_stock"], x["in_transit"],
                x["net_to_buy"]["current"], x["net_to_buy"]["next1"], x["net_to_buy"]["next2"], x["net_total"],
                x.get("avg_lead_time_days"), x.get("lead_total_days"), _plan_jcs_label(x),
                x.get("trade"), ", ".join(x.get("currencies", [])),
                "; ".join(x.get("suppliers", [])[:4]), ", ".join(x["fgs"]), _subs_label(x)]
               for x in lst],
              [30, 7, 30, 15, 16, 6, 11, 11, 11, 12, 11, 10, 11, 11, 11, 11, 12, 11, 11, 14, 10, 12, 40, 46, 44],
              {6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17})

    _cons_sheet("RM — Manufacturing", rp.get("consolidated_rm_manufacturing", []))
    _cons_sheet("RM — Repack-Relabel", rp.get("consolidated_rm_repack", []))
    _cons_sheet("RM — All (combined)", rp.get("consolidated_rm", []))
    # NOTE: Packing Material + Packing BOMs are exported in a SEPARATE workbook
    # (build_packing_workbook / "Supply_Packing_Plan.xlsx"), not in this RM plan.

    # Real RM Requirement — every intermediate exploded to its purchased (leaf) RMs,
    # names decoded. This is the true buy-list once encoded intermediates are resolved.
    def _real_rm_sheet(title, lst):
        _grid(wb, st, title,
              ["Item Description (decoded)", "Business", "Activity", "Status", "#Codes", "Item Codes", "Via intermediate(s)", "#FG",
               _c_cur, _c_n1, _c_n2, "Total", "Stock", "Stock by Org", "Encoded Name(s)", "of which Encoded", "In-transit", f"Net {_c_cur}", f"Net {_c_n1}",
               f"Net {_c_n2}", "Net Total", "Avg lead (d)", "Lead+7 (d)", "Plan / Buy", "Used in FGs"],
              [[x["rm_desc"], x.get("business", ""), x.get("activity", ""),
                ("⚠ Unresolved intermediate" if x.get("unresolved") else "Leaf RM"), x.get("code_count"),
                ", ".join(x.get("rm_codes", [])[:6]),
                ("; ".join(x.get("from_intermediates", [])) or "—"), x["fg_count"],
                x["gross"]["current"], x["gross"]["next1"], x["gross"]["next2"], x["gross_total"],
                x["main_stock"], x.get("stock_orgs", ""), x.get("encoded_names", ""), (x.get("encoded_stock", 0) or ""), x["in_transit"],
                x["net_to_buy"]["current"], x["net_to_buy"]["next1"], x["net_to_buy"]["next2"], x["net_total"],
                x.get("avg_lead_time_days"), x.get("lead_total_days"), _plan_jcs_label(x),
                ", ".join(x["fgs"])]
               for x in lst],
              [30, 12, 16, 22, 7, 30, 34, 6, 11, 11, 11, 12, 11, 40, 24, 13, 11, 11, 11, 11, 12, 11, 11, 14, 46],
              {9, 10, 11, 12, 13, 16, 17, 18, 19, 20, 21})

    _real_rm_sheet("Real RM Requirement", rp.get("real_rm_requirement", []))
    _real_rm_sheet("Real RM — Manufacturing", rp.get("real_rm_requirement_manufacturing", []))

    _pjc = rp.get("planning_jc") or 0
    _wk_lbl = f"JC{_pjc} Qty (WK1+2)" if _pjc else "JC Qty (WK1+2)"
    _cur_lbl = f"JC{_pjc} Qty" if _pjc else "Current"
    _n1_lbl = f"JC{_pjc + 1} Qty" if _pjc else "Next 1"
    _n2_lbl = f"JC{_pjc + 2} Qty" if _pjc else "Next 2"
    _grid(wb, st, "Products",
          ["#", "Product", "Activity", "Policy", _wk_lbl, "Overall SOC", "MFG SOC Pending", _cur_lbl, _n1_lbl, _n2_lbl,
           "Total", "MSL buffer", "Mfg Required (Current)", "Mfg Required (3 JC)", "Producible (PTS-first)",
           f"Producible {_c_cur}", f"Producible {_c_cur} Status", "Producible covers",
           "Avg 3JC sales", "vs proj", "BOM Assembly", "Warehouse On-Hand Qty (Kg)", "Branch", "RM to buy"],
          [[i, p["name"], _ACT_LABEL.get(p.get("bom_class"), p.get("bom_class", "")), p.get("pts_pto", ""),
            p["projection"].get("current_target", p["projection"]["current"]),
            p["projection"].get("overall_soc", 0), p["projection"].get("mfg_soc", 0),
            p["projection"]["current"], p["projection"]["next1"],
            p["projection"]["next2"], p["projection"]["total"], p["projection"].get("msl", 0),
            p["projection"].get("mfg_required", 0),
            p["projection"].get("mfg_required_3jc", 0), (p.get("producible_qty", 0) if p["has_bom"] else ""),
            (p["projection"].get("producible_current", 0) if p["has_bom"] else ""),
            (p["projection"].get("producible_status_current", "") if p["has_bom"] else ""),
            (p["projection"].get("producible_cover", "") if p["has_bom"] else ""),
            p["avg_3jc_sales"], p["proj_flag"],
            (f"{p['boms'][0]['assembly_item']}{' (Overridden)' if p.get('overridden') else ''}" if p["has_bom"] else ""),
            (p["boms"][0]["fg_stock"]["warehouse"] if p["has_bom"] else ""),
            (p["boms"][0]["fg_stock"]["branch"] if p["has_bom"] else ""),
            round(sum(c["net_total"] for c in p["boms"][0]["components"]) if p["has_bom"] else 0, 1)]
           for i, p in enumerate(products, 1)],
          [5, 32, 15, 8, 15, 11, 14, 11, 11, 11, 12, 12, 15, 15, 15, 15, 16, 18, 13, 10, 18, 24, 12, 12],
          {5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 19, 22, 23, 24})

    # selected BOM (system-preferred OR user-overridden) -> raw materials required.
    # Each FG's DIRECT components are listed as "Direct BOM"; RMs hidden inside an
    # intermediate component are exploded and listed as "Intermediate BOM" (with the
    # via-intermediate) so materials like PIGMACOLOR BINDER AFF become visible here too.
    from openpyxl.styles import PatternFill
    _INT_FILL = PatternFill("solid", fgColor="FFF4DA")
    _dm_codes = {str(c).upper() for c in (rp.get("rules") or {}).get("dm_water_codes", [])}

    def _is_dm(code, desc):
        d = str(desc or "").upper()
        return str(code or "").upper() in _dm_codes or "DM WATER" in d or "DEMINERALIZED" in d or "DEMINERALISED" in d

    rm_rows = []
    for p in products:
        if not p["has_bom"]:
            continue
        b = p["boms"][0]
        bom_src = "Overridden" if p.get("overridden") else "Preferred"
        for c in b["components"]:
            if _is_dm(c["rm_code"], c["rm_desc"]):     # DM water excluded from planning
                continue
            rm_rows.append([
                p["name"], "Direct BOM", "-", bom_src, b["assembly_item"], b.get("org_code", ""),
                b.get("designator", ""), c["seq"], c["rm_code"], c["rm_desc"], c["qty_per_unit"],
                c["gross"]["current"], c["gross"]["next1"], c["gross"]["next2"], c["gross_total"],
                c["main_stock"], c["substitute_stock"], c["in_transit"], c["net_total"],
                ", ".join(f"{su.get('desc') or su['code']} [{su['code']}] ({su['stock']})"
                          for su in c["substitutes"]),
            ])
        for im in p.get("intermediate_bom_rms", []):
            rm_rows.append([
                p["name"], "Intermediate BOM", im["via"], bom_src, b["assembly_item"],
                b.get("org_code", ""), b.get("designator", ""), "", im["rm_code"], im["rm_desc"], "",
                im["gross"]["current"], im["gross"]["next1"], im["gross"]["next2"], im["gross_total"],
                "", "", "", "", "",
            ])
    _grid(wb, st, "Selected BOM RMs",
          ["Product", "Source", "Via intermediate", "BOM", "BOM Assembly", "Org", "Designator",
           "Seq", "RM Code", "RM Description", "Qty / unit",
           f"Gross {_c_cur}", f"Gross {_c_n1}", f"Gross {_c_n2}", "Gross Total",
           "Stock", "Sub stk", "In-transit", "Net to buy", "Substitutes (name · stock)"],
          rm_rows,
          [28, 15, 26, 11, 16, 6, 13, 5, 16, 26, 10, 11, 11, 11, 12, 11, 9, 10, 12, 44],
          {11, 12, 13, 14, 15, 16, 17, 18, 19}, fill_map={"Intermediate BOM": _INT_FILL}, flag_col=2)

    # Packing BOMs now live in the separate packing workbook (build_packing_workbook).
    _add_stock_audit(wb, st, products, stock_lots)
    _add_intransit_audit(wb, st, rp, intransit_lots)
    wb.move_sheet("Summary", -(len(wb.sheetnames) - 1))
    return _save(wb)


def build_packing_workbook(rp, cycle="") -> bytes:
    """Standalone PACKING plan: 'Packing Material' (consolidated packing components by
    description) + 'Packing BOMs' (per-FG packing components). Split out of the main
    Supply_RM_Plan so the RM buy-list and the packing buy-list are downloaded separately."""
    st = _styles()
    s = rp.get("summary", {})
    products = rp.get("products", [])
    packing = rp.get("consolidated_rm_packing", [])
    wb = _book()
    _cpjc = rp.get("planning_jc") or 0
    _c_cur = f"JC{_cpjc}" if _cpjc else "Curr"
    _c_n1 = f"JC{_cpjc + 1}" if _cpjc else "Next1"
    _c_n2 = f"JC{_cpjc + 2}" if _cpjc else "Next2"

    def _plan_jcs_label(x):
        if not x.get("to_buy", x.get("net_total", 0) > 0):
            return "Available"
        m = {"current": _c_cur, "next1": _c_n1, "next2": _c_n2}
        return "Buy " + ", ".join(m.get(k, k) for k in (x.get("buy_jcs") or x.get("planned_jcs", [])))

    def _subs_label(x):
        return "; ".join(f"{su.get('desc') or su['code']} ({su['stock']})"
                         for su in x.get("substitutes", []))

    _n_pk_boms = sum(len(p.get("packing_boms", [])) for p in products)
    _cover(wb, st, "Packing Material — Planning Report",
           f"Cycle {cycle} · stock: {s.get('stock_source', '—')} · packing components split from the RM plan",
           [("Packing materials", len(packing), _BAND),
            ("Products with packing BOMs", sum(1 for p in products if p.get("packing_boms")), _BAND),
            ("Packing BOMs", _n_pk_boms, _BAND),
            ("Packing to buy (Net KG)", round(sum(x.get("net_total", 0) for x in packing), 1), _AMBER)])

    _grid(wb, st, "Packing Material",
          ["Item Description", "#Codes", "Item Codes", "Business", "Activity", "#FG", _c_cur, _c_n1, _c_n2,
           "Total", "Stock", "Sub stk", "In-transit", f"Net {_c_cur}", f"Net {_c_n1}", f"Net {_c_n2}", "Net Total",
           "Avg lead (d)", "Lead+7 (d)", "Plan / Buy", "Trade", "Currency", "Suppliers", "Used in FGs",
           "Substitutes (name · stock)"],
          [[x["rm_desc"], x.get("code_count"), ", ".join(x.get("rm_codes", [])[:6]), x.get("business", ""),
            x.get("activity", ""), x["fg_count"],
            x["gross"]["current"], x["gross"]["next1"], x["gross"]["next2"], x["gross_total"],
            x["main_stock"], x["substitute_stock"], x["in_transit"],
            x["net_to_buy"]["current"], x["net_to_buy"]["next1"], x["net_to_buy"]["next2"], x["net_total"],
            x.get("avg_lead_time_days"), x.get("lead_total_days"), _plan_jcs_label(x),
            x.get("trade"), ", ".join(x.get("currencies", [])),
            "; ".join(x.get("suppliers", [])[:4]), ", ".join(x["fgs"]), _subs_label(x)]
           for x in packing],
          [30, 7, 30, 15, 16, 6, 11, 11, 11, 12, 11, 10, 11, 11, 11, 11, 12, 11, 11, 14, 10, 12, 40, 46, 44],
          {6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17})

    rows = []
    for p in products:
        for pb in p.get("packing_boms", []):
            for c in pb["components"]:
                rows.append([p["name"], pb["assembly_item"], pb.get("org_code", ""), pb.get("designator", ""),
                             pb.get("created") or "", c["seq"], c["rm_code"], c["rm_desc"],
                             c["qty_per_unit"], c["gross_total"], c["main_stock"], c["in_transit"], c["net_total"]])
    _grid(wb, st, "Packing BOMs",
          ["Product", "Packing Assembly", "Org", "Designator", "Created", "Seq",
           "Component Code", "Component / packing material", "Qty/unit", "Gross total",
           "Stock", "In-transit", "Net to buy"],
          rows, [30, 18, 6, 14, 12, 5, 16, 30, 9, 12, 11, 11, 12], {9, 10, 11, 12, 13})

    wb.move_sheet("Summary", -(len(wb.sheetnames) - 1))
    return _save(wb)


def _safe_sheet_name(base: str, suffix: str, used: set) -> str:
    """Excel-safe, unique sheet name (<=31 chars) '<base>-<suffix>'."""
    bad = set(r'\/?*[]:')
    clean = "".join(c for c in (base or "Seg") if c not in bad).strip() or "Seg"
    room = 31 - len(suffix) - 1
    name = f"{clean[:room].strip()}-{suffix}"
    n, k = name, 2
    while n.lower() in used:
        tag = f"~{k}"
        n = f"{clean[:room - len(tag)].strip()}{tag}-{suffix}"
        k += 1
    used.add(n.lower())
    return n


def _seg_hdr(pjc=0):
    cur = f"JC{pjc} Qty(Kg)" if pjc else "Current Qty(Kg)"
    jcl = f"JC{pjc}" if pjc else "Current"
    return ["#", "Product", "Segment 3", "Activity", "Policy",
            (f"JC{pjc} Qty (WK1+2) Qty(Kg)" if pjc else "JC Qty (WK1+2) Qty(Kg)"),
            "Overall SOC Qty(Kg)", "MFG SOC Pending Qty(Kg)",
            cur, (f"JC{pjc + 1} Qty(Kg)" if pjc else "Next 1 JC Qty(Kg)"),
            (f"JC{pjc + 2} Qty(Kg)" if pjc else "Next 2 JC Qty(Kg)"), "Total Qty(Kg)",
            "MSL buffer Qty(Kg)",
            "Mfg Required (Current) Qty(Kg)", "Mfg Required (3 JC) Qty(Kg)",
            "Producible (PTS-first) Qty(Kg)", f"Producible {jcl} Qty(Kg)", f"Producible {jcl} Status",
            "Producible covers", "Avg 3JC sales Qty(Kg)", "vs proj",
            "Warehouse Stock Qty (Kg)", "Branch Stock Qty (Kg)"]


_SEG_WID = [5, 34, 22, 15, 8, 16, 13, 15, 12, 13, 13, 12, 13, 15, 15, 15, 15, 16, 18, 14, 9, 20, 18]
_SEG_NUM = {6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 20, 22, 23}


def _seg_current(p):
    return (p.get("projection") or {}).get("current", 0) or 0


def _segment_rows(prods):
    out = []
    for i, p in enumerate(sorted(prods, key=lambda x: (-_seg_current(x), x["name"])), 1):
        pj = p.get("projection") or {}
        b = p["boms"][0] if (p.get("has_bom") and p.get("boms")) else None
        out.append([
            i, p["name"], p.get("segment3", ""),
            _ACT_LABEL.get(p.get("bom_class"), p.get("bom_class", "")), p.get("pts_pto", ""),
            pj.get("current_target", pj.get("current", 0)),
            pj.get("overall_soc", 0), pj.get("mfg_soc", 0),
            pj.get("current", 0), pj.get("next1", 0), pj.get("next2", 0), pj.get("total", 0),
            pj.get("msl", 0),
            pj.get("mfg_required", 0), pj.get("mfg_required_3jc", 0),
            (p.get("producible_qty", 0) if p.get("has_bom") else ""),
            (pj.get("producible_current", 0) if p.get("has_bom") else ""),
            (pj.get("producible_status_current", "") if p.get("has_bom") else ""),
            (pj.get("producible_cover", "") if p.get("has_bom") else ""),
            p.get("avg_3jc_sales", 0), p.get("proj_flag", ""),
            (b["fg_stock"]["warehouse"] if b else ""), (b["fg_stock"]["branch"] if b else ""),
        ])
    return out


def _uniq_sheet(name, used) -> str:
    """Excel-safe, unique sheet name (<=31 chars)."""
    bad = set(r'\/?*[]:')
    clean = "".join(c for c in (name or "Sheet") if c not in bad).strip()[:31] or "Sheet"
    n, k = clean, 2
    while n.lower() in used:
        tag = f" ({k})"
        n = clean[:31 - len(tag)] + tag
        k += 1
    used.add(n.lower())
    return n


def _stock_audit_rows(products, lots, loc):
    """Lot-wise stock rows for the report's FGs at ``loc`` (warehouse/branch),
    matched by item DESCRIPTION so every code (bulk + packed) of the product shows."""
    fg = {}
    for p in products:
        if p.get("has_bom"):
            fg.setdefault(_pf._squash(p["name"]), p)
    rows = []
    for lt in (lots or []):
        if lt.get("location") != loc:
            continue
        p = fg.get(_pf._squash(lt.get("item_desc")))
        if not p:
            continue
        # Segment 2/3 come from the stock item's own CRM category (populated on the lot);
        # fall back to the plan product's segments if the lot lookup was empty.
        seg2 = lt.get("segment2") or p.get("segment2", "")
        seg3 = lt.get("segment3") or p.get("segment3", "")
        # Item Name sits right after Item Code and is the description of THAT specific
        # code (bulk / packed variant), so each code's real name is visible.
        rows.append([p["name"], seg2, seg3,
                     _ACT_LABEL.get(p.get("bom_class"), p.get("bom_class", "")),
                     lt.get("item_code"), lt.get("item_desc"),
                     lt.get("org"), lt.get("subinv"), lt.get("lot"),
                     round(lt.get("qty") or 0), lt.get("aging_date", "")])
    rows.sort(key=lambda r: (r[0], -r[9]))
    return rows


def _add_stock_audit(wb, st, products, lots):
    """Two lot-wise audit sheets — Warehouse Stock and Branch Stock — listing each
    FG's on-hand lots with Organisation, Sub Inventory, Lot and qty (audit trail)."""
    hdr = ["Product", "Segment 2", "Segment 3", "Activity", "Item Code", "Item Name",
           "Organisation", "Sub Inventory", "Lot Number", "Qty (Kg)", "Aging Date"]
    wid = [30, 20, 18, 13, 16, 30, 24, 16, 20, 12, 12]
    for loc, title in (("warehouse", "Warehouse Stock"), ("branch", "Branch Stock")):
        _grid(wb, st, title, hdr, _stock_audit_rows(products, lots, loc), wid, {10}, qty_fmt="#,##0")


def _add_intransit_audit(wb, st, rp, lots):
    """Audit sheets — one row per open PO line behind the in-transit totals (PO#,
    vendor, ordered/received/cancelled/balance). 'PO In-transit' = lines for RMs used
    in this report; 'In-transit Unmatched' = the same detail for lines NOT tied to any
    planned BOM RM (validation). Only for the live-CRM source."""
    if not lots:
        return
    used_codes, used_descs = set(), set()
    for key in ("consolidated_rm", "consolidated_rm_packing"):
        for x in rp.get(key, []):
            used_codes.update(x.get("rm_codes", []))
            used_descs.add(_pf._squash(x.get("rm_desc", "")))

    def _matched(lt):
        return lt.get("item_code") in used_codes or _pf._squash(lt.get("item_desc")) in used_descs

    # RM-source orgs (the raw-material filtering orgs) — the Unmatched sheet is limited
    # to these so it surfaces only RM-relevant in-transit, not depot/port/marketplace POs.
    rm_orgs = {_pf._norm(o).upper() for o in (rp.get("rules") or {}).get("rm_source_orgs", [])}
    # decode map: encoded item name -> real name (e.g. RDNBP101 -> PUREPRINT AFT),
    # keyed by both item code and squashed name (multi-code intermediates).
    dmap = rp.get("decode_map") or {}

    def _decoded(lt):
        return dmap.get(lt["item_code"]) or dmap.get(_pf._squash(lt.get("item_desc"))) or lt.get("item_desc")

    wid = [30, 22, 16, 14, 12, 12, 26, 24, 15, 12, 12, 11, 13]

    def _sheet(title, first_hdr, pred):
        rows = sorted((lt for lt in lots if pred(lt)),
                      key=lambda r: (r.get("org", ""), r.get("item_desc", ""), -r.get("in_transit", 0)))
        if not rows:
            return
        _grid(wb, st, title,
              [first_hdr, "Encoded Name", "Item Code", "Business", "PO Number", "PO Date", "Vendor", "Organisation",
               "Procurement", "Ordered", "Received", "Cancelled", "In-transit"],
              [[_decoded(r), r["item_desc"], r["item_code"], r.get("business", ""),
                r["po_number"], r["po_date"], r["vendor"], r["org"], r["procurement_type"],
                r["ordered"], r["received"], r["cancelled"], r["in_transit"]]
               for r in rows],
              wid, {10, 11, 12, 13}, qty_fmt="#,##0")

    def _rm_org(lt):
        return not rm_orgs or _pf._norm(lt.get("org")).upper() in rm_orgs

    def _not_packing(lt):
        # packing material (item code starting 'P') is excluded from these RM in-transit sheets
        return not _pf._pack_code(lt.get("item_code") or "")

    # Both sheets are limited to the RM orgs (MFG + trading) and exclude packing material,
    # so they show only raw-material-relevant in-transit (not depot/port/marketplace or packing).
    _sheet("PO In-transit", "RM Name (decoded)", lambda lt: _matched(lt) and _rm_org(lt) and _not_packing(lt))
    _sheet("In-transit Unmatched", "Item Name (decoded)", lambda lt: (not _matched(lt)) and _rm_org(lt) and _not_packing(lt))


def _add_reference_sheet(wb, st, rules):
    """A colour-templated 'Reference' sheet for the BU-shared files: an understanding
    note + the organization matrix (which orgs are in each RM/FG filtration list)."""
    from openpyxl.styles import Font, PatternFill, Alignment
    rules = rules or {}
    ws = wb.create_sheet("Reference")
    ws.sheet_view.showGridLines = False
    NAVY = PatternFill("solid", fgColor=_NAVY)
    TEAL = PatternFill("solid", fgColor=_TEAL)
    NOTEBG = PatternFill("solid", fgColor="FFF9E6")
    YES = PatternFill("solid", fgColor="C6EFCE")
    MFGF = PatternFill("solid", fgColor="FCE4D6")
    wrap = Alignment(wrap_text=True, vertical="center")
    ctr = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A1:G2")
    t = ws["A1"]; t.value = "Reference — Projection Confirmation (to share with BU)"
    t.fill = NAVY; t.font = Font(bold=True, color="FFFFFF", size=15)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 18; ws.row_dimensions[2].height = 18

    def band(row, text):
        ws.merge_cells(f"A{row}:G{row}")
        c = ws[f"A{row}"]; c.value = text; c.fill = TEAL
        c.font = Font(bold=True, color="FFFFFF", size=12)
        c.alignment = Alignment(indent=1, vertical="center")
        ws.row_dimensions[row].height = 22

    r = 4
    band(r, "Understanding Note"); r += 1
    notes = [
        'Projection Quantity does NOT include Pending SOC  (column name: "JC5 Qty (WK1+2) Qty(Kg)").',
        "Branch always serves Branch SOCs.",
        "For Planning, Confirmed Quantity is considered EXCLUDING SOC.",
        "For Planning, Pending SOCs of MFG and our Warehouse Orgs only are considered.",
    ]
    for i, note in enumerate(notes, 1):
        n = ws.cell(row=r, column=1, value=i); n.alignment = ctr; n.fill = NOTEBG
        n.font = Font(bold=True, color=_NAVY)
        ws.merge_cells(f"B{r}:G{r}")
        c = ws.cell(row=r, column=2, value=note); c.alignment = wrap; c.fill = NOTEBG; c.font = Font(size=11)
        ws.row_dimensions[r].height = 28
        r += 1

    r += 1
    band(r, "Organization Matrix"); r += 1
    ws.merge_cells(f"A{r}:G{r}")
    sub = ws.cell(row=r, column=1, value="Organisations included in each RM / FG filtration list used by the plan.")
    sub.font = Font(italic=True, color="718096", size=10); r += 1

    rm_src = set(rules.get("rm_source_orgs", []))
    interm = set(rules.get("intermediate_stock_orgs", []))
    itr = set(rules.get("intransit_rm_only_orgs", []))
    wh = set(rules.get("warehouse_orgs", []))
    soc = set(rules.get("mfg_soc_orgs", []))
    universe = sorted(rm_src | interm | itr | wh | soc)

    hdr = ["Organization", "MFG org (name)", "RM Source", "Intermediate Stock",
           "In-transit RM-only", "Warehouse", "MFG SOC Pending"]
    hr = r
    for c, name in enumerate(hdr, 1):
        cell = ws.cell(row=hr, column=c, value=name)
        cell.fill = NAVY; cell.font = st["white"]; cell.alignment = st["ctr"]; cell.border = st["border"]
    ws.row_dimensions[hr].height = 30
    rr = hr + 1
    is_mfg = lambda o: "mfg" in (o or "").lower()
    for o in universe:
        ws.cell(row=rr, column=1, value=o).border = st["border"]
        for j, f in enumerate([is_mfg(o), o in rm_src, o in interm, o in itr, o in wh, o in soc], 2):
            cell = ws.cell(row=rr, column=j, value="Yes" if f else "")
            cell.alignment = ctr; cell.border = st["border"]
            if f:
                cell.fill = MFGF if j == 2 else YES
        rr += 1
    if not universe:
        ws.cell(row=rr, column=1, value="(no organisation lists configured)").font = Font(italic=True, color="718096")

    _widths(ws, [34, 15, 12, 16, 16, 12, 15])
    ws.freeze_panes = f"A{hr + 1}"
    return ws


def _seg_coll_hdr(pjc=0):
    """Header for the product x collector projection sheets."""
    return ["#", "Product", "Segment 3", "Collector (customer)",
            (f"JC{pjc} Qty (WK1+2) Qty(Kg)" if pjc else "JC Qty (WK1+2) Qty(Kg)"),
            (f"JC{pjc + 1} Qty(Kg)" if pjc else "Next 1 JC Qty(Kg)"),
            (f"JC{pjc + 2} Qty(Kg)" if pjc else "Next 2 JC Qty(Kg)"), "Total Qty(Kg)"]


_SEG_COLL_WID = [5, 34, 22, 34, 18, 14, 14, 13]
_SEG_COLL_NUM = {5, 6, 7, 8}


def _collector_rows(prods, proj_by_name):
    """Per (product x collector) projection rows for a product bucket, so the business
    head can see how much projection each collector gave for each product. Products are
    ordered by Current qty; within a product, collectors by their Current qty."""
    out = []
    i = 0
    for p in sorted(prods, key=lambda x: -_seg_current(x)):
        crows = sorted(proj_by_name.get(_pf._squash(p["name"]), []),
                       key=lambda r: -(r.get("current") or 0))
        for r in crows:
            cur = r.get("current", 0) or 0
            n1 = r.get("next1", 0) or 0
            n2 = r.get("next2", 0) or 0
            if (cur + n1 + n2) <= 0:
                continue
            i += 1
            out.append([i, p["name"], p.get("segment3", ""), r.get("collector") or "—",
                        round(cur, 1), round(n1, 1), round(n2, 1), round(cur + n1 + n2, 1)])
    return out


_SOC_COLL_HDR = ["#", "Product", "Segment 3", "Collector (customer)", "MFG SOC Pending Qty(Kg)"]
_SOC_COLL_WID = [5, 34, 22, 34, 24]
_SOC_COLL_NUM = {5}


def _soc_collector_rows(prods, soc_by_name):
    """Per (product x collector) MFG SOC pending rows for a product bucket, so the
    business head sees how much pending sale-order qty each collector holds per product."""
    out = []
    i = 0
    for p in sorted(prods, key=lambda x: -_seg_current(x)):
        srows = sorted(soc_by_name.get(_pf._squash(p["name"]), []),
                       key=lambda r: -(r.get("qty") or 0))
        for r in srows:
            q = r.get("qty", 0) or 0
            if q <= 0:
                continue
            i += 1
            out.append([i, p["name"], p.get("segment3", ""), r.get("collector") or "—", round(q, 1)])
    return out


def _one_segment_workbook(s2, prods, cycle="", stock_lots=None, pjc=0, rules=None,
                          proj_by_name=None, soc_by_name=None) -> bytes:
    """One Segment-2 workbook: a separate sheet per Segment 3 for MANUFACTURING
    items, plus a single Others (rest) sheet. Each product sheet is followed by a
    product x collector projection sheet and a product x collector MFG SOC sheet."""
    from collections import defaultdict
    st = _styles()
    wb = _book()
    seg_hdr = _seg_hdr(pjc)
    mfg = [p for p in prods if p.get("bom_class") == "manufacturing"]
    oth = [p for p in prods if p.get("bom_class") != "manufacturing"]
    by_s3: dict = defaultdict(list)
    for p in mfg:
        by_s3[p.get("segment3") or "Unspecified"].append(p)
    _cover(wb, st, f"Supply & RM — {s2}",
           f"Cycle {cycle} · Manufacturing split by Segment 3 · Others in one sheet",
           [("Products", len(prods), _BAND), ("Manufacturing", len(mfg), _GREEN),
            ("Segment-3 sheets (Mfg)", len(by_s3), _BAND), ("Others (rest)", len(oth), _AMBER)])
    _add_reference_sheet(wb, st, rules)
    used = {"summary", "reference"}
    coll_hdr = _seg_coll_hdr(pjc)

    def _coll_sheet(label, bucket):
        # per-collector sheets right after the product sheet: projection, then MFG SOC
        if proj_by_name:
            crows = _collector_rows(bucket, proj_by_name)
            if crows:
                _grid(wb, st, _uniq_sheet(f"{label[:20]} Collector", used), coll_hdr, crows,
                      _SEG_COLL_WID, _SEG_COLL_NUM, qty_fmt="#,##0")
        if soc_by_name:
            srows = _soc_collector_rows(bucket, soc_by_name)
            if srows:
                _grid(wb, st, _uniq_sheet(f"{label[:16]} MFG SOC", used), _SOC_COLL_HDR, srows,
                      _SOC_COLL_WID, _SOC_COLL_NUM, qty_fmt="#,##0")

    # Segment-3 (Manufacturing) sheets ordered by highest total Current qty first
    for s3 in sorted(by_s3, key=lambda k: -sum(_seg_current(p) for p in by_s3[k])):
        _grid(wb, st, _uniq_sheet(s3, used), seg_hdr, _segment_rows(by_s3[s3]),
              _SEG_WID, _SEG_NUM, qty_fmt="#,##0")
        _coll_sheet(s3, by_s3[s3])
    if oth:
        _grid(wb, st, _uniq_sheet("Others", used), seg_hdr, _segment_rows(oth),
              _SEG_WID, _SEG_NUM, qty_fmt="#,##0")
        _coll_sheet("Others", oth)
    _add_stock_audit(wb, st, prods, stock_lots)
    wb.move_sheet("Summary", -(len(wb.sheetnames) - 1))
    return _save(wb)


def _safe_filename(name: str) -> str:
    bad = set('\\/:*?"<>|')
    return "".join(c for c in (name or "Unspecified") if c not in bad).strip()[:80] or "Unspecified"


# Some Segment-2 divisions are exported as SEPARATE files split by Segment 3, instead
# of one combined file. For each such division: a list of (file_name, [segment3 values]);
# the sentinel "*" collects every product not matched by an earlier bucket into that file.
# Each resulting file still gets its own Segment-3 sheets, Warehouse/Branch Stock and
# Reference sheet. Matching is on the squashed Segment 3 (case/space/punctuation-insensitive).
SEGMENT2_FILE_SPLIT = {
    "Textile & Paper Division": [
        ("Textile(PTJ)", ["Textile(PTJ)"]),
        ("Paper Pureco", ["Paper Pureco"]),
        ("Textile Pure", ["Textile Pure"]),
        # rest of the division (e.g. the 'Textile' segment-3); named "Textiles-Others"
        # so it never clashes with a standalone "Others" division's own Others.xlsx.
        ("Textiles-Others", "*"),
    ],
}


def build_rm_by_segment_zip(rp, cycle="", stock_lots=None, proj_rows=None, soc_rows=None) -> bytes:
    """A ZIP with one Excel FILE per Segment 2 (each split Manufacturing / Others).
    Divisions listed in SEGMENT2_FILE_SPLIT are exported as several files split by
    Segment 3 instead of one combined file. When ``proj_rows`` (per item x collector
    projection) and/or ``soc_rows`` (per item x collector MFG SOC pending) are supplied,
    each product sheet is followed by the matching per-collector sheet(s)."""
    import io
    import zipfile
    from collections import defaultdict
    seg: dict = defaultdict(list)
    for p in rp.get("products", []):
        seg[(p.get("segment2") or "Unspecified")].append(p)
    proj_by_name: dict = defaultdict(list)
    for r in (proj_rows or []):
        proj_by_name[_pf._squash(r.get("name"))].append(r)
    soc_by_name: dict = defaultdict(list)
    for r in (soc_rows or []):
        soc_by_name[_pf._squash(r.get("name"))].append(r)
    buf = io.BytesIO()
    used: set = set()
    pjc = rp.get("planning_jc") or 0
    rules = rp.get("rules", {})

    def _uniq_name(label):
        fn = _safe_filename(label)
        name, k = fn, 2
        while name.lower() in used:
            name = f"{fn} ({k})"; k += 1
        used.add(name.lower())
        return name

    def _write(zf, file_label, title, prods):
        if prods:
            zf.writestr(f"{_uniq_name(file_label)}.xlsx",
                        _one_segment_workbook(title, prods, cycle, stock_lots, pjc, rules,
                                              proj_by_name, soc_by_name))

    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for s2 in sorted(seg):
            split = SEGMENT2_FILE_SPLIT.get(s2)
            if not split:
                _write(zf, s2, s2, seg[s2])
                continue
            # split this division into separate files by Segment 3
            assigned, rest_label = set(), None
            for label, match in split:
                if match == "*":
                    rest_label = label
                    continue
                wanted = {_pf._squash(x) for x in match}
                bucket = [p for p in seg[s2] if _pf._squash(p.get("segment3")) in wanted]
                assigned.update(id(p) for p in bucket)
                # bare file names (label.xlsx); labels are chosen to be collision-free.
                _write(zf, label, f"{s2} — {label}", bucket)
            if rest_label is not None:
                rest = [p for p in seg[s2] if id(p) not in assigned]
                _write(zf, rest_label, f"{s2} — {rest_label}", rest)
    return buf.getvalue()


def build_msl_workbook(meta: dict, rows: list, activity: str | None = None) -> bytes:
    """MSL (Minimum Stock Level) report — one row per finished product with the average
    one-JC sales, movement frequency, customer coverage and MSL (50% of avg one-JC)."""
    st = _styles()
    wb = _book()
    if activity:
        rows = [r for r in rows if r.get("activity") == activity]
    total_msl = round(sum(r.get("msl", 0) for r in rows), 1)
    _cover(wb, st, f"MSL — {meta.get('reference', 'current')}"
           + (f" · {activity}" if activity else ""),
           f"{meta.get('jc_label', '')} · {meta.get('n_jcs', 13)}-JC window "
           f"{meta.get('jc_from', '')} → {meta.get('jc_to', '')} · FY {meta.get('fy', '')}",
           [("Items", len(rows), _BAND), ("Total MSL (KG)", total_msl, _AMBER),
            ("JCs (window)", meta.get("n_jcs", 13), _BAND),
            ("Reference", meta.get("reference", ""), _GREEN)])
    hdr = ["Item Name", "Activity", "Business",
           "Avg Qty / JC (Avg JC Sales)", "Frequency (JCs moved / 13)",
           "Customer Coverage (unique)", "Total Qty (13 JC)", "MSL (50% of Avg JC)",
           "On-hand: Warehouse", "On-hand: Branch", "On-hand Total"]
    wid = [36, 16, 18, 22, 20, 20, 16, 18, 16, 16, 15]
    grid = [[r.get("item_name", ""), r.get("activity", ""), r.get("business", ""),
             r.get("avg_qty_per_jc", 0), f"{r.get('freq_jcs', 0)} / {meta.get('n_jcs', 13)}",
             r.get("customer_coverage", 0), r.get("total_qty", 0), r.get("msl", 0),
             r.get("warehouse_stock", 0), r.get("branch_stock", 0), r.get("onhand_stock", 0)]
            for r in rows]
    _grid(wb, st, "MSL", hdr, grid, wid, {4, 6, 7, 8, 9, 10, 11}, qty_fmt="#,##0")

    # per-JC dispatch breakdown (only when live rows carry jc_qty)
    if rows and rows[0].get("jc_qty"):
        jlabels = [j.get("label", f"JC{i+1}") + " " + (j.get("from", "")[:7]) for i, j in enumerate(meta.get("jcs", []))]
        jhdr = ["Item Name", "Activity"] + (jlabels or [f"JC{i+1}" for i in range(13)])
        jrows = [[r["item_name"], r["activity"], *r.get("jc_qty", [])] for r in rows]
        jwid = [36, 16] + [11] * len(jlabels or range(13))
        _grid(wb, st, "By JC (dispatch)", jhdr, jrows, jwid, set(range(3, 3 + len(jlabels or range(13)))), qty_fmt="#,##0")

    wb.move_sheet("Summary", -(len(wb.sheetnames) - 1))
    return _save(wb)


def _vooki_minmax(p):
    """(min, max) FG qty producible from current RM stock: min uses only main RM
    on hand; max uses main + substitutes + in-transit. Bottleneck = scarcest RM."""
    b = p["boms"][0] if p.get("boms") else None
    if not b:
        return 0.0, 0.0
    mn = mx = None
    for c in b["components"]:
        q = c["qty_per_unit"]
        if q and q > 0:
            mn = c["main_stock"] / q if mn is None else min(mn, c["main_stock"] / q)
            mx = c["available"] / q if mx is None else min(mx, c["available"] / q)
    return round(mn or 0.0, 1), round(mx or 0.0, 1)


def build_vooki_workbook(vp, quantities=None, cycle="", product=None, stock_rows=None, intransit_lots=None) -> bytes:
    st = _styles()
    q = quantities or {}
    products = vp.get("products", [])
    if product:
        products = [p for p in products if p["name"] == product]
    s = vp.get("summary", {})
    wb = _book()
    single = bool(product)   # per-FG report: product name is in the Summary, so
                             # the redundant Product column is dropped elsewhere

    cons: dict[str, dict] = {}
    rm_rows = []
    total_net, planned = 0.0, 0
    for p in products:
        qty = float(q.get(p["name"], 0) or 0)
        if qty > 0:
            planned += 1
        b = p["boms"][0] if p.get("boms") else None
        p_net = 0.0
        if b and qty > 0:
            for c in b["components"]:
                if _pf._pack_code(c["rm_code"]):    # packing material excluded from Vooki planning
                    continue
                gross = round(qty * c["qty_per_unit"], 1)
                net = round(max(0.0, gross - c["available"]), 1)
                p_net += net
                rm_rows.append([p["name"], b["assembly_item"], b.get("org_code", ""), b.get("designator", ""),
                                c["seq"], c["rm_code"], c["rm_desc"], c["qty_per_unit"], gross,
                                c["main_stock"], c["substitute_stock"], c["in_transit"], c["available"],
                                c.get("producible", 0), net, c.get("lead_time"),
                                ", ".join(f"{su['code']} ({su['stock']})" for su in c["substitutes"]),
                                ", ".join(f"{su['desc']} ({su['stock']})" for su in c["substitutes"])])
                ckey = (c["rm_desc"] or c["rm_code"]).upper()   # consolidate by item description
                a = cons.setdefault(ckey, {"code": c["rm_code"], "desc": c["rm_desc"], "gross": 0.0,
                                           "available": c["available"], "fgs": set()})
                a["gross"] += gross
                a["available"] = max(a["available"], c["available"])
                a["fgs"].add(p["name"])
        total_net += p_net

    cons_rows = []
    for _k, a in sorted(cons.items(), key=lambda kv: -(kv[1]["gross"] - kv[1]["available"])):
        cons_rows.append([a["code"], a["desc"], len(a["fgs"]), round(a["gross"], 1), round(a["available"], 1),
                          round(max(0.0, a["gross"] - a["available"]), 1), ", ".join(sorted(a["fgs"])[:20])])

    # Real RM Requirement — every BOM intermediate (e.g. PURE INFINITY) exploded to its
    # purchased leaf RMs. Gross scales with the entered FG qty; consolidated across FGs.
    real: dict = {}
    for p in products:
        pq = float(q.get(p["name"], 0) or 0)
        if pq <= 0:
            continue
        for e in p.get("real_rm", []):
            gross = round(pq * e["per_unit"], 1)
            if gross <= 0:
                continue
            rk = _pf._squash(e["desc"]) or e["code"]
            a = real.setdefault(rk, {"code": e["code"], "desc": e["desc"], "gross": 0.0,
                                     "available": e["available"], "via": set(), "fgs": set(),
                                     "unresolved": bool(e.get("unresolved"))})
            a["gross"] += gross
            a["available"] = max(a["available"], e["available"])
            a["via"].update(e.get("via", []))
            a["fgs"].add(p["name"])
            a["unresolved"] = a["unresolved"] or bool(e.get("unresolved"))
    real_rows, real_net_total = [], 0.0
    for _k, a in sorted(real.items(), key=lambda kv: -(kv[1]["gross"] - kv[1]["available"])):
        net = round(max(0.0, a["gross"] - a["available"]), 1)
        real_net_total += net
        real_rows.append([a["code"], a["desc"],
                          "Unresolved intermediate" if a["unresolved"] else ("Via intermediate" if a["via"] else "Direct RM"),
                          "; ".join(sorted(a["via"])[:6]) or "-",
                          len(a["fgs"]), round(a["gross"], 1), round(a["available"], 1), net,
                          ", ".join(sorted(a["fgs"])[:20])])
    real_net_total = round(real_net_total, 1)

    if product and products:
        p0 = products[0]
        mn, mx = _vooki_minmax(p0)
        _cover(wb, st, f"Vooki Planning — {p0['name']}",
               f"Cycle {cycle} · plan qty {round(float(q.get(p0['name'], 0) or 0), 1)} · stock: {s.get('stock_source', '—')}",
               [("Plan qty", round(float(q.get(p0["name"], 0) or 0), 1), _BAND),
                ("Min producible", mn, _GREEN),
                ("Max producible", mx, _GREEN),
                ("RM to buy", round(total_net, 1), _AMBER),
                ("Real RM to buy", real_net_total, _AMBER),
                ("FG stock (units)", p0.get("fg_units", 0), _BAND),
                ("FG stock (KG/Lit)", p0.get("fg_volume_l", 0), _BAND)])
    else:
        _cover(wb, st, "Vooki Planning — RM Requirement",
               f"Cycle {cycle} · FG {vp.get('rules', {}).get('fg_business', 'Vooki Division')} · stock: {s.get('stock_source', '—')}",
               [("Vooki products", s.get("products"), _BAND),
                ("Products planned", planned, _BAND),
                ("RM to buy (total)", round(total_net, 1), _AMBER),
                ("Real RM to buy", real_net_total, _AMBER),
                ("RM items in stock", s.get("rm_items_in_stock"), _BAND),
                ("FG stock (units)", s.get("fg_stock_units"), _GREEN),
                ("FG stock (KG/Lit)", s.get("fg_stock_volume_l"), _GREEN)])

    r_hdr = ["Product", "BOM Assembly", "Org", "Designator", "Seq", "RM Code", "RM Description",
             "Qty/unit", "Gross", "Stock", "Sub stk", "In-transit", "Available", "Producible",
             "Net to buy", "Lead time (d)", "Substitutes (stock)", "Substitute Item Desc"]
    r_w = [34, 16, 6, 12, 5, 16, 26, 10, 12, 11, 9, 10, 11, 12, 12, 12, 30, 34]
    r_q, lead_col = {8, 9, 10, 11, 12, 13, 14, 15, 16}, 16
    if single:
        r_hdr, r_w, rm_rows, r_q = _drop_first_col(r_hdr, r_w, rm_rows, r_q)
        lead_col = 15
    ws_rm = _grid(wb, st, "RM Requirement", r_hdr, rm_rows, r_w, r_q)
    _color_lead(ws_rm, lead_col)   # colour the Lead time column by threshold

    c_hdr = ["RM Code", "RM Description", "#FG", "Gross", "Available", "Net to buy", "Used in FGs"]
    c_w = [16, 28, 6, 12, 12, 12, 50]
    if single:   # single product -> the "Used in FGs" column is redundant
        c_hdr, c_w, cons_rows = c_hdr[:-1], c_w[:-1], [r[:-1] for r in cons_rows]
    _grid(wb, st, "Consolidated RM", c_hdr, cons_rows, c_w, {4, 5, 6})

    # Real RM Requirement — intermediates exploded to the true purchased (leaf) RMs.
    rr_hdr = ["RM Code", "RM Description", "Source", "Via intermediate(s)", "#FG",
              "Gross", "Available", "Net to buy", "Used in FGs"]
    rr_w, rr_rows = [16, 30, 20, 30, 6, 12, 12, 12, 46], real_rows
    if single:   # single product -> the "Used in FGs" column is redundant
        rr_hdr, rr_w, rr_rows = rr_hdr[:-1], rr_w[:-1], [r[:-1] for r in real_rows]
    _grid(wb, st, "Real RM Requirement", rr_hdr, rr_rows, rr_w, {5, 6, 7, 8})

    # Producible (current RM) — how many units of each FG can be made from the RM in
    # stock right now (MFG orgs, intermediates exploded to leaf RMs), and the bottleneck.
    pr_rows = []
    for p in products:
        if not p.get("has_bom"):
            continue
        pn = p.get("producible_now", 0) or 0
        pr_rows.append([p["name"], round(pn, 1), p.get("limiting_rm") or "—",
                        p.get("limiting_rm_available"), round(float(q.get(p["name"], 0) or 0), 1)])
    pr_rows.sort(key=lambda r: -r[1])
    if pr_rows:
        _grid(wb, st, "Producible (current RM)",
              ["Vooki Product", "Producible now (with current RM)", "Limiting RM",
               "Limiting RM available", "Plan Qty"],
              pr_rows, [34, 24, 30, 18, 12], {2, 4, 5})

    # RM In-transit — open-PO in-transit detail for the Vooki RMs (same per-PO-line
    # audit as the Supply & RM Plan page). Packing material is excluded.
    if intransit_lots:
        rm_codes, rm_descs = set(), set()
        for p in products:
            for e in p.get("real_rm", []):          # true (leaf) RMs incl. exploded intermediates
                if e.get("code"):
                    rm_codes.add(e["code"])
                rm_descs.add(_pf._squash(e["desc"]))
            for b in (p.get("boms") or [])[:1]:      # direct RM components too
                for c in b["components"]:
                    if _pf._pack_code(c["rm_code"]):
                        continue
                    rm_codes.add(c["rm_code"])
                    rm_descs.add(_pf._squash(c["rm_desc"]))
        it_rows = []
        for lt in intransit_lots:
            if _pf._pack_code(lt.get("item_code") or ""):
                continue
            if not (lt.get("item_code") in rm_codes or _pf._squash(lt.get("item_desc")) in rm_descs):
                continue
            it_rows.append([lt.get("item_desc", ""), lt.get("item_code", ""), lt.get("business", ""),
                            lt.get("po_number", ""), lt.get("po_date", ""), lt.get("vendor", ""),
                            lt.get("org", ""), lt.get("procurement_type", ""), lt.get("ordered", 0),
                            lt.get("received", 0), lt.get("cancelled", 0), lt.get("in_transit", 0)])
        it_rows.sort(key=lambda r: (r[0], -(r[11] or 0)))
        if it_rows:
            _grid(wb, st, "RM In-transit",
                  ["RM Name", "Item Code", "Business", "PO Number", "PO Date", "Vendor", "Organisation",
                   "Procurement", "Ordered", "Received", "Cancelled", "In-transit"],
                  it_rows, [30, 16, 16, 14, 12, 26, 24, 15, 12, 12, 11, 13], {9, 10, 11, 12}, qty_fmt="#,##0")

    # Available Stock — Vooki FG on-hand, shown both as manufactured bulk (KG/Lit) and
    # as consumer units (units counted from combo & pack SKUs, not just single units).
    fsr = vp.get("fg_stock_rows", [])
    if product:
        _pk = _pf._squash(product)
        fsr = [r for r in fsr if _pf._squash(r["product"]) == _pk] or fsr
    if fsr:
        av_hdr = ["Product", "SKU Code", "SKU Description", "Pack type", "Group", "UOM",
                  "Unit size (ml)", "Units/UOM", "Stock Qty", "Available Units", "Available Bulk (KG/Lit)"]
        av_w = [30, 16, 30, 22, 16, 8, 12, 10, 12, 14, 20]
        av_rows = [[r["product"], r["sku_code"], r["sku_name"], r["pack_type"], r["group"], r["uom"],
                    r["unit_ml"], r["units_each"], r["qty"], r["units"], r["bulk_l"]] for r in fsr]
        _grid(wb, st, "Available Stock", av_hdr, av_rows, av_w, {7, 8, 9, 10, 11})

    # Raw inventory: physical stock (org + sub-inventory) behind each planning RM
    if stock_rows:
        excl = {x.lower() for x in _pf.EXCLUDE_SUBINV}
        plan_keys = set()
        for p in products:
            if not p.get("boms"):
                continue
            for c in p["boms"][0]["components"]:
                if _pf._pack_code(c["rm_code"]):    # packing material excluded
                    continue
                plan_keys.add(_pf._squash(c["rm_desc"]))
                for su in c["substitutes"]:
                    plan_keys.add(_pf._squash(su["desc"]))
        inv_rows = []
        for r in stock_rows:
            desc = _pf._norm(r.get("ItemDesc"))
            if _pf._squash(desc) not in plan_keys:
                continue
            qty = round(_pf._num(r.get("Qty")), 1)
            if qty == 0:
                continue
            sub = _pf._norm(r.get("SubInv"))
            inv_rows.append([_pf._norm(r.get("ItemCode")), desc, _pf._norm(r.get("Organization")),
                             sub, qty, round(_pf._num(r.get("ItemCost")), 2),
                             "No" if sub.lower() in excl else "Yes"])
        inv_rows.sort(key=lambda x: (x[1], x[2], x[3]))
        if inv_rows:
            _grid(wb, st, "Raw Inventory",
                  ["RM Code", "RM Description", "Organization", "Sub Inventory", "Qty",
                   "Item Cost", "Counted for planning"],
                  inv_rows, [16, 30, 24, 18, 12, 11, 14], {5, 6})

    wb.move_sheet("Summary", -(len(wb.sheetnames) - 1))
    return _save(wb)


def build_plan_template_workbook(items, segment2="", segment3="") -> bytes:
    """Blank planning-input template. Headers: S.No, Item Description, Qty (Kg),
    Current JC, Next JC1, Next JC2. The Item Description column is a dropdown (LOV)
    of all FG in the chosen segment, and the rows are pre-listed with those items."""
    from openpyxl.worksheet.datavalidation import DataValidation
    st = _styles()
    items = list(items)
    wb = _book()
    ws = wb.create_sheet("Template")
    ws.sheet_view.showGridLines = False

    headers = ["S.No", "Item Description", "Current JC Qty (Kg)", "Next JC1 Qty (Kg)", "Next JC2 Qty (Kg)"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _hdr_row(ws, 1, len(headers), st)

    # rows are left blank for the user to fill; number format on the qty columns
    rows = max(500, len(items) + 50)
    for r in range(2, rows + 1):
        for col in (3, 4, 5):
            ws.cell(row=r, column=col).number_format = _QTY
    _widths(ws, [8, 48, 18, 18, 18])
    ws.freeze_panes = "A2"

    # hidden sheet holding the LOV; referenced by the dropdown data-validation
    lov = wb.create_sheet("Items")
    for i, nm in enumerate(items, 1):
        lov.cell(row=i, column=1, value=nm)
    lov.sheet_state = "hidden"

    if items:
        dv = DataValidation(type="list", formula1=f"=Items!$A$1:$A${len(items)}", allow_blank=True)
        dv.prompt = "Select an item from the list"
        dv.promptTitle = "Item Description"
        ws.add_data_validation(dv)
        dv.add(f"B2:B{rows}")

    wb.move_sheet("Template", -(len(wb.sheetnames) - 1))
    return _save(wb)


def _color_priority(ws, col):
    from openpyxl.styles import PatternFill, Font
    colors = {1: "1A7D4F", 2: "2A9D8F", 3: "D68910", 4: "E67E22", 5: "1768C4", 6: "C0392B"}
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=col)
        v = cell.value
        if isinstance(v, (int, float)) and int(v) in colors:
            cell.fill = PatternFill("solid", fgColor=colors[int(v)])
            cell.font = Font(bold=True, color="FFFFFF")


def build_production_schedule_workbook(rp, cycle="") -> bytes:
    from openpyxl.styles import PatternFill, Font
    st = _styles()
    s = rp.get("summary", {})
    plan = rp.get("plan", {})
    wb = _book()
    _cover(wb, st, "Production Job Scheduling",
           f"Plan #{plan.get('plan_id', '—')} · JC{plan.get('jc_number', '')} ({plan.get('plan_type', '')}) · "
           f"JC start {rp.get('jc_start', '')} · generated {rp.get('today', '')}",
           [("Manufacturing jobs", s.get("scheduled_jobs"), _BAND),
            ("Total batches", s.get("total_batches"), _BAND),
            ("Plant utilisation %", s.get("utilisation_pct"), _GREEN),
            (f"Equipment ({s.get('horizon_days','')} d horizon)", s.get("equipment_used"), _BAND),
            ("Skipped repack / no-BOM", f"{s.get('skipped_non_manufacturing', 0)} / {s.get('skipped_no_bom', 0)}", _AMBER),
            ("Calendar", f"{s.get('date_from','')} → {s.get('date_to','')}", _GREEN)])

    jobs = rp.get("jobs", [])
    rows = [[j["priority"], j["item"], j["organization"], j["product_type"], j["equipment"],
             j["scenario"], "Yes" if j["rm_available"] else "No", j["qty"], j["batches"],
             j["batch_size"], j["cycle_hrs"], j.get("dur_hours", ""), ("" if j["rm_available"] else j["lead_days"]),
             j["start"], j["end"]] for j in jobs]
    ws = _grid(wb, st, "Schedule",
               ["Priority", "Item", "Organization", "Product Type", "Equipment", "Scenario",
                "RM Available", "Qty (Kg)", "Batches", "Batch Size (Kg)", "Cycle (h)", "Job hrs", "Lead (d)", "Start", "End"],
               rows, [8, 30, 20, 14, 12, 13, 11, 12, 9, 14, 9, 9, 8, 12, 12], {8, 9, 10, 11, 12, 13})
    _color_priority(ws, 1)

    # Equipment utilisation over the horizon (bottleneck view)
    util = rp.get("utilisation", [])
    if util:
        _grid(wb, st, "Equipment Utilisation",
              ["Equipment", "Jobs", "Busy hours", "Busy days", f"Utilisation % (of {s.get('horizon_days','')} d)"],
              [[u["equipment"], u["jobs"], u["busy_hours"], u["busy_days"], u["util_pct"]] for u in util],
              [16, 8, 12, 11, 22], {1, 2, 3, 4})

    un = rp.get("unscheduled", [])
    if un:
        _grid(wb, st, "Unscheduled",
              ["Priority", "Item", "Scenario", "RM Available", "Qty (Kg)", "Reason"],
              [[u.get("priority"), u.get("item"), u.get("scenario"), "Yes" if u.get("rm_available") else "No",
                u.get("qty"), u.get("reason")] for u in un],
              [8, 32, 14, 11, 12, 26], {5})

    # priority legend on the Summary sheet
    cov = wb["Summary"]
    base = 12
    cov.cell(row=base, column=1, value="Priority legend").font = Font(bold=True, color=_NAVY)
    legend = ["Pending SOC + RM available", "Future SOC + RM available", "Pending SOC + RM not available",
              "Future SOC + RM not available", "No SOC + RM available", "No SOC + RM not available"]
    cmap = {1: "1A7D4F", 2: "2A9D8F", 3: "D68910", 4: "E67E22", 5: "1768C4", 6: "C0392B"}
    for i, txt in enumerate(legend, 1):
        c = cov.cell(row=base + i, column=1, value=f"P{i}")
        c.fill = PatternFill("solid", fgColor=cmap[i]); c.font = Font(bold=True, color="FFFFFF")
        cov.cell(row=base + i, column=2, value=txt)

    wb.move_sheet("Summary", -(len(wb.sheetnames) - 1))
    return _save(wb)


def build_aged_rm_report_workbook(rep, cycle="") -> bytes:
    """Aged-RM excess analysis report (the Report_Aged_RM template): per aged raw material,
    aged qty/value vs last-3-JC consumption, last-3-JC sales requirement and projection
    requirement, with an Excess / OK / Critical status."""
    st = _styles()
    wb = _book()
    jcs = rep.get("jc_numbers") or [3, 4, 5]
    while len(jcs) < 3:
        jcs = list(jcs) + [0]
    summ = rep.get("summary", {})
    _cover(wb, st, "Aged RM — Excess Analysis",
           f"Cycle {cycle} · aged stock vs last-3-JC consumption / sales / projection",
           [("Aged RM items", summ.get("items"), _BAND),
            ("Aged Qty (KG)", summ.get("total_q90"), _AMBER),
            ("Aged Value", summ.get("total_v90"), _AMBER),
            ("Critical (no demand)", summ.get("critical"), _AMBER),
            ("Excess", summ.get("excess"), _BAND)])
    hdr = ["Item Desc", "Category", "Q>90", "V>90",
           f"JC{jcs[0]}_consumption\nQty(Kg)", f"JC{jcs[1]}_consumption\nQty(Kg)",
           f"JC{jcs[2]}_consumption\nQty(Kg)", "Avg_Consumption\n(3JC) Qty",
           "Excess_By_3JCs\n_Avg_Con(Qty)", "Excess%_by_Avg_Con\n(%)",
           "RM_Req_for_3JCs(Sum of Sales)\nQty", "%Excess on Sales Sum\nQty", "Total FG Used",
           "FG Names", "Projection_Requirement\nQty", "Excess_By_Projection\nQty", "Excess%_by_Projection\n%",
           "Status", "Remarks"]

    def _c(r, i):
        c = r.get("jc_consumption", [])
        return c[i] if i < len(c) else 0

    grid = [[r["item_desc"], r["category"], r["q90"], r["v90"],
             _c(r, 0), _c(r, 1), _c(r, 2), r["avg_consumption"],
             r["excess_avg"], r["excess_avg_pct"], r["rm_req_sales"], r["excess_sales_pct"],
             r["fg_used"], r.get("fg_names", ""), r["proj_req"], r["excess_proj"], r["excess_proj_pct"],
             r["status"], r["remarks"]]
            for r in rep.get("rows", [])]
    wid = [30, 15, 12, 14, 13, 13, 13, 15, 16, 16, 20, 16, 11, 50, 18, 16, 16, 13, 46]
    _grid(wb, st, "Aged RM Report", hdr, grid, wid, {3, 4, 5, 6, 7, 8, 9, 11, 13, 15, 16}, qty_fmt="#,##0")
    wb.move_sheet("Summary", -(len(wb.sheetnames) - 1))
    return _save(wb)


def build_aged_rm_workbook(rp, cycle="") -> bytes:
    st = _styles()
    s = rp.get("summary", {})
    wb = _book()
    _cover(wb, st, f"Aged RM → FG Production (aged > {rp.get('aged_days', 90)}d)",
           f"Cycle {cycle} · {rp.get('rm_filter','')} · maximise aged-inventory consumption",
           [("Aged RM items", s.get("aged_rm_items"), _BAND),
            ("Aged RM qty (KG)", s.get("aged_rm_qty"), _BAND),
            ("FGs producible", s.get("fgs_producible_from_aged"), _BAND),
            ("FGs needing purchase", s.get("fgs_needing_purchase"), _AMBER),
            ("Aged consumed (KG)", s.get("aged_consumed_qty"), _GREEN),
            ("Utilisation %", s.get("utilisation_pct"), _GREEN)])
    _grid(wb, st, "Recommended Production",
          ["#", "Finished good", "Assembly", "Produce (units)", "Aged consumed (KG)",
           "Aged value", "Cumulative aged (KG)", "Aged RMs used"],
          [[i, r["name"], r["assembly_item"], r["produce_units"], r["aged_consumed"],
            r["aged_value_consumed"], r["cumulative_aged_consumed"],
            ", ".join(f"{u['rm_desc']}" for u in r["rms_used"][:8])]
           for i, r in enumerate(rp.get("recommended", []), 1)],
          [5, 32, 18, 14, 16, 14, 16, 60], {4, 5, 6, 7})
    _grid(wb, st, "Producible from Aged",
          ["Finished good", "Assembly", "Producible (units)", "Aged consumable (KG)", "Inputs", "Components"],
          [[r["name"], r["assembly_item"], r["producible_units"], r["aged_consumed"],
            "needs fresh" if r["needs_fresh"] else "all aged", len(r["components"])]
           for r in rp.get("producible", [])],
          [32, 18, 16, 18, 12, 11], {3, 4})
    _grid(wb, st, "Needs Purchase",
          ["Finished good", "Assembly", "# missing RM", "Missing raw materials"],
          [[r["name"], r["assembly_item"], len(r["missing"]), ", ".join(r["missing"][:8])]
           for r in rp.get("blocked", [])], [32, 18, 12, 70], {3})
    _grid(wb, st, "Unused Aged RM",
          ["RM Code", "Description", "Aged qty (KG)", "Oldest age (days)", "Value"],
          [[r["rm_code"], r["rm_desc"], r["qty"], r["max_age"], r["value"]]
           for r in rp.get("unused_aged_rm", [])], [16, 34, 14, 16, 14], {3, 4, 5})
    return _save(wb)


def build_projection_sales_workbook(rp, cycle="") -> bytes:
    from openpyxl.styles import PatternFill
    st = _styles()
    s = rp.get("summary", {})
    flag_fill = {"over": PatternFill("solid", fgColor=_RED), "under": PatternFill("solid", fgColor=_AMBER),
                 "ontrack": PatternFill("solid", fgColor=_GREEN), "new": PatternFill("solid", fgColor="E6E6FA")}
    wb = _book()
    _cover(wb, st, "Projection vs Sales — under / over projection",
           f"Cycle {cycle} · projection (JC4 WK1+WK2) vs last {rp.get('n_jc',3)}-JC dispatched · ±{rp.get('band_pct',20)}%",
           [("Items", s.get("items"), _BAND), ("Over-projected", s.get("item_over"), _RED),
            ("Under-projected", s.get("item_under"), _AMBER), ("On track", s.get("item_ontrack"), _GREEN),
            ("Manufactured / Traded", f"{s.get('manufactured')} / {s.get('traded')}", _BAND),
            ("Collector×Item rows", s.get("collector_items"), _BAND)])
    _grid(wb, st, "Consolidated Item-based",
          ["Item", "Type", "Segment 2", "Segment 3", "Current (JC)", "Next 1", "Next 2",
           "Avg 3-JC sales", "Variance", "Var %", "Flag", "Warehouse", "Branch", "Stock total"],
          [[i["name"], i.get("item_type"), i["segment2"], i["segment3"], i["current"], i["next1"],
            i["next2"], i["avg_3jc_sales"], i["variance"], i.get("variance_pct"), i["flag"],
            i["warehouse"], i["branch"], i["stock_total"]] for i in rp.get("items", [])],
          [30, 13, 20, 18, 12, 11, 11, 13, 12, 9, 9, 12, 12, 12], {5, 6, 7, 8, 9, 12, 13, 14}, flag_fill, 11)
    _grid(wb, st, "Collector-Item based",
          ["Collector", "Item", "Type", "Segment 2", "Segment 3", "Current (JC)", "Next 1", "Next 2",
           "Avg 3-JC sales", "Variance", "Flag"],
          [[c["collector"], c["name"], c.get("item_type"), c["segment2"], c["segment3"], c["current"],
            c["next1"], c["next2"], c["avg_3jc_sales"], c["variance"], c["flag"]]
           for c in rp.get("collector_items", [])],
          [22, 30, 13, 20, 18, 12, 11, 11, 13, 12, 9], {6, 7, 8, 9, 10}, flag_fill, 11)
    return _save(wb)


def build_projection_accuracy_workbook(rp, cycle="") -> bytes:
    from openpyxl.styles import PatternFill
    st = _styles()
    s = rp.get("summary", {})
    sc = rp.get("scope", {})
    wb = _book()
    _cover(wb, st, "Projection Accuracy — projection vs actual production",
           f"Cycle {cycle} · {sc.get('acc_year', '')} · {sc.get('label', '')} · "
           f"actual = Output Qty per unique job · files: {', '.join(sc.get('files', []))}",
           [("Items", s.get("n_items"), _BAND),
            ("Projected (KG)", s.get("projected"), _BAND),
            ("Actual produced (KG)", s.get("actual"), _BAND),
            ("Accuracy %", s.get("accuracy_pct"), _GREEN),
            ("Bias %", s.get("bias_pct"), _AMBER),
            ("WMAPE / MAPE %", f"{s.get('wmape')} / {s.get('mape')}", _RED)])

    def _accfill(v):
        try:
            v = float(v)
        except (TypeError, ValueError):
            return None
        return PatternFill("solid", fgColor=_GREEN if v >= 80 else _AMBER if v >= 50 else _RED)

    ws = _grid(wb, st, "Item accuracy",
               ["Item", "Item Code", "Division", "Product", "UoM", "#Jobs", "Projected",
                "Actual", "Variance", "Var %", "Abs Err %", "Accuracy %", "Bias %", "Status"],
               [[i["item_desc"], i.get("item_code", ""), i["division"], i["product"],
                 i.get("uom", ""), i.get("jobs", 0), i["projected"], i["actual"], i["variance"],
                 i.get("variance_pct"), i.get("abs_pct_err"), i.get("accuracy_pct"),
                 i.get("bias_pct"), i["status"]] for i in rp.get("items", [])],
               [30, 16, 22, 20, 6, 7, 13, 13, 12, 9, 10, 11, 9, 24], {6, 7, 8, 9, 10, 11, 12, 13})
    for r in range(2, len(rp.get("items", [])) + 2):
        f = _accfill(ws.cell(row=r, column=12).value)
        if f:
            ws.cell(row=r, column=12).fill = f

    def _grp(title, lst):
        w = _grid(wb, st, title,
                  ["Name", "#Items", "Projected", "Actual", "Variance", "Var %",
                   "MAPE %", "WMAPE %", "Accuracy %", "Bias %"],
                  [[g["name"], g["n_items"], g["projected"], g["actual"], g["variance"],
                    g.get("variance_pct"), g.get("mape"), g.get("wmape"),
                    g.get("accuracy_pct"), g.get("bias_pct")] for g in lst],
                  [28, 8, 13, 13, 12, 9, 9, 9, 11, 9], {2, 3, 4, 5, 6, 7, 8, 9, 10})
        for r in range(2, len(lst) + 2):
            f = _accfill(w.cell(row=r, column=9).value)
            if f:
                w.cell(row=r, column=9).fill = f

    _grp("Division accuracy", rp.get("divisions", []))
    _grp("Product accuracy", rp.get("products", []))
    return _save(wb)


def build_srdms_reports_workbook(tat_rows, dash) -> bytes:
    """SRDMS reporting workbook: TAT, Pending & Ageing, Holds-overdue."""
    st = _styles()
    wb = _book()
    tot = (dash or {}).get("totals", {})
    _cover(wb, st, "SRDMS — Sample Request & Dispatch reports",
           "TAT (turn-around-time) and pending / ageing across sample requests",
           [("Open", tot.get("open"), _BAND), ("Closed", tot.get("closed"), _GREEN),
            ("Lines on hold", tot.get("held_lines"), _AMBER),
            ("Ack SLA breached", tot.get("ack_breach"), _RED),
            ("Holds overdue", tot.get("hold_overdue"), _RED),
            ("Avg total TAT (h)", (dash or {}).get("avg_total_tat_h"), _BAND)])

    _grid(wb, st, "TAT",
          ["SR No", "Requester", "Plant", "Priority", "Status", "Submitted", "Acknowledged",
           "Closed", "Ack TAT (h)", "Total TAT (h)", "Open age (d)", "Lines"],
          [[r["sr_no"], r["requester"], r["plant"], r["priority"], r["status"], r["submitted_at"],
            r["acknowledged_at"], r["closed_at"], r["ack_tat_h"], r["total_tat_h"],
            r["open_age_days"], r["lines"]] for r in (tat_rows or [])],
          [20, 22, 20, 10, 16, 18, 18, 18, 12, 13, 12, 7], {9, 10, 11, 12})

    _grid(wb, st, "Pending & Ageing",
          ["SR No", "Requester", "Plant", "Priority", "Status", "Open lines", "Age (days)"],
          [[r["sr_no"], r["requester_name"], r["plant_name"], r["priority"], r["status"],
            r["open_lines"], r["tat"]["open_age_days"]] for r in (dash or {}).get("open_requests", [])],
          [20, 22, 20, 10, 16, 11, 11], {6, 7})

    _grid(wb, st, "Holds overdue",
          ["SR No", "Item", "Planned date", "Reason", "Requester"],
          [[h["sr_no"], h["item"], h["planned_date"], h["reason"], h["requester"]]
           for h in (dash or {}).get("hold_overdue_lines", [])],
          [20, 30, 14, 24, 22], set())
    return _save(wb)


def build_supplier_scorecard_workbook(rp, cycle="") -> bytes:
    from openpyxl.styles import PatternFill
    st = _styles()
    s = rp.get("summary", {})
    suppliers = rp.get("suppliers", [])

    def sfill(v):
        try:
            v = float(v)
        except (TypeError, ValueError):
            return None
        return PatternFill("solid", fgColor=_GREEN if v >= 75 else _AMBER if v >= 50 else _RED)

    wb = _book()
    _cover(wb, st, "Supplier Scorecard — RM purchasing", f"Cycle {cycle} · {s.get('weights','')}",
           [("Suppliers", s.get("suppliers"), _BAND), ("Rated", s.get("rated_suppliers"), _BAND),
            ("Items supplied", s.get("items_supplied"), _BAND), ("Imports", s.get("imports"), _AMBER),
            ("Critical / sole-source", f"{s.get('critical')} / {s.get('sole_source')}", _RED),
            ("Avg score", s.get("avg_score"), _GREEN)])
    ws = _grid(wb, st, "Suppliers",
               ["Vendor", "Score", "OTD %", "OTIF %", "Fill %", "Avg lead (d)", "Price vs mkt %",
                "PO lines", "Items", "Spend (INR)", "Trade", "Criticality", "Sole-source #",
                "Criticality reasons", "Currency", "Locations"],
               [[x["vendor"], x["score"], x["otd"], x["otif"], x["fill_rate"], x["avg_lead_time"],
                 x["price_vs_market"], x["po_lines"], x["item_count"], x["spend"], x["trade"],
                 x["criticality"], x["sole_source_count"], "; ".join(x["criticality_reasons"]),
                 ", ".join(x["currencies"]), ", ".join(x["locations"])] for x in suppliers],
               [38, 8, 8, 8, 8, 12, 13, 9, 8, 14, 10, 11, 11, 36, 10, 28], {10})
    for r in range(2, len(suppliers) + 2):
        f = sfill(ws.cell(row=r, column=2).value)
        if f:
            ws.cell(row=r, column=2).fill = f
    rows = []
    for x in suppliers:
        for it in x["items"]:
            rows.append([x["vendor"], x["trade"], it["code"], it["name"], it["lines"], it["received"],
                         it["spend"], it["avg_price"], it["market_price"], it["price_vs_market"], it["avg_lead"]])
    _grid(wb, st, "Supplier-Item",
          ["Vendor", "Trade", "Item Code", "Item", "Lines", "Received", "Spend (INR)",
           "Avg price (INR)", "Market price", "Price vs mkt %", "Avg lead (d)"],
          rows, [34, 10, 16, 30, 7, 12, 14, 14, 13, 13, 11], {6, 7, 8, 9})
    return _save(wb)


def build_adhoc_workbook(rp, cycle="") -> bytes:
    st = _styles()
    s = rp.get("summary", {})
    wb = _book()
    _cover(wb, st, "Adhoc Planning — RM for open SOC not in projection", f"Cycle {cycle}",
           [("SOC items", s.get("soc_items"), _BAND), ("Adhoc items", s.get("adhoc_items"), _AMBER),
            ("Adhoc with BOM", s.get("adhoc_with_bom"), _BAND), ("Consolidated RMs", s.get("consolidated_rms"), _BAND),
            ("RMs to buy", s.get("rms_to_buy"), _RED), ("Total RM to buy (KG)", s.get("total_buy_qty"), _AMBER)])
    _grid(wb, st, "Adhoc Items",
          ["Item", "SOC qty", "# SOCs", "Adhoc?", "Has BOM", "RM to buy"],
          [[p["name"], p["soc_qty"], p["soc_count"], "Adhoc" if p["is_adhoc"] else "Projected",
            "Yes" if p["has_bom"] else "—", p["net_total"]] for p in rp.get("products", [])],
          [34, 13, 8, 10, 9, 13], {2, 6})
    _grid(wb, st, "Consolidated RM",
          ["RM Code", "RM Description", "#Items", "Gross (KG)", "Stock", "Sub stk", "Available", "Net to buy", "Used in items"],
          [[x["rm_code"], x["rm_desc"], x["item_count"], x["gross"], x["main_stock"], x["substitute_stock"],
            x["available"], x["net_to_buy"], ", ".join(x["items"])] for x in rp.get("consolidated_rm", [])],
          [16, 28, 8, 13, 12, 10, 12, 12, 50], {4, 5, 6, 7, 8})
    return _save(wb)


def build_ppv_workbook(rp, cycle="") -> bytes:
    from openpyxl.styles import PatternFill
    st = _styles()
    s = rp.get("summary", {})
    fav, unf = PatternFill("solid", fgColor=_GREEN), PatternFill("solid", fgColor=_RED)
    wb = _book()
    _cover(wb, st, f"Purchase Price Variance — JC-wise (std = FY{rp.get('std_fy','')} WAP)",
           f"Cycle {cycle} · {s.get('note','')[:90]}",
           [("Std items", s.get("std_items"), _BAND), ("Total spend ₹", s.get("total_spend"), _BAND),
            ("Timing overspend ₹", s.get("timing_overspend"), _RED), ("Overspend %", s.get("timing_overspend_pct"), _AMBER),
            ("Best JC", s.get("best_jc"), _GREEN), ("Worst JC", s.get("worst_jc"), _RED)])
    _grid(wb, st, "JC Performance",
          ["JC", "Qty", "Spend (₹)", "PPV (₹)", "PPV %", "Status"],
          [[j["jc"], j["qty"], j["spend"], j["ppv"], j["ppv_pct"], j["status"]] for j in rp.get("jc_performance", [])],
          [8, 16, 16, 16, 9, 14], {2, 3, 4}, {"favourable": fav, "unfavourable": unf}, 6)
    _grid(wb, st, "Items (volatility)",
          ["Item Code", "Item", "Std price (₹)", "Min", "Max", "Volatility %",
           "JCs above", "JCs below", "Timing overspend (₹)", "Worst JC", "Spend (₹)", "Qty"],
          [[x["code"], x["name"], x["std_price"], x["min_price"], x["max_price"], x["volatility_pct"],
            x["jcs_above"], x["jcs_below"], x["timing_overspend"], x["worst_jc"], x["spend"], x["qty"]]
           for x in rp.get("items", [])],
          [16, 30, 12, 10, 10, 11, 9, 9, 16, 9, 14, 12], {3, 4, 5, 9, 11, 12})
    return _save(wb)
