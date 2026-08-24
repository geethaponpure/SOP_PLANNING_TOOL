"""Business-Plan / projection loader (PC Business Plan JC-wise report export).

This is the output of SP_SCBusinessPlan_GetDetailedReportJCWise (which needs
EXECUTE rights the read login lacks). The report is per customer x item; we
aggregate to per-item annual budget and join to CRM items BY NAME (the report
carries ItemName, not Item Code -- ~99.8% match the CRM item_description).

Returns name(upper) -> {annual_budget, annual_potential, lms_potential,
jc_target} in KG/year. The adapter converts annual budget to a monthly
projection (÷12) to compare with the monthly statistical baseline.
"""
from __future__ import annotations

# header label -> logical field (labels confirmed from the export)
FIELDS = {
    "name": "ItemName",
    "customer": "CustomerNumber",
    "annual_budget": "AnnualBudgetQty (KG)",
    "annual_potential": "AnnualPotentialQty (KG)",
    "lms_potential": "LMSPotentialQTY (KG)",
    "jc_target": "JC4TargetQty (KG)",
    "accyear": "Accyear",
}


def _num(v) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def load_business_plan(path: str, sheet: str | None = None,
                       accyear: str | None = None,
                       classification: dict[str, str] | None = None) -> dict[str, dict]:
    """If ``classification`` (customer_number -> tier letter A/B/C/D/E) is given,
    also accumulates each item's budget by customer tier and a key-customer share."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    # prefer the Approved sheet if present (the locked plan), else the main report
    if sheet and sheet in wb.sheetnames:
        ws = wb[sheet]
    elif "Approved" in wb.sheetnames:
        ws = wb["Approved"]
    else:
        ws = wb.worksheets[0]

    rit = ws.iter_rows(values_only=True)
    ix = {}
    for r in rit:                       # find the header row
        labels = [str(x).strip() if x is not None else "" for x in r]
        if "ItemName" in labels and "AnnualBudgetQty (KG)" in labels:
            ix = {name: labels.index(col) for name, col in FIELDS.items() if col in labels}
            break

    agg: dict[str, dict] = {}
    for r in rit:
        nm = r[ix["name"]] if ix.get("name") is not None else None
        if not nm:
            continue
        if accyear and ix.get("accyear") is not None and str(r[ix["accyear"]]).strip() != accyear:
            continue
        key = str(nm).strip().upper()
        a = agg.setdefault(key, {"annual_budget": 0.0, "annual_potential": 0.0,
                                 "lms_potential": 0.0, "jc_target": 0.0, "rows": 0,
                                 "class_budget": {}})
        budget = _num(r[ix["annual_budget"]]) if ix.get("annual_budget") is not None else 0.0
        a["annual_budget"] += budget
        a["annual_potential"] += _num(r[ix["annual_potential"]]) if ix.get("annual_potential") is not None else 0.0
        a["lms_potential"] += _num(r[ix["lms_potential"]]) if ix.get("lms_potential") is not None else 0.0
        a["jc_target"] += _num(r[ix["jc_target"]]) if ix.get("jc_target") is not None else 0.0
        a["rows"] += 1
        if classification and ix.get("customer") is not None and budget > 0:
            cust = r[ix["customer"]]
            cls = classification.get(str(cust).strip()) if cust is not None else None
            if cls:
                a["class_budget"][cls] = a["class_budget"].get(cls, 0.0) + budget

    # derive a key-customer share (A+B class budget / total) and dominant tier
    for a in agg.values():
        cb = a["class_budget"]
        tot = sum(cb.values())
        a["key_customer_share"] = round((cb.get("A", 0) + cb.get("B", 0)) / tot, 3) if tot else 0.0
        a["top_tier"] = max(cb, key=cb.get) if cb else None
    return agg
