"""Planning Filtration Technique for the Supply & RM page (client spec).

Self-contained: reads the Excel/CSV exports + live CRM rows directly (no engine
dataset dependency). Provides the Supply & RM plan, Aged-RM -> FG plan,
Projection-vs-Sales, Supplier Scorecard, Adhoc planning and PPV builders.
"""
from __future__ import annotations

import csv
import hashlib
import os
import pickle
import re
from datetime import datetime, date
from functools import lru_cache

# --- constants ----------------------------------------------------------------
WAREHOUSE_ORGS = {
    "PCM - Madhavaram MFG", "PCM - Puzhal MFG", "POI - Alathur MFG",
    "PSM - Thervoykandigai MFG", "PCM - Illalor MFG", "PPM - Puzhal MFG",
    "PPM - Gummudipoondi MFG", "PPM - Illalor MFG", "POI - Hosur Mfg",
    "PPC - Bavla(MFG)", "PSM - Hosur Mfg", "PSM - Hosur Mfg (BM)",
}
RM_SOURCE_ORGS = {
    "PPC - Puzhal", "PCT - Puzhal Trading", "PCT - Illalor Trading",
    "PCT - Madhavaram Trading", "PST - Thervoykandigai", "PCM - Madhavaram MFG",
    "PSM - Thervoykandigai MFG", "PCM - Puzhal MFG", "POI - Alathur MFG",
    "PCM - Illalor MFG", "PPM - Puzhal MFG", "POI - Hosur Mfg", "PPM - Illalor MFG",
    "POI - Alathur", "POT - Hosur Trading", "PCC - Madhavaram Trading",
    "PST - Hosur Trading", "PSM - Hosur Mfg (BM)", "PSM - Hosur Mfg", "PCM- madhavaram",
}
EXCLUDE_SUBINV = {s.lower() for s in
                  ["NO SALE", "Return", "Clr/NonSal", "Scrap", "Expired", "Color",
                   "UNRECON", "Rework", "Rejected", "Re-Process"]}


def _norm(v) -> str:
    return re.sub(r"\s+", " ", str(v).replace("\xa0", " ")).strip() if v is not None else ""


def _num(v) -> float:
    try:
        return float(str(v).replace(",", ""))
    except (TypeError, ValueError):
        return 0.0


def _is_raw_material(business: str, rm_label: str = "Raw Material") -> bool:
    return _norm(business).lower() == _norm(rm_label).lower()


def _squash(s) -> str:
    return re.sub(r"[^A-Z0-9]", "", _norm(s).upper())


# An "encoded" item name = a short code like RDNBP101 / ASPP001 / POAP025 / LTLP005
# (letters then digits, no spaces) — vs a real chemical name (has spaces / no digits).
_ENC_NAME_RE = re.compile(r"^[A-Z]{2,6}[0-9]{2,4}[A-Z]?$")


def _is_bulk_hdlk(assembly_item: str) -> bool:
    a = (assembly_item or "").upper()
    return "BULK" in a or "HDLK" in a


# ── BOM activity classification (Manufacturing vs Repack/Relabel vs internal) ──
# Manufacturing  : BOM_TYPE=MFG (authoritative), OR blank/coded type on a bulk
#                  (BULK/HDLK) assembly with no packing component (un-tagged recipe).
# Repack/Relabel : BOM_TYPE in {REPACK, RELABLE}.
# Internal       : BOM_TYPE in {CONVERSION, DECODE} — encode/decode, not planned.
# Unclassified   : anything else (blank/coded non-bulk, or bulk-with-packing).
_REPACK_TYPES = {"REPACK", "RELABLE", "RELABEL"}
_INTERNAL_TYPES = {"CONVERSION", "DECODE"}


def _pack_code(code) -> bool:
    """Item code identifies PACKING material — it starts with 'P'. This is Pure
    Chemicals' item-numbering convention: packing (bottles/caps/labels/barrels/…)
    are coded P…, while chemical raw materials are coded M…/T…. The DESCRIPTION is
    deliberately NOT used, because chemical names (PURONAL, POLYDADMAC, PUREPRINT)
    also start with 'P'."""
    return (code or "").lstrip()[:1].upper() == "P"


def _is_packing_comp(c: dict) -> bool:
    """Packing material = COMPONENT_ITEM (item code) starts with 'P' (see _pack_code)."""
    return _pack_code(c.get("comp_code"))


def _has_packing_component(variant: dict) -> bool:
    """True if any component is packing material (item code starts with 'P')."""
    return any(_is_packing_comp(c) for c in variant.get("components", []))


def _has_qty1_component(variant: dict) -> bool:
    """True if ANY component (packing or non-packing) is consumed 1:1
    (COMPONENT_QUANTITY == 1) — the signature of a repack/relabel
    (one base unit / one pack in → one unit out)."""
    for c in variant.get("components", []):
        q = c.get("qty")
        if q is not None and abs(float(q) - 1.0) < 1e-9:
            return True
    return False


def _is_recipe_bom(variant: dict) -> bool:
    """True if the BOM is a formulation/mixing recipe: two or more NON-packing
    components, none consumed 1:1, at least one fractional (< 1). This is the
    manufacturing signature for coded/blank BOMs on non-BULK assembly codes
    (e.g. a 5-RM formula) that the MFG-type / BULK rules would otherwise miss."""
    qtys = []
    for c in variant.get("components", []):
        if _is_packing_comp(c):
            continue
        q = c.get("qty")
        if q is not None:
            qtys.append(float(q))
    if len(qtys) < 2:
        return False
    return all(abs(q - 1.0) > 1e-9 for q in qtys) and any(0 < q < 1 for q in qtys)


def classify_bom(variant: dict) -> str:
    """Return 'manufacturing' | 'repack_relabel' | 'internal' | 'unclassified'.
    (No-BOM finished goods are tagged 'trading' at the product level, not here.)"""
    bt = (variant.get("bom_type") or "").strip().upper()
    if bt in _INTERNAL_TYPES:
        return "internal"
    if bt == "MFG":
        return "manufacturing"
    if bt in _REPACK_TYPES:
        return "repack_relabel"
    # untagged (blank) or coded BOM_TYPE:
    #  · bulk recipe with no packing → manufacturing
    #  · a multi-RM formulation recipe (fractional qtys, no 1:1) → manufacturing
    if _is_bulk_hdlk(variant.get("assembly_item")) and not _has_packing_component(variant):
        return "manufacturing"
    if _is_recipe_bom(variant):
        return "manufacturing"
    # otherwise, any component used 1:1 (base or packing) → repack/relabel behaviour
    if _has_qty1_component(variant):
        return "repack_relabel"
    return "unclassified"


# --- 1. projection (3-JC): current = JCn WK1 + WK2 (n auto-detected) ----------
# The approved-projection file is compiled per JC — every column is prefixed with
# that JC (e.g. JC4WK1Qty / JC5WK1Qty). We detect the JC number from the header so
# the plan follows whichever approved projection is loaded (JC4, then JC5, …).
_JC_WK1_RE = re.compile(r"^JC(\d+)WK1Qty \(KG\)$")


def _find_projection_header(rit):
    for r in rit:
        labels = [_norm(x) for x in r]
        if "ItemName" not in labels:
            continue
        for lab in labels:
            m = _JC_WK1_RE.match(lab)
            if m:
                return {lab2: i for i, lab2 in enumerate(labels)}, int(m.group(1))
    return {}, None


def detect_projection_jc(path: str) -> int | None:
    """The JC number the approved-projection file is compiled for (from its columns)."""
    if not path:
        return None
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Approved"] if "Approved" in wb.sheetnames else wb.worksheets[0]
    _, jc_n = _find_projection_header(ws.iter_rows(values_only=True))
    wb.close()
    return jc_n


def load_projection_3jc(path: str, accyear: str | None = None,
                        drop_zero: bool = True) -> dict[str, dict]:
    import os as _os
    if not path or not _os.path.exists(path):
        return {}
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Approved"] if "Approved" in wb.sheetnames else wb.worksheets[0]
    rit = ws.iter_rows(values_only=True)
    ix, jc_n = _find_projection_header(rit)
    if not ix or jc_n is None:
        wb.close()
        return {}
    wk1, wk2 = ix.get(f"JC{jc_n}WK1Qty (KG)"), ix.get(f"JC{jc_n}WK2Qty (KG)")
    n1, n2 = ix.get(f"JC{jc_n}Next1JC (KG)"), ix.get(f"JC{jc_n}Next2JC (KG)")
    nm, ay = ix.get("ItemName"), ix.get("Accyear")
    s2, s3 = ix.get("Segment2"), ix.get("Segment3")
    out: dict[str, dict] = {}
    for r in rit:
        name = _norm(r[nm]) if nm is not None else ""
        if not name:
            continue
        if accyear and ay is not None and _norm(r[ay]) != accyear:
            continue
        a = out.setdefault(name.upper(), {"name": name, "current": 0.0, "next1": 0.0,
                                          "next2": 0.0, "segment2": "", "segment3": ""})
        a["current"] += (_num(r[wk1]) if wk1 is not None else 0.0) \
            + (_num(r[wk2]) if wk2 is not None else 0.0)
        a["next1"] += _num(r[n1]) if n1 is not None else 0.0
        a["next2"] += _num(r[n2]) if n2 is not None else 0.0
        if not a["segment2"] and s2 is not None:
            a["segment2"] = _norm(r[s2])
        if not a["segment3"] and s3 is not None:
            a["segment3"] = _norm(r[s3])
    if drop_zero:
        return {k: v for k, v in out.items()
                if (v["current"] + v["next1"] + v["next2"]) > 0}
    return out


def projection_from_crm(rows, drop_zero: bool = True) -> dict[str, dict]:
    """Convert CRM projection rows (crm_sources.business_plan_projection) into the
    same {NAME: {name,current,next1,next2,segment2,segment3}} shape as
    ``load_projection_3jc`` — so the RM planner is source-agnostic."""
    out: dict[str, dict] = {}
    for r in rows or []:
        name = _norm(r.get("ItemName"))
        if not name:
            continue
        a = out.setdefault(name.upper(), {"name": name, "current": 0.0, "next1": 0.0,
                                          "next2": 0.0, "segment2": "", "segment3": ""})
        a["current"] += _num(r.get("CurrentQ"))
        a["next1"] += _num(r.get("Next1Q"))
        a["next2"] += _num(r.get("Next2Q"))
        if not a["segment2"]:
            a["segment2"] = _norm(r.get("Segment2"))
        if not a["segment3"]:
            a["segment3"] = _norm(r.get("Segment3"))
    if drop_zero:
        return {k: v for k, v in out.items()
                if (v["current"] + v["next1"] + v["next2"]) > 0}
    return out


def projection_rows_from_crm(rows):
    """CRM projection rows (crm_sources.business_plan_projection_rows) → the same
    per-(item, collector) shape ``load_projection_rows`` yields."""
    for r in rows or []:
        name = _norm(r.get("ItemName"))
        if not name:
            continue
        yield {"name": name, "collector": _norm(r.get("Collector")) or "—",
               "segment2": _norm(r.get("Segment2")), "segment3": _norm(r.get("Segment3")),
               "current": _num(r.get("CurrentQ")), "next1": _num(r.get("Next1Q")),
               "next2": _num(r.get("Next2Q"))}


def load_projection_rows(path: str, accyear: str | None = None):
    """Per-row projection KEEPING collector + segments (for Projection vs Sales)."""
    import os as _os
    if not path or not _os.path.exists(path):
        return
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Approved"] if "Approved" in wb.sheetnames else wb.worksheets[0]
    rit = ws.iter_rows(values_only=True)
    ix, jc_n = _find_projection_header(rit)
    if not ix or jc_n is None:
        wb.close()
        return
    wk1, wk2 = ix.get(f"JC{jc_n}WK1Qty (KG)"), ix.get(f"JC{jc_n}WK2Qty (KG)")
    n1, n2 = ix.get(f"JC{jc_n}Next1JC (KG)"), ix.get(f"JC{jc_n}Next2JC (KG)")
    nm, ay = ix.get("ItemName"), ix.get("Accyear")
    cl, s2, s3 = ix.get("CollectorName"), ix.get("Segment2"), ix.get("Segment3")
    for r in rit:
        name = _norm(r[nm]) if nm is not None else ""
        if not name:
            continue
        if accyear and ay is not None and _norm(r[ay]) != accyear:
            continue
        cur = (_num(r[wk1]) if wk1 is not None else 0.0) + (_num(r[wk2]) if wk2 is not None else 0.0)
        yield {
            "name": name, "collector": (_norm(r[cl]) if cl is not None else "") or "—",
            "segment2": _norm(r[s2]) if s2 is not None else "",
            "segment3": _norm(r[s3]) if s3 is not None else "",
            "current": cur, "next1": _num(r[n1]) if n1 is not None else 0.0,
            "next2": _num(r[n2]) if n2 is not None else 0.0,
        }


# --- 2. BOM (detailed, with selection preference + substitutes) ---------------
@lru_cache(maxsize=4)
def load_bom_detailed(path: str) -> dict[str, list]:
    """assembly_desc(upper) -> list of BOM variants. A variant = (assembly_item,
    org, designator) so distinct BOMs aren't merged. Each carries BOM_TYPE +
    BOM_CREATION_DATE and is tagged packing when its assembly is non-BULK/HDLK."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Output"] if "Output" in wb.sheetnames else wb.active
    rit = ws.iter_rows(values_only=True)
    h = list(next(rit))
    j = {n: i for i, n in enumerate(h)}

    def col(r, name, default=None):
        i = j.get(name)
        return r[i] if i is not None else default

    variants: dict[tuple, dict] = {}
    for r in rit:
        if col(r, "DISABLE_DATE"):
            continue
        ai = _norm(col(r, "ASSEMBLY_ITEM"))
        if not ai:
            continue
        org = _norm(col(r, "ORGANIZATION_CODE"))
        desig = _norm(col(r, "ALTERNATE_BOM_DESIGNATOR"))
        v = variants.setdefault((ai, org, desig), {
            "assembly_item": ai, "assembly_desc": _norm(col(r, "ASSEMBLY_DESC")),
            "org_code": org, "designator": desig, "bom_type": _norm(col(r, "BOM_TYPE")),
            "created": col(r, "BOM_CREATION_DATE"), "is_packing": not _is_bulk_hdlk(ai),
            "seqs": {},
        })
        seq = col(r, "COMPONENT_ITEM_SEQ")
        comp = _norm(col(r, "COMPONENT_ITEM"))
        sub = _norm(col(r, "SUBSTITUTE_ITEM"))
        entry = v["seqs"].setdefault(seq, {
            "seq": seq, "comp_code": comp, "comp_desc": _norm(col(r, "COMP_ITEM_DESC")),
            "qty": _num(col(r, "COMPONENT_QUANTITY")), "substitutes": [],
        })
        if sub:
            entry["substitutes"].append({"code": sub, "desc": _norm(col(r, "SUBSTITUTE_ITEM_DESC")),
                                         "qty": _num(col(r, "SUBSTITUTE_ITEM_QUANTITY"))})

    by_desc: dict[str, list] = {}
    by_squash: dict[str, list] = {}
    for v in variants.values():
        v["components"] = list(v["seqs"].values())
        v["created_iso"] = v["created"].isoformat()[:10] if hasattr(v["created"], "isoformat") else None
        del v["seqs"]
        v["bom_class"] = classify_bom(v)
        by_desc.setdefault(v["assembly_desc"].upper(), []).append(v)
        by_squash.setdefault(_squash(v["assembly_desc"]), []).append(v)
    return {"by_desc": by_desc, "by_squash": by_squash}


def select_bom(variants: list[dict], settings: dict | None = None) -> tuple[dict | None, list]:
    if not variants:
        return None, []
    s = settings or {}

    def score(v):
        ai = v["assembly_item"].upper()
        key = []
        if s.get("bom_prefer_pmo", True):
            key.append(0 if v["org_code"] == "PMO" else 1)
        if s.get("bom_prefer_bulk_hdlk", True):
            key.append(0 if ("BULK" in ai or "HDLK" in ai) else 1)
        if s.get("bom_prefer_creation_date", True):
            created = v.get("created")
            key.append(-created.timestamp() if hasattr(created, "timestamp") else float("inf"))
        if s.get("bom_prefer_primary", True):
            key.append(0 if v["designator"].lower() == "primary" else 1)
        return tuple(key)
    ordered = sorted(variants, key=score)
    return ordered[0], ordered[1:]


# --- 3. stock (FG warehouse/branch + RM availability) -------------------------
def _classify_stock(rows, settings=None, rm_codes=None, fg_codes=None, business_map=None) -> dict:
    s = settings or {}
    warehouse = set(s.get("warehouse_orgs", WAREHOUSE_ORGS))
    rm_source = set(s.get("rm_source_orgs", RM_SOURCE_ORGS))
    # For RMs whose Business is NOT Raw Material (e.g. GC1/GC2 intermediates) only
    # stock held at these orgs counts as available — trading/depot stock elsewhere is
    # excluded. Admin-configurable via `intermediate_stock_orgs` (default = the MFG
    # plants + selected trading orgs); falls back to the MFG plants in the warehouse set.
    inter_orgs = s.get("intermediate_stock_orgs")
    mfg_orgs = set(inter_orgs) if inter_orgs else {o for o in warehouse if "mfg" in o.lower()}
    excluded = {x.lower() for x in s.get("excluded_subinv", EXCLUDE_SUBINV)}
    rm_label = s.get("raw_material_business", "Raw Material")
    rm_codes = rm_codes or set()
    fg_codes = fg_codes or set()
    business_map = business_map or {}
    fg: dict[str, dict] = {}
    rm: dict[str, dict] = {}
    rm_business: set = set()   # codes classified Raw Material by the Business field
    by_desc: dict[str, dict] = {}   # on-hand summed by item DESCRIPTION (all codes)
    for row in rows:
        sub = _norm(row.get("subinv"))
        if sub.lower() in excluded:
            continue
        code = _norm(row.get("code"))
        if not code:
            continue
        qty = _num(row.get("qty"))
        if qty == 0:
            continue
        org = _norm(row.get("org"))
        name = _norm(row.get("name"))
        dk = _squash(name)
        if dk:
            d = by_desc.setdefault(dk, {"warehouse": 0.0, "branch": 0.0, "name": name})
            if org in warehouse:
                d["warehouse"] += qty
            else:
                d["branch"] += qty
        business = row.get("business")
        if business is None and code in business_map:
            business = business_map[code]
        biz_rm = business is not None and _is_raw_material(business, rm_label)
        if business is not None:
            is_rm = biz_rm
            is_fg = not is_rm
        else:
            is_rm = (code in rm_codes) or (org in rm_source)
            is_fg = code in fg_codes
        if is_rm or org in rm_source:
            a = rm.setdefault(code, {"name": name, "qty": 0.0, "orgs": {},
                                     "business": _norm(business) if business else ""})
            if not a.get("business") and business:
                a["business"] = _norm(business)
            # Non-RM business (GC1/GC2 …): count only MFG-plant stock as available.
            non_rm_biz = business is not None and _norm(business) != "" and not biz_rm
            if not (non_rm_biz and org not in mfg_orgs):
                a["qty"] += qty
                if org:                       # track which org the counted stock sits in
                    a["orgs"][org] = a["orgs"].get(org, 0.0) + qty
            if biz_rm:
                rm_business.add(code)
        if is_fg:
            f = fg.setdefault(code, {"name": name, "warehouse": 0.0, "branch": 0.0})
            if org in warehouse:
                f["warehouse"] += qty
            else:
                f["branch"] += qty
    return {"fg": fg, "rm": rm, "rm_business": rm_business, "fg_by_desc": by_desc}


def _stock_rows_from_xlsx(path: str):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rit = ws.iter_rows(values_only=True)
    ix = {n: i for i, n in enumerate(list(next(rit)))}
    for r in rit:
        yield {"subinv": r[ix["Sub Inv"]], "code": r[ix["Item Code"]],
               "qty": r[ix["Qty"]], "org": r[ix["Organization"]],
               "name": r[ix["Item Desc"]], "business": r[ix["Business"]]}


def load_stock_filtered(path: str, settings=None) -> dict:
    return _classify_stock(_stock_rows_from_xlsx(path), settings)


def load_stock_crm(rows, settings=None, rm_codes=None, fg_codes=None, business_map=None) -> dict:
    norm = ({"subinv": r.get("SubInv"), "code": r.get("ItemCode"),
             "qty": r.get("Qty"), "org": r.get("Organization"),
             "name": r.get("ItemDesc"), "business": None} for r in rows)
    return _classify_stock(norm, settings, rm_codes, fg_codes, business_map)


# --- 4. PO pending / in-transit ----------------------------------------------
def load_po_pending(path: str) -> dict[str, dict]:
    with open(path, encoding="utf-8-sig", errors="replace", newline="") as fh:
        rows = list(csv.reader(fh))
    hi = next((i for i, r in enumerate(rows[:10]) if "Item Code" in r and "Po Number" in r), 0)
    ix = {n: i for i, n in enumerate(rows[hi])}

    def c(row, name):
        i = ix.get(name)
        return row[i] if i is not None and i < len(row) else None

    grp: dict[tuple, dict] = {}
    for row in rows[hi + 1:]:
        code = _norm(c(row, "Item Code"))
        if not code:
            continue
        po = _norm(c(row, "Po Number"))
        g = grp.setdefault((po, code), {"po_qty": 0.0, "recd": 0.0, "name": _norm(c(row, "Item Description")),
                                        "podate": _parse_date(c(row, "Po Date"))})
        g["po_qty"] = max(g["po_qty"], _num(c(row, "Po Quantity")))
        g["recd"] += _num(c(row, "Receipt Qty"))
    out: dict[str, dict] = {}
    for (po, code), g in grp.items():
        pend = max(0.0, g["po_qty"] - g["recd"])
        a = out.setdefault(code, {"name": g["name"], "in_transit": 0.0, "received": 0.0,
                                  "open_pos": 0, "pend_dates": []})
        a["received"] += g["recd"]
        if pend > 0:
            a["in_transit"] += pend
            a["open_pos"] += 1
            if g["podate"]:
                a["pend_dates"].append(g["podate"])
    return out


def load_po_pending_rows(rows) -> dict[str, dict]:
    """Pending / in-transit per item from pre-parsed PO rows (handles the .xls HTML
    register via read_po_rows). Same output shape as load_po_pending."""
    grp: dict[tuple, dict] = {}
    for r in rows or []:
        code = _norm(r.get("Item Code"))
        if not code:
            continue
        po = _norm(r.get("Po Number"))
        g = grp.setdefault((po, code), {"po_qty": 0.0, "recd": 0.0,
                                        "name": _norm(r.get("Item Description")),
                                        "podate": _parse_date(r.get("Po Date"))})
        g["po_qty"] = max(g["po_qty"], _num(r.get("Po Quantity")))
        g["recd"] += _num(r.get("Receipt Qty"))
    out: dict[str, dict] = {}
    for (po, code), g in grp.items():
        pend = max(0.0, g["po_qty"] - g["recd"])
        a = out.setdefault(code, {"name": g["name"], "in_transit": 0.0, "received": 0.0,
                                  "open_pos": 0, "pend_dates": []})
        a["received"] += g["recd"]
        if pend > 0:
            a["in_transit"] += pend
            a["open_pos"] += 1
            if g["podate"]:
                a["pend_dates"].append(g["podate"])
    return out


def load_po_intransit_crm(rows, settings=None) -> dict[str, dict]:
    """Shape live CRM open-PO rows (from crm_sources.po_open_intransit) into the same
    per-item-code dict as load_po_pending_rows: {code: {name, in_transit, received,
    open_pos}}. Inter-company (group-company) vendors are dropped, same as the file path."""
    exclude = _intercompany_set(settings)
    out: dict[str, dict] = {}
    for r in rows or []:
        if _norm(r.get("Vendor_Name")).upper() in exclude:
            continue
        code = _norm(r.get("Item_Code"))
        if not code:
            continue
        a = out.setdefault(code, {"name": _norm(r.get("Item_Desc")), "in_transit": 0.0,
                                  "received": 0.0, "open_pos": 0, "pend_dates": []})
        if not a["name"]:
            a["name"] = _norm(r.get("Item_Desc"))
        a["in_transit"] += _num(r.get("InTransit"))
        a["received"] += _num(r.get("Received"))
        a["open_pos"] += int(_num(r.get("OpenLines")))
    return out


# --- 4b. PO intelligence: lead time, supplier, import/domestic, currency ------
def _parse_date(s):
    if s is None or s == "":
        return None
    if isinstance(s, datetime):
        return s.date()
    if isinstance(s, date):
        return s
    s = _norm(s)
    for fmt in ("%d-%b-%y", "%d-%b-%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


_PO_COLS = ("Item Code", "Item Description", "Po Date", "Receipt Date", "Vendor Name",
            "Vendor State", "Vendor Site", "Currency Code", "Po Number", "Receipt Qty",
            "Po Quantity", "Po Unit Price", "Rcpt Amt", "Vendor No", "Currency Rate",
            "Division", "Product", "Category",
            "Receipt No.", "Org Name", "Subinventory", "Lot Number")


_HTML_CELL_RE = re.compile(r"<t[dh][^>]*>(.*?)</t[dh]>", re.S | re.I)
_HTML_ROW_RE = re.compile(r"<tr[^>]*>(.*?)</tr>", re.S | re.I)
_HTML_TAG_RE = re.compile(r"<[^>]+>")


def _iter_html_po_rows(path: str):
    """Parse an Oracle BI Publisher HTML table exported with an .xls/.htm
    extension (one <tr> per receipt; the first row containing 'Item Code' is the
    header). Streams rows so the 70 MB+ export never fully materialises as cells."""
    import html as _html
    with open(path, encoding="utf-8", errors="replace") as fh:
        content = fh.read()
    ix = None
    for m in _HTML_ROW_RE.finditer(content):
        vals = [_html.unescape(_HTML_TAG_RE.sub("", c)).strip()
                for c in _HTML_CELL_RE.findall(m.group(1))]
        is_header = "Item Code" in vals and "Receipt Qty" in vals
        if ix is None:
            if is_header:
                ix = {n: i for i, n in enumerate(vals)}
            continue
        if is_header:          # repeated header on a later page -- skip
            continue
        yield {col: (vals[ix[col]] if col in ix and ix[col] < len(vals) else None)
               for col in _PO_COLS}


def _iter_po_rows(path: str):
    low = path.lower()
    if low.endswith(".xlsx"):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rit = ws.iter_rows(values_only=True)
        ix = None
        for r in rit:
            vals = [_norm(x) for x in r]
            if "Item Code" in vals:
                ix = {n: i for i, n in enumerate(vals)}
                break
        if ix:
            for r in rit:
                yield {col: (r[ix[col]] if col in ix and ix[col] < len(r) else None)
                       for col in _PO_COLS}
        wb.close()
    elif low.endswith((".xls", ".htm", ".html")):
        yield from _iter_html_po_rows(path)
    else:
        with open(path, encoding="utf-8-sig", errors="replace", newline="") as fh:
            reader = csv.reader(fh)
            ix = None
            for r in reader:
                if "Item Code" in r:
                    ix = {n: i for i, n in enumerate(r)}
                    break
            if ix:
                for r in reader:
                    yield {col: (r[ix[col]] if col in ix and ix[col] < len(r) else None)
                           for col in _PO_COLS}


# Inter-company / related-party vendors -- their PO receipts are internal stock
# transfers, not third-party purchases, so they are excluded from all PO
# analytics (lead time, supplier scorecard, PPV).
# Default group-company vendors (used if settings.intercompany_vendors is absent).
DEFAULT_INTERCOMPANY_VENDORS = [
    "PON PURE CHEMICAL INDIA PRIVATE LIMITED",
    "PURE CHEMICALS CO MFG",
    "COLOR CHEMICALS",
    "PON PURE SPECIALITY CHEMICAL PRIVATE LIMITED",
    "PURE CHEMICALS CO",
    "COLOR CHEMICALS AND DYES LLP",
    "PURE ORGANIC INDUSTRIES TRADING",
    "PON PURE SPECIALITY CHEMICAL PRIVATE LIMITED-MFG",
    "COLOR CHEMICALS AND DYES PRIVATE LIMITED",
    "PURE ORGANIC INDUSTRIES",
    "PURE CHEMICALS PTE LTD",
]
INTERCOMPANY_VENDORS = {re.sub(r"\s+", " ", v).strip().upper() for v in DEFAULT_INTERCOMPANY_VENDORS}


def _intercompany_set(settings=None) -> set:
    """Normalized group-company vendor names to exclude from PO analytics — from the
    admin-configurable `intercompany_vendors` setting, else the module default."""
    if settings is None:
        try:
            from . import planning_settings
            settings = planning_settings.load()
        except Exception:   # noqa: BLE001
            settings = {}
    lst = settings.get("intercompany_vendors")
    src = lst if lst is not None else DEFAULT_INTERCOMPANY_VENDORS
    return {re.sub(r"\s+", " ", str(v)).strip().upper() for v in src if v}


def _read_po_rows_raw(paths) -> list[dict]:
    if isinstance(paths, str):
        paths = [paths]
    paths = [p for p in paths if p]
    cache_file = None
    try:
        cols_ver = hashlib.md5((",".join(_PO_COLS) + "|dedup=v2-lot-subinv-qty").encode()).hexdigest()[:8]
        key = hashlib.md5(("|".join(f"{p}:{os.path.getmtime(p)}" for p in paths)
                           + f"|cols={cols_ver}").encode()).hexdigest()
        cdir = os.path.join(os.path.dirname(os.path.dirname(__file__)), ".cache")
        os.makedirs(cdir, exist_ok=True)
        cache_file = os.path.join(cdir, f"po_rows_{key}.pkl")
        if os.path.exists(cache_file):
            with open(cache_file, "rb") as f:
                return pickle.load(f)
    except OSError:
        cache_file = None
    out: list[dict] = []
    for path in paths:
        out.extend(_iter_po_rows(path))
    # de-duplicate receipt lines across overlapping periodic downloads. A single
    # Receipt No. can legitimately hold SEVERAL lines (a receipt split by lot /
    # sub-inventory / quantity — e.g. GRN 2373 for one PO booked as 1,294 + 521,128),
    # so the key must include lot, sub-inventory and receipt qty; otherwise real
    # split receipts collapse into one and their received qty is lost (inflating
    # in-transit). Rows without a Receipt No. are kept as-is.
    seen: set = set()
    deduped: list[dict] = []
    for r in out:
        rn = _norm(r.get("Receipt No."))
        if rn:
            k = (rn, _norm(r.get("Po Number")), _norm(r.get("Item Code")),
                 _norm(r.get("Lot Number")), _norm(r.get("Subinventory")),
                 _num(r.get("Receipt Qty")))
            if k in seen:
                continue
            seen.add(k)
        deduped.append(r)
    out = deduped
    if cache_file:
        try:
            with open(cache_file, "wb") as f:
                pickle.dump(out, f, protocol=pickle.HIGHEST_PROTOCOL)
        except OSError:
            pass
    return out


def read_po_rows(paths, settings=None) -> list[dict]:
    """Parse all PO-receipt rows once (cached to disk by file mtime), then drop
    inter-company (group-company) vendors so only third-party purchases reach the
    analytics. The excluded vendor list is admin-configurable (settings
    `intercompany_vendors`); vendor filtering is applied AFTER the row cache, so
    editing the list takes effect without rebuilding the cache."""
    raw = _read_po_rows_raw(paths)
    exclude = _intercompany_set(settings)
    if not exclude:
        return raw
    return [r for r in raw if _norm(r.get("Vendor Name")).upper() not in exclude]


def load_po_intel(paths, rows=None) -> dict[str, dict]:
    """Per RM: avg lead time over the latest 5 purchases, suppliers, locations,
    import vs domestic (non-INR currency) and currency codes."""
    agg: dict[str, dict] = {}
    for row in (rows if rows is not None else read_po_rows(paths)):
        if True:
            code = _norm(row.get("Item Code"))
            if not code:
                continue
            a = agg.setdefault(code, {"name": _norm(row.get("Item Description")), "pos": {},
                                      "suppliers": set(), "locations": set(),
                                      "currencies": set(), "receipts": 0})
            pod, rcd = _parse_date(row.get("Po Date")), _parse_date(row.get("Receipt Date"))
            if pod and rcd and 0 <= (rcd - pod).days <= 730:
                pono = _norm(row.get("Po Number")) or f"_row{len(a['pos'])}"
                p = a["pos"].setdefault(pono, {"po_date": pod, "leads": []})
                p["po_date"] = max(p["po_date"], pod)
                p["leads"].append((rcd - pod).days)
            a["receipts"] += 1
            v = _norm(row.get("Vendor Name"))
            if v:
                a["suppliers"].add(v)
            loc = _norm(row.get("Vendor State")) or _norm(row.get("Vendor Site"))
            if loc:
                a["locations"].add(loc)
            cur = _norm(row.get("Currency Code")).upper()
            if cur:
                a["currencies"].add(cur)

    out: dict[str, dict] = {}
    for code, a in agg.items():
        pos = sorted(a["pos"].values(), key=lambda p: p["po_date"], reverse=True)[:5]
        lead = round(sum(sum(p["leads"]) / len(p["leads"]) for p in pos) / len(pos), 1) if pos else None
        is_import = bool(a["currencies"] - {"INR", ""})
        out[code] = {
            "name": a["name"], "avg_lead_time_days": lead, "receipts": a["receipts"],
            "lead_basis_pos": len(pos),
            "suppliers": sorted(a["suppliers"])[:10], "supplier_count": len(a["suppliers"]),
            "locations": sorted(a["locations"])[:10], "currencies": sorted(a["currencies"]),
            "trade": "Import" if is_import else "Domestic",
        }
    return out


# --- 4c. dispatched trailing average across JCs -------------------------------
def aggregate_dispatch(rows, n_jc: int = 3) -> dict:
    by_name: dict[str, dict] = {}
    by_ci: dict[tuple, dict] = {}
    for r in rows or []:
        name = _norm(r.get("ItemName"))
        if not name:
            continue
        key = _squash(name)
        jc = [_num(r.get(f"jc{i}")) for i in range(n_jc)]
        a = by_name.setdefault(key, {"name": name, "jc": [0.0] * n_jc})
        for i in range(n_jc):
            a["jc"][i] += jc[i]
        coll = _norm(r.get("Collector")) or "—"
        b = by_ci.setdefault((coll, key), {"collector": coll, "name": name, "jc": [0.0] * n_jc})
        for i in range(n_jc):
            b["jc"][i] += jc[i]
    for a in by_name.values():
        a["avg"] = round(sum(a["jc"]) / n_jc, 1)
    for b in by_ci.values():
        b["avg"] = round(sum(b["jc"]) / n_jc, 1)
    return {"by_name": by_name, "by_collector_item": by_ci, "n_jc": n_jc}


def _proj_flag(projection: float, avg_sales: float, band: float = 0.2) -> str:
    if avg_sales <= 0:
        return "new" if projection > 0 else "none"
    if projection > avg_sales * (1 + band):
        return "over"
    if projection < avg_sales * (1 - band):
        return "under"
    return "ontrack"


# --- 5. assemble the Supply & RM page data -----------------------------------
def _net_3jc(gross: dict, available: float) -> dict:
    left = available
    net = {}
    for k in ("current", "next1", "next2"):
        n = max(0.0, gross[k] - left)
        left = max(0.0, left - gross[k])
        net[k] = round(n, 1)
    return net


def _lead_jcs(avg_days, preproc: float = 7.0) -> tuple:
    """JC buckets to plan for an RM by its TOTAL lead time (avg + preprocessing):
    ≤30d → Current; 31–60d → Current + Next1; >60d → all 3. Unknown lead → all 3."""
    if avg_days is None:
        return ("current", "next1", "next2")
    total = avg_days + preproc
    if total <= 30:
        return ("current",)
    if total <= 60:
        return ("current", "next1")
    return ("current", "next1", "next2")


def _producible_from_rm(bom) -> float:
    """FG qty producible from currently-available RAW MATERIAL (chemical RMs only —
    PACKING material is excluded, see _pack_code). Bottleneck = scarcest RM: min over
    non-packing components of available / qty_per_unit, where available = main stock +
    substitutes + in-transit. A required RM with zero availability caps output at 0."""
    if not bom or not bom.get("components"):
        return 0.0
    cap = None
    for c in bom["components"]:
        if _pack_code(c.get("rm_code")):          # packing material — not an RM constraint
            continue
        q = c.get("qty_per_unit") or 0.0
        if q <= 0:
            continue
        units = (c.get("available") or 0.0) / q
        cap = units if cap is None else min(cap, units)
    return round(cap or 0.0, 1)


def _components_for(variant, proj, rm_stock, po, consider_subs=True,
                    po_intel=None, preproc: float = 7.0, sub_exclude=None, lead_fn=None) -> list[dict]:
    rows = []
    dm_exclude = sub_exclude or set()

    def _sub_ok(code):
        # drop substitutes that are data-entry errors: packing material (code starts
        # 'P') or DM water (admin-listed codes) — they can't substitute a raw material.
        c = _norm(code).upper()
        return c[:1] != "P" and c not in dm_exclude

    for comp in variant["components"]:
        gross = {k: round(proj[k] * comp["qty"], 1) for k in ("current", "next1", "next2")}
        subs = [{"code": s["code"], "desc": s["desc"],
                 "stock": round(rm_stock.get(s["code"], {}).get("qty", 0.0), 1)}
                for s in comp["substitutes"] if _sub_ok(s["code"])]
        main_stock = rm_stock.get(comp["comp_code"], {}).get("qty", 0.0)
        sub_stock = sum(s["stock"] for s in subs) if consider_subs else 0.0
        po_item = po.get(comp["comp_code"], {})
        in_transit = po_item.get("in_transit", 0.0)
        received = po_item.get("received", 0.0)
        available = main_stock + sub_stock + in_transit
        net = _net_3jc(gross, available)
        if lead_fn is not None:                    # decode-aware lead (fills encoded items)
            avg_lead, _ = lead_fn([comp["comp_code"]], comp["comp_desc"])
        else:
            avg_lead = (po_intel or {}).get(comp["comp_code"], {}).get("avg_lead_time_days")
        planned = _lead_jcs(avg_lead, preproc)
        rows.append({
            "seq": comp["seq"], "rm_code": comp["comp_code"], "rm_desc": comp["comp_desc"],
            "qty_per_unit": comp["qty"], "gross": gross,
            "gross_total": round(gross["current"] + gross["next1"] + gross["next2"], 1),
            "main_stock": round(main_stock, 1), "substitute_stock": round(sub_stock, 1),
            "received": round(received, 1), "in_transit": round(in_transit, 1),
            "available": round(available, 1), "net_to_buy": net,
            # net_total buys only the JC buckets the lead time allows planning for
            "net_total": round(sum(net[k] for k in planned), 1),
            "planned_jcs": list(planned),
            # actionable plan: buy only where there's a real shortfall in the lead horizon.
            # available (no shortfall) -> to_buy False -> no purchase to plan.
            "buy_jcs": [k for k in planned if net[k] > 0],
            "to_buy": any(net[k] > 0 for k in planned),
            "lead_total_days": round(avg_lead + preproc, 1) if avg_lead is not None else None,
            "substitutes": subs,
        })
    return rows


def parse_plan_template(data: bytes) -> list[dict]:
    """Parse a filled plan-input template (S.No, Item Description, Current JC Qty
    (Kg), Next JC1 Qty (Kg), Next JC2 Qty (Kg)) into overrides
    [{name, current, next1, next2}]. Old headers (Current JC / Qty (Kg)) still work."""
    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = wb["Template"] if "Template" in wb.sheetnames else wb.active
    rit = ws.iter_rows(values_only=True)
    header = None
    for r in rit:
        vals = [_norm(x) for x in r]
        if "Item Description" in vals:
            header = {v: i for i, v in enumerate(vals)}
            break
    if not header:
        return []

    def cell(row, *names):
        for nm in names:
            i = header.get(nm)
            if i is not None and i < len(row):
                return row[i]
        return None

    out = []
    for r in rit:
        name = _norm(cell(r, "Item Description"))
        if not name:
            continue
        cur = _num(cell(r, "Current JC Qty (Kg)", "Current JC", "Qty (Kg)"))
        n1 = _num(cell(r, "Next JC1 Qty (Kg)", "Next JC1"))
        n2 = _num(cell(r, "Next JC2 Qty (Kg)", "Next JC2"))
        if cur <= 0 and n1 <= 0 and n2 <= 0:
            continue
        out.append({"name": name, "current": round(cur, 1), "next1": round(n1, 1), "next2": round(n2, 1)})
    return out


def _variant_key(v) -> str:
    return f"{v['assembly_item']}|{v['org_code']}|{v['designator']}"


def build_rm_planning(bp_path, bom_path, stock_path, po_path, accyear=None, settings=None,
                      stock_rows=None, pending_rows=None, soc_window=None, business_map=None,
                      dispatch_avg=None, po_intel=None, overrides=None, plan_mode="crm",
                      projection=None, bom_overrides=None, po_pending=None,
                      mfg_pending_rows=None, pto_map=None, msl_map=None,
                      division_map=None, plan_divisions=None) -> dict:
    from . import planning_settings
    s = settings or planning_settings.load()
    max_products = int(s.get("max_products", 400))
    consider_subs = s.get("consider_substitutes", True)
    preproc = float(s.get("preprocessing_days", 7))
    # substitutes to drop (data-entry errors): DM water codes (+ packing, handled in
    # _components_for by the 'P' code prefix)
    sub_exclude = {_norm(c).upper() for c in s.get("dm_water_codes", []) if c}

    def _is_dm_water(code, desc):
        """DM / demineralised water — excluded from the RM planning sheets (never bought)."""
        d = _norm(desc).upper()
        return (_norm(code).upper() in sub_exclude
                or "DM WATER" in d or "DEMINERALIZED" in d or "DEMINERALISED" in d)
    jc_on = {"current": s.get("plan_current", True), "next1": s.get("plan_next1", True),
             "next2": s.get("plan_next2", True)}
    drop_zero = s.get("drop_all_zero_projection", True)

    # projection source: CRM (SP replication, passed in) takes precedence; else the
    # legacy PCBusinessPlan Excel export.
    if projection is not None:
        proj = dict(projection)
    else:
        proj = load_projection_3jc(bp_path, accyear, drop_zero=False) if bp_path else {}

    # user-uploaded Excel plan: 'excel_only' plans purely from the Excel (drop CRM
    # projection); 'consolidate' overrides/adds items from the Excel on top of the
    # CRM projection. Overridden items use the Excel qty as the final target.
    overrides = overrides or []
    excel_keys: set = set()
    if plan_mode == "excel_only":
        proj = {}
    if overrides:
        sq_to_key = {_squash(v["name"]): k for k, v in proj.items()}
        for ov in overrides:
            sk = _squash(ov["name"])
            excel_keys.add(sk)
            k = sq_to_key.get(sk) or ov["name"].upper()
            v = proj.get(k)
            if v is None:
                v = proj[k] = {"name": ov["name"], "current": 0.0, "next1": 0.0, "next2": 0.0,
                               "segment2": "", "segment3": ""}
                sq_to_key[sk] = k
            v["current"], v["next1"], v["next2"] = ov["current"], ov["next1"], ov["next2"]
    bom_idx = load_bom_detailed(bom_path) if bom_path else {"by_desc": {}, "by_squash": {}}
    bom_by_desc, bom_by_squash = bom_idx["by_desc"], bom_idx["by_squash"]

    bom_fg_codes, bom_rm_codes = set(), set()
    for variants in bom_by_desc.values():
        for v in variants:
            bom_fg_codes.add(v["assembly_item"])
            for comp in v["components"]:
                bom_rm_codes.add(comp["comp_code"])
                for su in comp["substitutes"]:
                    bom_rm_codes.add(su["code"])

    if stock_rows is not None:
        stock = load_stock_crm(stock_rows, s, bom_rm_codes, bom_fg_codes, business_map)
    elif stock_path:
        stock = load_stock_filtered(stock_path, s)
    else:
        stock = {"fg": {}, "rm": {}}
    po = po_pending if po_pending is not None else (load_po_pending(po_path) if po_path else {})
    fg_stock, rm_stock = stock["fg"], stock["rm"]
    fg_by_desc = stock.get("fg_by_desc", {})
    rm_business = stock.get("rm_business", set())
    # date span of the PENDING POs (same source as the pending-items count)
    _pend_dates = [d for v in po.values() for d in v.get("pend_dates", [])]
    po_window = {"from": min(_pend_dates).isoformat(), "to": max(_pend_dates).isoformat()} \
        if _pend_dates else None

    def _soc_by_desc(rows):
        out: dict[str, float] = {}
        tot = 0.0
        for r in (rows or []):
            q = _num(r.get("PendingQty"))
            if q <= 0:
                continue
            k = _squash(r.get("ItemDesc"))
            out[k] = out.get(k, 0.0) + q
            tot += q
        return out, tot

    overall_by_sq, overall_total = _soc_by_desc(pending_rows)   # all orgs (info)
    mfg_by_sq, mfg_total = _soc_by_desc(mfg_pending_rows)        # planning orgs -> Current
    pend_by_sq, pending_total = overall_by_sq, overall_total     # legacy alias

    # ── Division (Segment1) scope ────────────────────────────────────────────────
    # The plan can be restricted to one or more divisions (e.g. only "Performance
    # Chemicals"). Disabled — everything passes — when no filter or no division map is
    # supplied, so the plan never silently empties if the CRM lookup is unavailable.
    _allowed_div = {str(d).strip() for d in (plan_divisions or []) if str(d).strip()}
    division_map = division_map or {}

    def _in_scope(name):
        if not _allowed_div or not division_map:
            return True
        return division_map.get(_squash(name)) in _allowed_div

    for v in proj.values():
        # Excel-overridden items (and excel_only mode) use the Excel qty as-is —
        # SOC is not added again; other items add MFG SOC pending (planning orgs).
        if plan_mode == "excel_only" or _squash(v["name"]) in excel_keys:
            v["mfg_soc"] = 0.0
            v["overall_soc"] = 0.0
        else:
            v["mfg_soc"] = round(mfg_by_sq.get(_squash(v["name"]), 0.0), 1)
            v["overall_soc"] = round(overall_by_sq.get(_squash(v["name"]), 0.0), 1)
        v["pending_soc"] = v["mfg_soc"]   # legacy alias: Current uses MFG SOC

    def enabled_sum(v):
        base = sum(v[k] for k in ("next1", "next2") if jc_on[k])
        if jc_on["current"]:
            base += v["current"] + v.get("pending_soc", 0.0)
        return base

    # Negligible-quantity filter: plan an item only if a JC projection (any of the 3)
    # OR the pending SOC is more than `min_plan_qty` (default 25 KG). e.g. projection 0
    # but SOC 40 -> still planned; all JCs and SOC <= 25 -> dropped as negligible.
    min_qty = float(s.get("min_plan_qty", 25))

    def _qualifies(v):
        proj_ok = any(v[k] > min_qty for k in ("current", "next1", "next2") if jc_on[k])
        return proj_ok or v.get("pending_soc", 0.0) > min_qty

    # ── MSL-only items ────────────────────────────────────────────────────────────
    # Valid MSL products (freq > 10 AND unique customers > 5) that would NOT otherwise be
    # planned this JC — either absent from the projection, OR present with only a
    # negligible (below min_plan_qty) projection — are still replenished to their MSL,
    # but only when:
    #   (a) they are in the plan's division scope (Segment1),
    #   (b) they carry a BOM (Manufacturing / Repack; pure Trading has no recipe → skip),
    #   (c) on-hand FG (Warehouse + Branch) is BELOW their MSL (a real shortfall).
    # Demand is (projection + pending SOC + MSL) − on-hand; for a pure MSL-only item the
    # projection part is ~0, so it reduces to (MSL − on-hand). Flagged msl_only so the
    # negligible-qty / division / cap filters don't drop it.
    # In 'excel_only' mode the plan must show ONLY the products in the uploaded file, so
    # no MSL-only top-up items are injected (they would add products the user didn't attach).
    msl_only_keys: set = set()
    if msl_map and jc_on["current"] and plan_mode != "excel_only":
        _by_sq = {_squash(v["name"]): v for v in proj.values()}
        for _sq, _mslv in msl_map.items():
            if _mslv <= 0:
                continue
            v0 = _by_sq.get(_sq)
            if v0 is not None and _qualifies(v0):
                continue                                       # already planned normally
            if _allowed_div and division_map and division_map.get(_sq) not in _allowed_div:
                continue                                       # out of division scope
            _vars = bom_by_squash.get(_sq)                     # has a recipe? (Mfg/Repack)
            if not _vars:
                continue
            _fg = fg_by_desc.get(_sq, {})
            if (_fg.get("warehouse", 0.0) + _fg.get("branch", 0.0)) >= _mslv:
                continue                                       # already covered → skip
            if v0 is None:                                     # not in projection → inject
                _nm = _vars[0].get("assembly_desc") or _sq
                _mfgsoc = round(mfg_by_sq.get(_sq, 0.0), 1)
                proj[_nm.upper()] = {"name": _nm, "current": 0.0, "next1": 0.0, "next2": 0.0,
                                     "segment2": _vars[0].get("segment2", ""), "segment3": "",
                                     "mfg_soc": _mfgsoc, "pending_soc": _mfgsoc,
                                     "overall_soc": round(overall_by_sq.get(_sq, 0.0), 1)}
            msl_only_keys.add(_sq)

    # Excel-uploaded items (and injected MSL-only items) sort first and are NEVER
    # dropped by the max_products cap, so they always appear in the plan.
    def _priority(v):
        sq = _squash(v["name"])
        return sq not in excel_keys and sq not in msl_only_keys
    items = sorted(proj.values(),
                   key=lambda v: (_priority(v), -enabled_sum(v)))
    # ── Real-RM explosion helpers (hoisted so the products loop's _components_for
    #    can use the decode-aware lead-time fallback) ─────────────────────────────
    # Some "RMs" are really encoded INTERMEDIATES with their own BOM (e.g. RDNBP101).
    # For the true purchasing list we recursively explode any component that is itself
    # an assembly down to its leaf (purchased) raw materials. A component is an
    # intermediate iff its code is a known ASSEMBLY_ITEM (or its desc a known assembly).
    asm_by_code: dict[str, list] = {}
    for _vl in bom_by_squash.values():
        for _v in _vl:
            asm_by_code.setdefault(_v["assembly_item"], []).append(_v)
    _sel_cache: dict[str, dict] = {}
    # A BOM is only a real recipe worth exploding if it's MANUFACTURING/REPACK AND not
    # a code ALIAS. 'internal' BOMs (DECODE/CONVERSION) and 1:1 single-component aliases
    # (e.g. ASPP001 = 1× P.T.S ACID, no packing — a rename, and P.T.S ACID aliases back)
    # are NOT intermediates; such items are treated as leaf RMs (not exploded/flagged).
    _REAL_BOM = {"manufacturing", "repack_relabel"}

    def _is_alias_bom(v):
        comps = v.get("components", [])
        return (len(comps) == 1 and not _is_packing_comp(comps[0])
                and abs((comps[0].get("qty") or 0.0) - 1.0) < 1e-9)

    def _is_real_recipe(v):
        return v.get("bom_class") in _REAL_BOM and not _is_alias_bom(v)

    def _sel_variant(code, desc):
        # Match an intermediate to explode. By CODE: any real recipe (mfg or repack).
        # By DESCRIPTION (fallback, when the component code isn't the assembly's own
        # code): ONLY a manufacturing recipe — never a repack. A repack SKU shares its
        # DESCRIPTION with its own bulk base (packed "FLEXOL WT 500" = 1× bulk FLEXOL
        # WT 500 + barrel); matching the bulk to that repack by desc would loop the bulk
        # back into itself and wrongly flag a purchased RM as an unresolved intermediate.
        by_code = [v for v in asm_by_code.get(code, []) if _is_real_recipe(v)]
        if by_code:
            vlist = by_code
        else:
            vlist = [v for v in bom_by_squash.get(_squash(desc), [])
                     if v.get("bom_class") == "manufacturing" and not _is_alias_bom(v)]
        if not vlist:
            return None                      # leaf RM (no recipe / alias / decode)
        ck = code or _squash(desc)
        if ck not in _sel_cache:
            # Prefer the real MANUFACTURING recipe over a repack/multi-source variant.
            mfg = [v for v in vlist if v.get("bom_class") == "manufacturing"]
            sel, _ = select_bom(mfg or vlist, s)
            _sel_cache[ck] = sel
        return _sel_cache[ck]

    def _is_assembly(code, desc):
        # A yielded leaf is an "unresolved intermediate" only if it genuinely has a
        # recipe to explode (same rule as _sel_variant) but couldn't be resolved
        # (circular / depth-capped). A leaf with no real recipe is just a purchased RM.
        return _sel_variant(code, desc) is not None

    # Decode map: an ENCODED item name -> its real name, by following the item's
    # encode/decode ALIAS BOM (1:1 single non-packing component, or internal
    # DECODE/CONVERSION). e.g. RDNBP101 -> PUREPRINT AFT, ASPP001 -> P.T.S ACID.
    # Only names that LOOK encoded (a code like ABCDE123, no spaces) are decoded, so a
    # real chemical name (TRI SODIUM PHOSPHATE, ACETIC ACID) is never touched. Keyed by
    # both item code and squashed name, so every code of a multi-code intermediate resolves.
    def _looks_encoded(name):
        return bool(_ENC_NAME_RE.match(_norm(name).upper()))

    def _decode_name(code, desc, depth=0, seen=frozenset()):
        if not _looks_encoded(desc):
            return desc                      # already a real name -> stop (avoids the
        node = _squash(desc) or code         # encoded<->real alias cycle recursing back)
        if depth > 8 or node in seen:
            return desc
        vlist = asm_by_code.get(code) or bom_by_squash.get(_squash(desc)) or []
        aliases = [v for v in vlist if _is_alias_bom(v) or v.get("bom_class") == "internal"]
        if not aliases:
            return desc
        nonpack = [c for c in aliases[0]["components"] if not _is_packing_comp(c)]
        if not nonpack:
            return desc
        return _decode_name(nonpack[0]["comp_code"], nonpack[0]["comp_desc"], depth + 1, seen | {node})

    decode_map: dict[str, str] = {}
    for _code, _vl in asm_by_code.items():
        _ad = _vl[0]["assembly_desc"]
        if not _looks_encoded(_ad):
            continue
        if not any(_is_alias_bom(v) or v.get("bom_class") == "internal" for v in _vl):
            continue
        _dec = _decode_name(_code, _ad)
        if _squash(_dec) != _squash(_ad) and not _looks_encoded(_dec):   # resolved to a REAL name
            decode_map[_code] = _dec
            decode_map.setdefault(_squash(_ad), _dec)                    # also by name (multi-code)

    # Lead time by squashed DESCRIPTION — a fallback for when an RM's own item code has
    # no PO history. Purchases are often booked under a different code of the same
    # material, or under the DECODED real name of an encoded item (e.g. LTLP005 =
    # LUTENSOL TO5). Aggregated from the code-keyed po_intel by item name.
    po_lead_by_desc: dict[str, dict] = {}
    for _cd, _in in (po_intel or {}).items():
        _nm = _squash(_in.get("name"))
        _l = _in.get("avg_lead_time_days")
        if _nm and _l is not None:
            _e = po_lead_by_desc.setdefault(_nm, {"sum": 0.0, "n": 0, "pos": 0})
            _e["sum"] += _l
            _e["n"] += 1
            _e["pos"] += _in.get("lead_basis_pos", 0)

    def _rm_row_lead(codes, desc):
        """(avg lead days, basis_pos) for an RM, filling missing encoded-item leads via
        3 tiers: (1) the item's own PO history by code; (2) by description (same material
        under another code); (3) decode the encoded name to the real item, use its lead."""
        leads, pos = [], 0
        for cd in codes or []:
            intel = (po_intel or {}).get(cd) or {}
            if intel.get("avg_lead_time_days") is not None:
                leads.append(intel["avg_lead_time_days"])
                pos += intel.get("lead_basis_pos", 0)
        if leads:
            return round(sum(leads) / len(leads), 1), pos
        d = po_lead_by_desc.get(_squash(desc))
        if not d:
            dec = decode_map.get((codes or [""])[0]) or decode_map.get(_squash(desc))
            if dec:
                d = po_lead_by_desc.get(_squash(dec))
        if d and d["n"]:
            return round(d["sum"] / d["n"], 1), d["pos"]
        return None, 0

    products, matched, pending_applied, out_of_scope = [], 0, 0, 0
    for v in items:
        is_excel = _squash(v["name"]) in excel_keys
        is_msl_only = _squash(v["name"]) in msl_only_keys
        # Division scope: drop items outside the allowed division(s) — Excel-uploaded and
        # MSL-only items (already scoped at injection) are exempt.
        if not is_excel and not is_msl_only and not _in_scope(v["name"]):
            out_of_scope += 1
            continue
        if not is_excel and not is_msl_only and drop_zero and not _qualifies(v):
            continue
        if not is_excel and not is_msl_only and max_products and len(products) >= max_products:
            break
        variants = bom_by_desc.get(v["name"].upper()) or bom_by_squash.get(_squash(v["name"]), [])
        mfg = [x for x in variants if not x["is_packing"]]
        packing = [x for x in variants if x["is_packing"]]
        if mfg:
            selected, alts = select_bom(mfg, s)
        else:
            selected, alts = select_bom(packing, s)
            packing = []
        # user BOM override: force the chosen variant as the effective (preferred) BOM
        overridden = False
        ov = bom_overrides.get(_squash(v["name"])) if bom_overrides else None
        if ov and variants:
            chosen = next((x for x in variants if _variant_key(x) == ov), None)
            if chosen is not None and chosen is not selected:
                pool = mfg if chosen in mfg else variants
                selected, alts = chosen, [x for x in pool if x is not chosen]
                packing = [x for x in packing if x is not chosen]
                overridden = True
        mfg_soc = v.get("mfg_soc", 0.0) if jc_on["current"] else 0.0
        overall_soc = v.get("overall_soc", 0.0) if jc_on["current"] else 0.0
        target = round(v["current"], 1) if jc_on["current"] else 0.0
        # MSL safety-stock buffer — added to the Current-JC demand so the plan tops the
        # item back up to its minimum stock level. Only valid items are in msl_map (freq
        # > 10 AND unique customers > 5, filtered upstream); everything else buffers 0.
        msl_buf = round((msl_map or {}).get(_squash(v["name"]), 0.0), 1) if jc_on["current"] else 0.0
        # Current JC = projection WK1+WK2 + MFG SOC pending (planning orgs)
        pj = {"current_target": target, "mfg_soc": mfg_soc, "overall_soc": overall_soc,
              "pending_soc": mfg_soc,   # legacy alias (Current uses MFG SOC)
              "msl": msl_buf,
              "current": round(target + mfg_soc, 1),
              "next1": round(v["next1"], 1) if jc_on["next1"] else 0.0,
              "next2": round(v["next2"], 1) if jc_on["next2"] else 0.0}
        pj["total"] = round(pj["current"] + pj["next1"] + pj["next2"], 1)
        # Manufacturing required (MSL-driven) = (projection + MFG SOC + MSL) − on-hand FG
        # stock at BOTH the warehouse and the branches, for the Current JC and the full
        # 3-JC horizon (Total). MSL is a one-time buffer added once to each.
        _fg = fg_by_desc.get(_squash(v["name"]), {})
        _wh = _fg.get("warehouse", 0.0)
        _br = _fg.get("branch", 0.0)
        _onhand = _wh + _br
        pj["onhand_warehouse"] = round(_wh, 1)
        pj["onhand_branch"] = round(_br, 1)
        pj["onhand"] = round(_onhand, 1)
        pj["mfg_required"] = round(max(0.0, pj["current"] + msl_buf - _onhand), 1)
        pj["mfg_required_3jc"] = round(max(0.0, pj["total"] + msl_buf - _onhand), 1)
        if mfg_soc > 0:
            pending_applied += 1

        def _bom_entry(variant, preferred):
            # FG on-hand by item DESCRIPTION (all codes of this product — bulk +
            # packed variants), not just the BOM assembly code.
            fg = fg_by_desc.get(_squash(v["name"]), {})
            return {
                "assembly_item": variant["assembly_item"], "org_code": variant["org_code"],
                "designator": variant["designator"], "bom_type": variant.get("bom_type", ""),
                "bom_class": variant.get("bom_class", "unclassified"),
                "created": variant.get("created_iso"), "preferred": preferred,
                "fg_stock": {"warehouse": round(fg.get("warehouse", 0.0), 1),
                             "branch": round(fg.get("branch", 0.0), 1)},
                "components": _components_for(variant, pj, rm_stock, po, consider_subs,
                                              po_intel=po_intel, preproc=preproc,
                                              sub_exclude=sub_exclude, lead_fn=_rm_row_lead),
            }

        davg = (dispatch_avg or {}).get(_squash(v["name"]), {})
        avg_sales = round(davg.get("avg", 0.0), 1)
        prod = {"name": v["name"], "segment2": v.get("segment2", ""),
                "segment3": v.get("segment3", ""), "projection": pj, "pending_soc": mfg_soc,
                "avg_3jc_sales": avg_sales, "sales_variance": round(target - avg_sales, 1),
                "proj_flag": _proj_flag(target, avg_sales),
                "has_bom": selected is not None, "alternatives": len(alts),
                "bom_class": selected.get("bom_class", "unclassified") if selected else "trading",
                "bom_variant": _variant_key(selected) if selected else None,
                "overridden": overridden, "msl_only": is_msl_only,
                "boms": [], "packing_boms": []}
        if selected:
            matched += 1
            for k, variant in enumerate([selected] + alts):
                prod["boms"].append(_bom_entry(variant, k == 0))
        for variant in sorted(packing, key=lambda x: (x["org_code"] != "PMO", x["assembly_item"])):
            prod["packing_boms"].append(_bom_entry(variant, False))
        # Producible FG qty from current available RM (bottleneck: scarcest component).
        # producible_solo = independent value (full stock); producible_qty is replaced
        # below by the shared-RM, PTS-first allocation (capped at Current demand).
        prod["producible_solo"] = _producible_from_rm(prod["boms"][0]) if prod["boms"] else 0.0
        prod["producible_qty"] = prod["producible_solo"]
        products.append(prod)

    # ── Producible allocation (RAW MATERIAL only — packing is not a constraint) ───
    # An RM can feed several FGs. Serve PTS (Plan-To-Stock) FGs before PTO, each
    # consuming shared RM only up to its Current-JC Mfg-Required demand; the leftover
    # flows to the next FG. producible_qty becomes the qty this FG actually gets to make.
    # Packing material (P… codes) is excluded so producible reflects RM availability.
    _pto = pto_map or {}
    pool: dict[str, float] = {}
    for p in products:
        if p.get("has_bom"):
            for c in p["boms"][0]["components"]:
                if not _pack_code(c["rm_code"]):
                    pool.setdefault(c["rm_code"], c.get("available", 0.0))

    def _policy(p):
        return _pto.get(_squash(p["name"]), "PTO")

    for p in products:
        p["pts_pto"] = _policy(p) if p.get("has_bom") else ""
    # PTS before PTO; within a policy, larger Current demand first (deterministic).
    for p in sorted((p for p in products if p.get("has_bom")),
                    key=lambda p: (0 if _policy(p) == "PTS" else 1,
                                   -p["projection"].get("mfg_required", 0.0))):
        comps = [c for c in p["boms"][0]["components"] if not _pack_code(c["rm_code"])]
        demand = p["projection"].get("mfg_required", 0.0)
        cap = None
        for c in comps:
            q = c.get("qty_per_unit") or 0.0
            if q <= 0:
                continue
            u = pool.get(c["rm_code"], 0.0) / q
            cap = u if cap is None else min(cap, u)
        cap = max(0.0, cap if cap is not None else 0.0)
        # producible = makeable from the RM still available to this FG (after higher-
        # priority FGs consumed theirs); consumption is capped at its Current demand.
        p["producible_qty"] = round(cap, 1)
        make = min(demand, cap)
        for c in comps:
            q = c.get("qty_per_unit") or 0.0
            if q > 0:
                pool[c["rm_code"]] = max(0.0, pool.get(c["rm_code"], 0.0) - make * q)

    # Express Producible against the manufacturing demand: how much of the Current-JC
    # and the full 3-JC Mfg-Required it covers (Producible itself is an RM-driven
    # capacity, not a JC horizon — this frames it in demand terms for the planner).
    for p in products:
        if not p.get("has_bom"):
            continue
        pj = p["projection"]
        prodq = p.get("producible_qty", 0.0)
        req1 = pj.get("mfg_required", 0.0)
        req3 = pj.get("mfg_required_3jc", 0.0)
        pj["producible_pct_current"] = round(100 * prodq / req1) if req1 > 0 else None
        pj["producible_pct_3jc"] = round(100 * prodq / req3) if req3 > 0 else None
        # Current-JC producible = how much of THIS JC's Mfg-Required can be made now from
        # available RM (capped at the requirement), plus a plain Satisfied/Need-plan flag.
        pj["producible_current"] = round(min(prodq, req1), 1) if req1 > 0 else 0.0
        if req1 <= 0:
            pj["producible_status_current"] = "No mfg needed"   # warehouse covers Current JC
        elif prodq >= req1:
            pj["producible_status_current"] = "Satisfied"       # RM on hand makes full Current JC
        else:
            pj["producible_status_current"] = "Need plan"       # RM short → must purchase
        if req3 <= 0:
            cover = "No mfg needed"          # warehouse stock already covers demand
        elif prodq >= req3:
            cover = "Covers 3 JC"
        elif req1 > 0 and prodq >= req1:
            cover = "Covers Current JC"
        elif prodq > 0:
            cover = f"Partial ({pj['producible_pct_current'] if req1 > 0 else pj['producible_pct_3jc']}%)"
        else:
            cover = "None (RM short)"
        pj["producible_cover"] = cover

    # ── consolidate RM by ITEM DESCRIPTION (not code) so the planning team sees one
    #    row per material name, summing demand + stock across all its item codes.
    #    Packing material (COMPONENT_ITEM code starting 'P') is split into its own
    #    plan (want_packing=True) and removed from the material plans.
    def _is_pack_code(code):
        return (code or "").lstrip()[:1].upper() == "P"

    # stock + PO rolled up by description (across every item code of that material)
    stock_by_desc: dict[str, float] = {}
    stock_org_by_desc: dict[str, dict] = {}   # {desc -> {org -> qty}} for the org breakdown
    stock_name_by_sq: dict[str, str] = {}     # squash -> original (display) stock name
    for _cd, _v in rm_stock.items():
        k = _squash(_v.get("name"))
        if k:
            stock_by_desc[k] = stock_by_desc.get(k, 0.0) + _num(_v.get("qty"))
            stock_name_by_sq.setdefault(k, _norm(_v.get("name")))
            _od = stock_org_by_desc.setdefault(k, {})
            for _org, _oq in (_v.get("orgs") or {}).items():
                _od[_org] = _od.get(_org, 0.0) + _num(_oq)

    # Decode-aware stock: a raw material may sit in CRM BiStock under its real (decoded)
    # name, under an ENCODED code's name, or BOTH (different lots/codes). Index every
    # encoded stock-name by the real name it decodes to, so the Real-RM inventory can
    # include both — and keep the encoded part separate so the user can see it.
    _enc_keys_for_real: dict[str, list] = {}
    for _sk in list(stock_org_by_desc):
        _real = decode_map.get(_sk)          # _sk is an encoded stock-name -> real name
        if _real:
            _enc_keys_for_real.setdefault(_squash(_real), []).append(_sk)

    def _merged_stock(key, codes=()):
        """(total, {org: qty}, {org: encoded_qty}, [encoded names]) merging real-name +
        encoded-name stock. org_enc / enc_names describe the part held under encoded name(s)."""
        real_name = decode_map.get(key) or decode_map.get((list(codes) or [""])[0])
        canon = _squash(real_name) if real_name else key
        org_total: dict = {}
        org_enc: dict = {}
        enc_names: set = set()
        for _o, _q in stock_org_by_desc.get(canon, {}).items():   # stock under the real name
            org_total[_o] = org_total.get(_o, 0.0) + _q
        enc = set(_enc_keys_for_real.get(canon, []))
        if real_name:                        # the leaf itself is an encoded name
            enc.add(key)
        for _ek in enc:                      # stock under encoded name(s) that decode here
            if _ek == canon or not stock_org_by_desc.get(_ek):
                continue
            enc_names.add(stock_name_by_sq.get(_ek, _ek))
            for _o, _q in stock_org_by_desc.get(_ek, {}).items():
                org_total[_o] = org_total.get(_o, 0.0) + _q
                org_enc[_o] = org_enc.get(_o, 0.0) + _q
        return round(sum(org_total.values()), 1), org_total, org_enc, sorted(enc_names)

    def _stock_org_str(org_total, org_enc):
        """Per-org breakdown, largest first, flagging orgs whose stock is under an encoded
        name, e.g. 'POI - Alathur MFG (400), PCM - Madhavaram MFG (50, encoded)'."""
        parts = sorted(((o, q) for o, q in org_total.items() if round(q, 1) > 0), key=lambda x: -x[1])
        out = []
        for o, q in parts:
            eq = round(org_enc.get(o, 0.0), 1)
            if eq >= round(q, 1):
                out.append(f"{o} ({round(q)}, encoded)")
            elif eq > 0:
                out.append(f"{o} ({round(q)}, incl. {round(eq)} encoded)")
            else:
                out.append(f"{o} ({round(q)})")
        return ", ".join(out)
    po_by_desc: dict[str, dict] = {}
    for _cd, _v in (po or {}).items():
        k = _squash(_v.get("name"))
        if not k:
            continue
        a = po_by_desc.setdefault(k, {"in_transit": 0.0, "received": 0.0})
        a["in_transit"] += _num(_v.get("in_transit"))
        a["received"] += _num(_v.get("received"))

    _bmap = business_map or {}

    def _biz_for(codes):
        """Distinct Business value(s) for a material's item codes (Raw Material,
        GC1, GC2 …) from the CRM item-business map. Usually one value."""
        seen = []
        for c in codes:
            b = _norm(_bmap.get(c) or _bmap.get(_norm(c)) or "")
            if b and b not in seen:
                seen.append(b)
        return ", ".join(seen)

    # Activity a material serves — the set of bom_classes of every FG that uses it —
    # so the "Activity" column is consistent for a material across all RM sheets.
    rm_activity: dict[str, set] = {}
    for _p in products:
        if not _p.get("has_bom"):
            continue
        _cls = _p.get("bom_class", "unclassified")
        for _c in _p["boms"][0]["components"]:
            _k = _squash(_c["rm_desc"]) or _c["rm_code"]
            rm_activity.setdefault(_k, set()).add(_cls)

    def _act_from(classes, packing=False):
        if packing:
            return "Packing"
        acts = {a for a in classes if a}
        parts = []
        if "manufacturing" in acts:
            parts.append("Manufacturing")
        if "repack_relabel" in acts:
            parts.append("Repack/Relabel")
        if acts - {"manufacturing", "repack_relabel"}:
            parts.append("Other")
        return " + ".join(parts) if parts else "Unclassified"

    # ── Real-RM explosion (helpers hoisted above the products loop) ─────────────
    # asm_by_code / _sel_variant / _is_assembly / decode_map / _rm_row_lead are now
    # defined before the products loop so _components_for can use the decode-aware
    # lead-time fallback. _explode_gross / _real_rm_rows follow below.

    def _explode_gross(code, desc, gross, depth=0, seen=frozenset(), trail=()):
        """Yield (leaf_code, leaf_desc, leaf_gross, intermediates_trail) for a
        component, exploding any intermediate that has its own BOM. leaf_gross is the
        3-JC demand for the leaf; trail = the intermediate names traversed to reach it.
        Cycles are detected by ITEM CODE (not name): a packed variant and its own bulk
        base share the same DESCRIPTION but are different codes, so a normal packed->bulk
        step must not look like a cycle. Genuine loops (the same code re-appearing) are
        still caught, and the depth cap backstops any name-collision runaway."""
        node = _norm(code) or _squash(desc)
        sel = None if (depth > 12 or node in seen) else _sel_variant(code, desc)
        if not sel or not sel.get("components"):
            yield (code, desc, gross, trail)
            return
        for cc in sel["components"]:
            sub_gross = {k: gross[k] * cc["qty"] for k in ("current", "next1", "next2")}
            yield from _explode_gross(cc["comp_code"], cc["comp_desc"], sub_gross,
                                      depth + 1, seen | {node}, trail + (desc or code,))

    def _real_rm_rows(prod_subset, want_packing=False):
        """Consolidated LEAF raw-material requirement (decoded names), after exploding
        every intermediate to its purchased RMs. Stock/net computed like _consolidate."""
        cons: dict[str, dict] = {}
        for p in prod_subset:
            if not p["has_bom"]:
                continue
            for c in p["boms"][0]["components"]:
                for lc, ld, lg, trail in _explode_gross(c["rm_code"], c["rm_desc"], c["gross"]):
                    if _is_pack_code(lc) != want_packing:
                        continue
                    if _is_dm_water(lc, ld):        # DM water excluded from planning
                        continue
                    # Key each leaf by its DECODED real name when the leaf is encoded, so
                    # encoded and real-name variants of the same material become ONE row
                    # (displayed under the real name, its stock counted once).
                    _real = decode_map.get(lc) or decode_map.get(_squash(ld))
                    _disp = _real or (ld or lc)
                    key = _squash(_disp) or lc
                    a = cons.setdefault(key, {
                        "rm_desc": _disp, "codes": set(), "intermediates": set(),
                        "gross": {"current": 0.0, "next1": 0.0, "next2": 0.0},
                        "fgs": set(), "activities": set(), "unresolved": False})
                    a["codes"].add(lc)
                    a["fgs"].add(p["name"])
                    a["activities"].add(p.get("bom_class", "unclassified"))
                    a["intermediates"].update(t for t in trail if t)
                    # a "leaf" that is itself an assembly = an intermediate the explosion
                    # could not fully resolve (circular / depth-capped encoded BOM).
                    if _is_assembly(lc, ld):
                        a["unresolved"] = True
                    for k in ("current", "next1", "next2"):
                        a["gross"][k] += lg[k]
        out = []
        for key, a in cons.items():
            gross = {k: round(a["gross"][k], 1) for k in ("current", "next1", "next2")}
            _stk_total, _stk_org, _stk_enc, _enc_names = _merged_stock(key, a["codes"])
            main_stock = _stk_total
            pod = po_by_desc.get(key, {})
            in_transit, received = pod.get("in_transit", 0.0), pod.get("received", 0.0)
            available = main_stock + in_transit
            net = _net_3jc(gross, available)
            codes = sorted(a["codes"])
            avg_lead, _lead_pos = _rm_row_lead(codes, a["rm_desc"])
            planned = _lead_jcs(avg_lead, preproc)
            out.append({
                "rm_code": codes[0], "rm_codes": codes, "code_count": len(codes),
                "rm_desc": a["rm_desc"], "business": _biz_for(codes),
                "activity": _act_from(a["activities"], packing=want_packing), "gross": gross,
                "gross_total": round(sum(gross.values()), 1),
                "main_stock": round(main_stock, 1), "stock_orgs": _stock_org_str(_stk_org, _stk_enc),
                "encoded_names": ", ".join(_enc_names),
                "encoded_stock": round(sum(_stk_enc.values()), 1),
                "has_encoded_stock": bool(_stk_enc),
                "in_transit": round(in_transit, 1),
                "received": round(received, 1), "available": round(available, 1),
                "net_to_buy": net, "net_total": round(sum(net[k] for k in planned), 1),
                "planned_jcs": list(planned),
                "buy_jcs": [k for k in planned if net[k] > 0],
                "to_buy": any(net[k] > 0 for k in planned),
                "lead_total_days": round(avg_lead + preproc, 1) if avg_lead is not None else None,
                "avg_lead_time_days": avg_lead,
                "fg_count": len(a["fgs"]), "fgs": sorted(a["fgs"])[:20],
                "from_intermediates": sorted(a["intermediates"])[:8],
                "via_intermediate": bool(a["intermediates"]),
                "unresolved": a.get("unresolved", False),
            })
        out.sort(key=lambda x: -x["net_total"])
        return out

    def _consolidate(prod_subset, want_packing=False):
        cons: dict[str, dict] = {}
        for p in prod_subset:
            if not p["has_bom"]:
                continue
            for c in p["boms"][0]["components"]:
                if _is_pack_code(c["rm_code"]) != want_packing:
                    continue
                if _is_dm_water(c["rm_code"], c["rm_desc"]):   # DM water excluded from planning
                    continue
                key = _squash(c["rm_desc"]) or c["rm_code"]
                a = cons.setdefault(key, {
                    "rm_desc": c["rm_desc"], "codes": set(),
                    "gross": {"current": 0.0, "next1": 0.0, "next2": 0.0}, "fgs": set(), "subs": {}})
                a["codes"].add(c["rm_code"])
                for k in ("current", "next1", "next2"):
                    a["gross"][k] += c["gross"][k]
                a["fgs"].add(p["name"])
                for su in c["substitutes"]:
                    a["subs"][su["code"]] = {"desc": su.get("desc", ""), "stock": su["stock"]}
        out = []
        for key, a in cons.items():
            gross = {k: round(a["gross"][k], 1) for k in ("current", "next1", "next2")}
            main_stock = stock_by_desc.get(key, 0.0)
            sub_stock = sum(v["stock"] for v in a["subs"].values()) if consider_subs else 0.0
            pod = po_by_desc.get(key, {})
            in_transit, received = pod.get("in_transit", 0.0), pod.get("received", 0.0)
            available = main_stock + sub_stock + in_transit
            net = _net_3jc(gross, available)
            # PO intelligence (lead time, suppliers) aggregated across the material's codes
            suppliers, locations, currencies, trade, npos = set(), set(), set(), None, 0
            for code in a["codes"]:
                intel = (po_intel or {}).get(code)
                if not intel:
                    continue
                suppliers |= set(intel.get("suppliers", []))
                locations |= set(intel.get("locations", []))
                currencies |= set(intel.get("currencies", []))
                npos += intel.get("lead_basis_pos", 0)
                trade = trade or intel.get("trade")
            codes = sorted(a["codes"])
            # lead time with the encoded-name decode fallback (fills missing leads)
            avg_lead, _lead_pos = _rm_row_lead(codes, a["rm_desc"])
            npos = npos or _lead_pos
            planned = _lead_jcs(avg_lead, preproc)   # JC buckets to buy by lead time
            out.append({
                "rm_code": codes[0], "rm_codes": codes, "code_count": len(codes),
                "rm_desc": a["rm_desc"], "business": _biz_for(codes),
                "activity": _act_from(rm_activity.get(key, ()), packing=want_packing), "gross": gross,
                "gross_total": round(sum(gross.values()), 1),
                "main_stock": round(main_stock, 1), "substitute_stock": round(sub_stock, 1),
                "received": round(received, 1), "in_transit": round(in_transit, 1),
                "available": round(available, 1), "net_to_buy": net,
                "net_total": round(sum(net[k] for k in planned), 1),
                "planned_jcs": list(planned),
                "buy_jcs": [k for k in planned if net[k] > 0],
                "to_buy": any(net[k] > 0 for k in planned),
                "lead_total_days": round(avg_lead + preproc, 1) if avg_lead is not None else None,
                "fg_count": len(a["fgs"]), "fgs": sorted(a["fgs"])[:20],
                "substitutes": [{"code": k, "desc": v["desc"], "stock": round(v["stock"], 1)}
                                for k, v in a["subs"].items()],
                "avg_lead_time_days": avg_lead,
                "lead_basis_pos": npos,
                "suppliers": sorted(suppliers)[:8], "supplier_count": len(suppliers),
                "locations": sorted(locations), "trade": trade, "currencies": sorted(currencies),
            })
        out.sort(key=lambda x: -x["net_total"])
        return out

    mfg_products = [p for p in products if p["bom_class"] == "manufacturing"]
    repack_products = [p for p in products if p["bom_class"] == "repack_relabel"]
    bom_products = [p for p in products if p["has_bom"]]
    # Per-FG leaf RMs hidden INSIDE an intermediate direct-component — surfaced for the
    # "Selected BOM RMs" sheet (tagged "Intermediate BOM", with the via-intermediate).
    for _p in bom_products:
        _im: dict = {}
        for _c in _p["boms"][0]["components"]:
            for _lc, _ld, _lg, _trail in _explode_gross(_c["rm_code"], _c["rm_desc"], _c["gross"]):
                if not _trail:
                    continue                       # a direct leaf — already a direct BOM row
                if _is_dm_water(_lc, _ld):         # DM water excluded from planning
                    continue
                _k = _squash(_ld) or _lc
                _a = _im.setdefault(_k, {"rm_code": _lc, "rm_desc": _ld or _lc, "via": set(),
                                         "gross": {"current": 0.0, "next1": 0.0, "next2": 0.0}})
                _a["via"].update(t for t in _trail if t)
                for _kk in ("current", "next1", "next2"):
                    _a["gross"][_kk] += _lg[_kk]
        _rows = []
        for _a in _im.values():
            _g = {k: round(_a["gross"][k], 1) for k in ("current", "next1", "next2")}
            _rows.append({"rm_code": _a["rm_code"], "rm_desc": _a["rm_desc"],
                          "via": ", ".join(sorted(_a["via"])[:5]),
                          "gross": _g, "gross_total": round(sum(_g.values()), 1)})
        _rows.sort(key=lambda x: -x["gross_total"])
        _p["intermediate_bom_rms"] = _rows
    consolidated = _consolidate(bom_products)                       # materials, all activities
    consolidated_mfg = _consolidate(mfg_products)                   # materials, manufacturing
    consolidated_repack = _consolidate(repack_products)             # materials, repack/relabel
    consolidated_packing = _consolidate(bom_products, want_packing=True)   # packing materials
    # Real (leaf) RM requirement — intermediates exploded to purchased RMs, decoded
    real_rm = _real_rm_rows(bom_products)                           # true RMs to buy
    real_rm_manufacturing = _real_rm_rows(mfg_products)

    n_projected = sum(1 for v in proj.values() if (not drop_zero or _qualifies(v)))
    pending_in_scope = round(sum(p["projection"].get("mfg_soc", 0.0) for p in products), 1)
    overall_in_scope = round(sum(p["projection"].get("overall_soc", 0.0) for p in products), 1)
    mfg_required_total = round(sum(p["projection"].get("mfg_required", 0.0) for p in products), 1)
    msl_total = round(sum(p["projection"].get("msl", 0.0) for p in products), 1)
    msl_items = sum(1 for p in products if p["projection"].get("msl", 0.0) > 0)
    msl_only_count = sum(1 for p in products if p.get("msl_only"))
    onhand_total = round(sum(p["projection"].get("onhand", 0.0) for p in products), 1)
    def _cons_summary(lst):
        unresolved = [x for x in lst if x.get("unresolved")]
        return {"distinct_rms": len(lst),
                "rms_to_buy": sum(1 for x in lst if x["net_total"] > 0),
                "total_buy_qty": round(sum(x["net_total"] for x in lst), 1),
                "unresolved_intermediates": len(unresolved),
                "unresolved_qty": round(sum(x["net_total"] for x in unresolved), 1)}

    return {
        "products": products, "consolidated_rm": consolidated,
        "consolidated_rm_manufacturing": consolidated_mfg,
        "consolidated_rm_repack": consolidated_repack,
        "consolidated_rm_packing": consolidated_packing,
        "consolidated_summary": _cons_summary(consolidated),
        "consolidated_summary_manufacturing": _cons_summary(consolidated_mfg),
        "consolidated_summary_repack": _cons_summary(consolidated_repack),
        "consolidated_summary_packing": _cons_summary(consolidated_packing),
        "real_rm_requirement": real_rm,
        "real_rm_requirement_manufacturing": real_rm_manufacturing,
        "real_rm_summary": _cons_summary(real_rm),
        "decode_map": decode_map,
        "decode_names": s.get("decode_encoded_names", True),
        "planned_jcs": [k for k in ("current", "next1", "next2") if jc_on[k]],
        "soc_window": {"from": soc_window[0], "to": soc_window[1]} if soc_window else None,
        "po_window": po_window,
        "summary": {
            "projected_products": n_projected, "shown": len(products), "with_bom": matched,
            "manufacturing": len(mfg_products), "repack_relabel": len(repack_products),
            "internal": sum(1 for p in products if p["bom_class"] == "internal"),
            "trading": sum(1 for p in products if p["bom_class"] == "trading"),
            "unclassified": sum(1 for p in products if p["bom_class"] == "unclassified"),
            "without_bom": len(products) - matched, "rm_items_in_stock": len(rm_business),
            "rm_items_in_stock_all": len(rm_stock),
            "fg_items_in_stock": len(fg_stock),
            "po_pending_items": sum(1 for k, v in po.items() if (v.get("in_transit") or 0) > 0 and not _pack_code(k)),
            "po_items_total": len(po),
            "with_packing_bom": sum(1 for p in products if p.get("packing_boms")),
            "packing_bom_count": sum(len(p.get("packing_boms", [])) for p in products),
            "stock_source": "CRM SPBiStockDetails (BiStockDetail)" if stock_rows is not None
                            else "Excel stock-agings",
            "pending_soc_total": round(mfg_total, 1), "pending_soc_in_plan": pending_in_scope,
            "overall_soc_total": round(overall_total, 1), "overall_soc_in_plan": overall_in_scope,
            "mfg_soc_in_plan": pending_in_scope, "mfg_required_total": mfg_required_total,
            "msl_total": msl_total, "msl_items": msl_items, "onhand_total": onhand_total,
            "msl_only_items": msl_only_count,
            "plan_divisions": sorted(_allowed_div), "out_of_scope_items": out_of_scope,
            "mfg_required_formula": "(Projection + Pending SOC + MSL) − On-hand (Warehouse + Branch)",
            "pending_soc_items": pending_applied,
            "preprocessing_days": preproc,
            "min_plan_qty": min_qty,
            "over_projected": sum(1 for p in products if p.get("proj_flag") == "over"),
            "under_projected": sum(1 for p in products if p.get("proj_flag") == "under"),
        },
        "rules": {
            "warehouse_orgs": sorted(s.get("warehouse_orgs", WAREHOUSE_ORGS)),
            "rm_source_orgs": sorted(s.get("rm_source_orgs", RM_SOURCE_ORGS)),
            "intermediate_stock_orgs": sorted(s.get("intermediate_stock_orgs", [])),
            "intransit_rm_only_orgs": sorted(s.get("intransit_rm_only_orgs", [])),
            "mfg_soc_orgs": sorted(s.get("mfg_soc_orgs", [])),
            "dm_water_codes": sorted(sub_exclude),
            "excluded_subinv": sorted(s.get("excluded_subinv", EXCLUDE_SUBINV)),
            "bom_preference": " -> ".join(
                (["PMO"] if s.get("bom_prefer_pmo", True) else [])
                + (["BULK/HDLK"] if s.get("bom_prefer_bulk_hdlk", True) else [])
                + (["newest BOM"] if s.get("bom_prefer_creation_date", True) else [])
                + (["Primary"] if s.get("bom_prefer_primary", True) else [])) or "(no preference)",
            "consider_substitutes": consider_subs,
        },
    }


# --- 6. Aged-RM -> FG production ----------------------------------------------
def _usable_pool(rows, excluded, by_code=None, business_map=None, rm_only=False,
                 rm_label="Raw Material"):
    pool: dict[str, dict] = by_code if by_code is not None else {}
    for r in rows or []:
        sub = _norm(r.get("SubInv"))
        if sub.lower() in excluded:
            continue
        code = _norm(r.get("ItemCode"))
        if not code:
            continue
        if rm_only and not _is_raw_material((business_map or {}).get(code, ""), rm_label):
            continue
        qty = _num(r.get("Qty"))
        if qty <= 0:
            continue
        a = pool.setdefault(code, {"desc": _norm(r.get("ItemDesc")), "qty": 0.0,
                                   "cost": _num(r.get("ItemCost")), "max_age": 0})
        a["qty"] += qty
        if not a["cost"]:
            a["cost"] = _num(r.get("ItemCost"))
        a["max_age"] = max(a["max_age"], int(_num(r.get("MaxAgeDays"))))
    return pool


def build_aged_rm_plan(bom_path, aged_rows, all_rows=None, settings=None,
                       aged_days=90, business_map=None) -> dict:
    from . import planning_settings
    s = settings or planning_settings.load()
    consider_subs = s.get("consider_substitutes", True)
    excluded = {x.lower() for x in s.get("excluded_subinv", EXCLUDE_SUBINV)}
    decode = s.get("decode_encoded_names", True)
    rm_label = s.get("raw_material_business", "Raw Material")

    aged = _usable_pool(aged_rows, excluded, business_map=business_map, rm_only=True, rm_label=rm_label)
    total = _usable_pool(all_rows, excluded) if all_rows is not None else dict(aged)

    bom_idx = load_bom_detailed(bom_path) if bom_path else {"by_desc": {}, "by_squash": {}}
    fgs = []
    for variants in bom_idx["by_desc"].values():
        mfg = [v for v in variants if not v["is_packing"]]
        sel, _alts = select_bom(mfg or variants, s)
        if not sel or not sel["components"]:
            continue
        comps = []
        for c in sel["components"]:
            subs = [su["code"] for su in c["substitutes"]] if consider_subs else []
            comps.append({"rm_code": c["comp_code"], "rm_desc": c["comp_desc"], "qty": c["qty"], "subs": subs})
        fgs.append({"name": sel["assembly_desc"], "assembly_item": sel["assembly_item"],
                    "org_code": sel["org_code"], "designator": sel["designator"], "components": comps})

    def avail(pool, comp):
        return pool.get(comp["rm_code"], {}).get("qty", 0.0) \
            + sum(pool.get(x, {}).get("qty", 0.0) for x in comp["subs"])

    def aged_codes(comp):
        return [comp["rm_code"]] + comp["subs"]

    producible, blocked = [], []
    for fg in fgs:
        rows, all_aged, any_aged, all_total = [], True, False, True
        units = float("inf")
        aged_value_per_unit = 0.0
        for c in fg["components"]:
            av_aged, av_total = avail(aged, c), avail(total, c)
            any_aged = any_aged or av_aged > 0
            all_aged = all_aged and av_aged > 0
            all_total = all_total and av_total > 0
            if c["qty"] > 0 and av_aged > 0:
                units = min(units, av_aged / c["qty"])
            cost = aged.get(c["rm_code"], {}).get("cost", 0.0)
            aged_value_per_unit += c["qty"] * cost
            rows.append({"rm_code": c["rm_code"], "rm_desc": c["rm_desc"], "qty_per_unit": c["qty"],
                         "aged_stock": round(av_aged, 1), "total_stock": round(av_total, 1),
                         "aged_age_days": aged.get(c["rm_code"], {}).get("max_age", 0),
                         "ok_aged": av_aged > 0, "ok_total": av_total > 0, "substitutes": c["subs"]})
        if not any_aged:
            continue
        units = 0.0 if units == float("inf") else units
        rec = {"name": fg["name"], "assembly_item": fg["assembly_item"], "org_code": fg["org_code"],
               "designator": fg["designator"], "producible_units": round(units, 1),
               "aged_consumed": round(units * sum(c["qty"] for c in fg["components"]), 1),
               "aged_value_consumed": round(units * aged_value_per_unit, 0),
               "all_aged": all_aged, "all_available": all_total,
               "missing": [r["rm_desc"] if decode else r["rm_code"] for r in rows if not r["ok_total"]],
               "needs_fresh": all_total and not all_aged, "components": rows}
        (producible if all_aged else blocked).append(rec)
    producible.sort(key=lambda x: -x["aged_consumed"])
    blocked.sort(key=lambda x: -sum(r["aged_stock"] for r in x["components"]))

    avail_pool = {code: a["qty"] for code, a in aged.items()}

    def _avail_now(comp):
        return sum(avail_pool.get(x, 0.0) for x in aged_codes(comp))

    def _draw(comp, need):
        val = 0.0
        for x in aged_codes(comp):
            if need <= 1e-9:
                break
            take = min(avail_pool.get(x, 0.0), need)
            if take > 0:
                avail_pool[x] -= take
                val += take * aged.get(x, {}).get("cost", 0.0)
                need -= take
        return val

    pending = [fg for fg in fgs if all(avail(aged, c) > 0 for c in fg["components"])]
    recommended, cum, cum_value = [], 0.0, 0.0
    while pending:
        best, best_units, best_consumed = None, 0.0, 0.0
        for fg in pending:
            u = min((_avail_now(c) / c["qty"]) for c in fg["components"] if c["qty"] > 0)
            consumed = u * sum(c["qty"] for c in fg["components"])
            if consumed > best_consumed:
                best, best_units, best_consumed = fg, u, consumed
        if not best or best_consumed <= 1e-6:
            break
        best_value = sum(_draw(c, best_units * c["qty"]) for c in best["components"])
        cum += best_consumed
        cum_value += best_value
        recommended.append({"name": best["name"], "assembly_item": best["assembly_item"],
                            "produce_units": round(best_units, 1), "aged_consumed": round(best_consumed, 1),
                            "aged_value_consumed": round(best_value, 0),
                            "cumulative_aged_consumed": round(cum, 1),
                            "rms_used": [{"rm_desc": c["rm_desc"] if decode else c["rm_code"],
                                          "qty": round(best_units * c["qty"], 1)} for c in best["components"]]})
        pending.remove(best)

    aged_rm = aged
    aged_rm_qty = sum(a["qty"] for a in aged_rm.values())
    aged_rm_value = sum(a["qty"] * a["cost"] for a in aged_rm.values())
    consumed_qty = sum(r["aged_consumed"] for r in recommended)
    leftover = round(aged_rm_qty - consumed_qty, 1)

    used_codes = set()
    for fg in fgs:
        if all(avail(aged, c) > 0 for c in fg["components"]):
            for c in fg["components"]:
                used_codes.update(aged_codes(c))
    unused = sorted(
        [{"rm_code": c, "rm_desc": a["desc"], "qty": round(a["qty"], 1),
          "max_age": a["max_age"], "value": round(a["qty"] * a["cost"], 0)}
         for c, a in aged_rm.items() if c not in used_codes and a["qty"] > 0],
        key=lambda x: -x["qty"])[:200]

    return {
        "aged_days": aged_days, "decode_names": decode,
        "rm_filter": f"Business = {rm_label}",
        "stock_source": "CRM SPBiStockDetails (BiStockDetail)",
        "producible": producible, "blocked": blocked, "recommended": recommended,
        "unused_aged_rm": unused,
        "summary": {
            "aged_rm_items": len(aged_rm), "aged_rm_qty": round(aged_rm_qty, 1),
            "aged_rm_value": round(aged_rm_value, 0),
            "fgs_producible_from_aged": len(producible), "fgs_needing_purchase": len(blocked),
            "recommended_fgs": len(recommended), "aged_consumed_qty": round(consumed_qty, 1),
            "aged_consumed_value": round(cum_value, 0), "aged_left_unused": leftover,
            "utilisation_pct": round(100 * consumed_qty / aged_rm_qty, 1) if aged_rm_qty else 0.0,
        },
    }


def build_aged_rm_report(rp, aged_rows, cons_by_jc, jc_nums, bom_path,
                         business_map=None, settings=None, segment_map=None) -> dict:
    """Aged-RM excess analysis (one row per aged raw material):
      aged qty/value (>aged_days) vs last-3-JC actual consumption, last-3-JC actual-sales
      RM requirement, and the projection RM requirement — with an Excess/OK/Critical status.
    cons_by_jc = [(jc_number, {squash(rm_desc): {...,'qty'}})] for the trailing JCs."""
    from . import planning_settings
    s = settings or planning_settings.load()
    bmap = business_map or {}
    excluded = {x.lower() for x in s.get("excluded_subinv", EXCLUDE_SUBINV)}
    dm_codes = {_norm(c).upper() for c in s.get("dm_water_codes", []) if c}

    def _is_dm(code, desc):
        d = _norm(desc).upper()
        return (_norm(code).upper() in dm_codes or "DM WATER" in d
                or "DEMINERALIZED" in d or "DEMINERALISED" in d)

    # 1. RM universe + Total-FG-used, from the full BOM (chemical components only)
    bom_idx = load_bom_detailed(bom_path) if bom_path else {"by_desc": {}}
    rm_descs: set = set()
    fg_by_rm: dict = {}
    for variants in bom_idx.get("by_desc", {}).values():
        for v in variants:
            fg = _norm(v.get("assembly_desc"))    # keep the FG display name
            for c in v.get("components", []):
                if _pack_code(c.get("comp_code")) or _is_dm(c.get("comp_code"), c.get("comp_desc")):
                    continue
                k = _squash(c.get("comp_desc"))
                if k:
                    rm_descs.add(k)
                    if fg:
                        fg_by_rm.setdefault(k, set()).add(fg)

    # 2. aged RM stock aggregated by squashed description (all orgs; excluded sub-inv out).
    #    ONLY true raw materials: Segment2 = 'Raw Material' (GC1/GC2 chemicals excluded).
    smap = segment_map or {}

    def _seg2(code):
        seg = smap.get(code)
        return " ".join(_norm(seg[1]).split()) if seg else ""    # collapse 'Raw  Material'

    aged: dict = {}
    for r in (aged_rows or []):
        if _norm(r.get("SubInv")).lower() in excluded:
            continue
        code = _norm(r.get("ItemCode"))
        qty = _num(r.get("Qty"))
        if not code or qty <= 0:
            continue
        k = _squash(r.get("ItemDesc"))
        if k not in rm_descs:                     # only chemicals used in a recipe
            continue
        seg2 = _seg2(code)
        if seg2 != "Raw Material":                # keep only Segment2 = Raw Material
            continue
        a = aged.setdefault(k, {"rm_desc": _norm(r.get("ItemDesc")), "category": seg2,
                                "qty": 0.0, "value": 0.0, "max_age": 0})
        a["qty"] += qty
        a["value"] += qty * _num(r.get("ItemCost"))
        a["max_age"] = max(a["max_age"], int(_num(r.get("MaxAgeDays"))))

    # 3. projection RM requirement (plan's 3-JC consolidated RM) by desc
    proj: dict = {}
    for x in rp.get("consolidated_rm", []):
        g = x.get("gross", {})
        proj[_squash(x.get("rm_desc"))] = round(_num(g.get("current")) + _num(g.get("next1")) + _num(g.get("next2")), 1)

    # 4. last-3-JC actual-sales RM requirement (FG 3-JC sales x BOM qty/unit)
    sales: dict = {}
    for p in rp.get("products", []):
        if not p.get("has_bom"):
            continue
        s3 = (p.get("avg_3jc_sales") or 0.0) * 3.0
        if s3 <= 0:
            continue
        for c in p["boms"][0]["components"]:
            k = _squash(c.get("rm_desc"))
            if k:
                sales[k] = sales.get(k, 0.0) + s3 * (c.get("qty_per_unit") or 0.0)

    def _cons(k, cm):
        v = cm.get(k)
        return round((v.get("qty") if isinstance(v, dict) else (v or 0.0)), 1)

    rows = []
    for k, d in aged.items():
        q90 = round(d["qty"], 1)
        if q90 <= 0:
            continue
        jc_cons = [_cons(k, cm) for (_jc, cm) in cons_by_jc]
        cons_sum = round(sum(jc_cons), 1)
        sreq = round(sales.get(k, 0.0), 1)
        preq = round(proj.get(k, 0.0), 1)
        has_plan, has_sales = preq > 0, sreq > 0
        if not has_plan and not has_sales:
            status = "Critical"
            remarks = "No projection demand and no recent sales - aged stock cannot be consumed"
        else:
            cover = max(preq, sreq, cons_sum)
            if q90 > cover:
                status, remarks = "Excess", f"Aged {q90:g} exceeds demand {cover:g}"
            else:
                status, remarks = "OK", "Within demand - consumable"
        rows.append({
            "item_desc": d["rm_desc"], "category": d["category"] or "-",
            "q90": q90, "v90": round(d["value"], 0),
            "jc_consumption": jc_cons, "avg_consumption": round(cons_sum / (len(jc_cons) or 1), 1),
            "excess_avg": round(q90 - cons_sum, 1),
            "excess_avg_pct": round(100 * (q90 - cons_sum) / cons_sum, 1) if cons_sum > 0 else None,
            "rm_req_sales": sreq,
            "excess_sales_pct": round(100 * (q90 - sreq) / sreq, 1) if sreq > 0 else None,
            "fg_used": len(fg_by_rm.get(k, ())),
            "fg_names": ", ".join(sorted(fg_by_rm.get(k, []))[:40]),
            "proj_req": preq,
            "excess_proj": round(q90 - preq, 1),
            "excess_proj_pct": round(100 * (q90 - preq) / preq, 1) if preq > 0 else None,
            "status": status, "remarks": remarks, "max_age": d["max_age"],
        })
    rows.sort(key=lambda r: -r["v90"])
    summary = {"items": len(rows), "total_q90": round(sum(r["q90"] for r in rows), 1),
               "total_v90": round(sum(r["v90"] for r in rows), 0),
               "critical": sum(1 for r in rows if r["status"] == "Critical"),
               "excess": sum(1 for r in rows if r["status"] == "Excess"),
               "ok": sum(1 for r in rows if r["status"] == "OK")}
    return {"rows": rows, "jc_numbers": list(jc_nums), "summary": summary}


# --- 6b. Supplier scorecard ---------------------------------------------------
def _median(xs):
    xs = sorted(xs)
    n = len(xs)
    if not n:
        return 0.0
    m = n // 2
    return xs[m] if n % 2 else (xs[m - 1] + xs[m]) / 2.0


def _clamp(v, lo=0.0, hi=100.0):
    return max(lo, min(hi, v))


def build_supplier_scorecard(paths, settings=None, rows=None) -> dict:
    from . import planning_settings
    s = settings or planning_settings.load()
    tol = float(s.get("otd_tolerance", 0.25))
    crit_lead = float(s.get("critical_lead_days", 30))

    lines: dict[tuple, dict] = {}
    for row in (rows if rows is not None else read_po_rows(paths)):
        code = _norm(row.get("Item Code"))
        vend = _norm(row.get("Vendor Name"))
        if not code or not vend:
            continue
        pono = _norm(row.get("Po Number")) or f"_{code}_{vend}"
        L = lines.setdefault((pono, code), {
            "vendor": vend, "vendor_no": _norm(row.get("Vendor No")), "code": code,
            "name": _norm(row.get("Item Description")), "po_date": None, "po_qty": 0.0,
            "received": 0.0, "last_rcpt": None, "price": 0.0, "amt": 0.0,
            "currency": "", "location": ""})
        pod, rcd = _parse_date(row.get("Po Date")), _parse_date(row.get("Receipt Date"))
        if pod and (L["po_date"] is None or pod < L["po_date"]):
            L["po_date"] = pod
        if rcd and (L["last_rcpt"] is None or rcd > L["last_rcpt"]):
            L["last_rcpt"] = rcd
        L["po_qty"] = max(L["po_qty"], _num(row.get("Po Quantity")))
        L["received"] += _num(row.get("Receipt Qty"))
        L["amt"] += _num(row.get("Rcpt Amt"))
        pr = _num(row.get("Po Unit Price"))
        if pr > 0:
            rate = _num(row.get("Currency Rate")) or 1.0
            L["price"] = pr * rate
        cur = _norm(row.get("Currency Code")).upper()
        if cur:
            L["currency"] = cur
        loc = _norm(row.get("Vendor State")) or _norm(row.get("Vendor Site"))
        if loc:
            L["location"] = loc

    by_item_price: dict[str, list] = {}
    by_item_lead: dict[str, list] = {}
    for L in lines.values():
        lead = ((L["last_rcpt"] - L["po_date"]).days if L["po_date"] and L["last_rcpt"] else None)
        L["lead"] = lead if (lead is not None and 0 <= lead <= 730) else None
        L["fill"] = (L["received"] / L["po_qty"]) if L["po_qty"] > 0 else None
        L["spend"] = round(L["received"] * L["price"], 2)
        if L["price"] > 0:
            by_item_price.setdefault(L["code"], []).append(L["price"])
        if L["lead"] is not None:
            by_item_lead.setdefault(L["code"], []).append(L["lead"])
    item_mkt = {c: _median(v) for c, v in by_item_price.items()}
    item_exp_lead = {c: _median(v) for c, v in by_item_lead.items()}

    sup: dict[str, dict] = {}
    for L in lines.values():
        a = sup.setdefault(L["vendor"], {
            "vendor": L["vendor"], "vendor_no": L["vendor_no"], "lines": 0, "items": set(),
            "currencies": set(), "locations": set(), "leads": [], "ord": 0.0, "recv": 0.0,
            "ontime": 0, "infull": 0, "otif": 0, "elig": 0, "pvm": [], "spend": 0.0, "by_item": {}})
        a["lines"] += 1
        a["items"].add(L["code"])
        if L["currency"]:
            a["currencies"].add(L["currency"])
        if L["location"]:
            a["locations"].add(L["location"])
        a["spend"] += L["spend"]
        if L["lead"] is not None:
            a["leads"].append(L["lead"])
        exp = item_exp_lead.get(L["code"])
        on_time = L["lead"] is not None and exp is not None and L["lead"] <= exp * (1 + tol) + 2
        in_full = L["fill"] is not None and L["fill"] >= 0.999
        if L["po_qty"] > 0 and L["lead"] is not None:
            a["elig"] += 1
            a["ord"] += L["po_qty"]
            a["recv"] += min(L["received"], L["po_qty"])
            if on_time:
                a["ontime"] += 1
            if in_full:
                a["infull"] += 1
            if on_time and in_full:
                a["otif"] += 1
        mkt = item_mkt.get(L["code"])
        pvm = ((L["price"] / mkt - 1) * 100) if (mkt and L["price"] > 0) else None
        if pvm is not None:
            a["pvm"].append(pvm)
        it = a["by_item"].setdefault(L["code"], {"code": L["code"], "name": L["name"], "lines": 0,
                                                 "recv": 0.0, "prices": [], "leads": [], "mkt": mkt, "spend": 0.0})
        it["lines"] += 1
        it["recv"] += L["received"]
        it["spend"] += L["spend"]
        if L["price"] > 0:
            it["prices"].append(L["price"])
        if L["lead"] is not None:
            it["leads"].append(L["lead"])

    item_suppliers: dict[str, set] = {}
    for L in lines.values():
        item_suppliers.setdefault(L["code"], set()).add(L["vendor"])

    suppliers = []
    for a in sup.values():
        elig = a["elig"] or 1
        otd = 100.0 * a["ontime"] / elig
        otif = 100.0 * a["otif"] / elig
        infull = 100.0 * a["infull"] / elig
        fill = 100.0 * a["recv"] / a["ord"] if a["ord"] > 0 else 0.0
        avg_lead = round(sum(a["leads"]) / len(a["leads"]), 1) if a["leads"] else None
        avg_pvm = round(sum(a["pvm"]) / len(a["pvm"]), 1) if a["pvm"] else None
        lead_score = _clamp(100 * (1 - (avg_lead or 0) / 45.0)) if avg_lead is not None else 50.0
        price_score = _clamp(50 - (avg_pvm or 0) * 2.5)
        score = round(0.30 * otif + 0.15 * otd + 0.15 * fill + 0.15 * lead_score + 0.25 * price_score, 1)
        items = []
        for it in a["by_item"].values():
            ip = round(sum(it["prices"]) / len(it["prices"]), 2) if it["prices"] else None
            ipvm = round((ip / it["mkt"] - 1) * 100, 1) if (it["mkt"] and ip) else None
            items.append({"code": it["code"], "name": it["name"], "lines": it["lines"],
                          "received": round(it["recv"], 1), "spend": round(it["spend"], 0),
                          "avg_price": ip, "market_price": round(it["mkt"], 2) if it["mkt"] else None,
                          "price_vs_market": ipvm,
                          "avg_lead": round(sum(it["leads"]) / len(it["leads"]), 1) if it["leads"] else None})
        items.sort(key=lambda x: -x["spend"])
        is_import = bool(a["currencies"] - {"INR", ""})

        sole = [it for it in a["by_item"].values() if len(item_suppliers.get(it["code"], ())) == 1]
        sole.sort(key=lambda it: -it["spend"])
        reasons = []
        if sole:
            reasons.append(f"Sole source for {len(sole)} RM")
        if is_import:
            reasons.append("Import dependency")
        if avg_lead is not None and avg_lead >= crit_lead:
            reasons.append(f"Long lead time ({avg_lead}d)")
        if sole:
            level = "High"
        elif (is_import and (avg_lead or 0) >= 21) or (avg_lead or 0) >= 45:
            level = "Medium"
        elif reasons:
            level = "Low"
        else:
            level = "—"

        suppliers.append({
            "vendor": a["vendor"], "vendor_no": a["vendor_no"], "score": score,
            "otd": round(otd, 1), "otif": round(otif, 1), "in_full": round(infull, 1),
            "fill_rate": round(fill, 1), "avg_lead_time": avg_lead, "price_vs_market": avg_pvm,
            "po_lines": a["lines"], "item_count": len(a["items"]), "spend": round(a["spend"], 0),
            "trade": "Import" if is_import else "Domestic", "currencies": sorted(a["currencies"]),
            "locations": sorted(a["locations"])[:6],
            "critical": level in ("High", "Medium"), "criticality": level,
            "criticality_reasons": reasons, "sole_source_count": len(sole),
            "sole_source_items": [{"code": it["code"], "name": it["name"], "spend": round(it["spend"], 0)}
                                  for it in sole[:20]],
            "items": items[:50],
        })
    suppliers.sort(key=lambda x: (-x["score"], -x["spend"]))
    rated = [s for s in suppliers if s["po_lines"] >= 2]
    return {
        "suppliers": suppliers,
        "summary": {
            "suppliers": len(suppliers), "rated_suppliers": len(rated),
            "po_lines": sum(s["po_lines"] for s in suppliers),
            "items_supplied": len({c for a in sup.values() for c in a["items"]}),
            "imports": sum(1 for s in suppliers if s["trade"] == "Import"),
            "critical": sum(1 for s in suppliers if s["critical"]),
            "sole_source": sum(1 for s in suppliers if s["sole_source_count"] > 0),
            "avg_score": round(sum(s["score"] for s in rated) / len(rated), 1) if rated else 0.0,
            "source": "PO receipts (2-yr)",
            "weights": "OTIF 30% · OTD 15% · Fill 15% · Lead 15% · Price-vs-mkt 25%",
        },
    }


# --- 7. Projection vs Sales ---------------------------------------------------
def _fg_stock_by_name(stock_rows, settings, business_map=None):
    """On-hand FG stock split warehouse/branch, keyed by ITEM DESCRIPTION ONLY.
    Same canonical rule as ``_classify_stock`` (``fg_by_desc``): every non-excluded
    stock lot is summed by its item description across all its item codes — no
    RM/FG classification filter — so warehouse/branch are consistent on every
    planning sheet. Only excluded sub-inventories and zero-qty lots are dropped."""
    s = settings or {}
    warehouse = set(s.get("warehouse_orgs", WAREHOUSE_ORGS))
    excluded = {x.lower() for x in s.get("excluded_subinv", EXCLUDE_SUBINV)}
    out: dict[str, dict] = {}
    for r in stock_rows or []:
        if _norm(r.get("SubInv")).lower() in excluded:
            continue
        qty = _num(r.get("Qty"))
        if qty == 0:
            continue
        name = _norm(r.get("ItemDesc"))
        dk = _squash(name)
        if not dk:
            continue
        a = out.setdefault(dk, {"name": name, "warehouse": 0.0, "branch": 0.0})
        if _norm(r.get("Organization")) in warehouse:
            a["warehouse"] += qty
        else:
            a["branch"] += qty
    return out


def build_projection_vs_sales(bp_path, dispatch_rows, stock_rows, settings=None,
                              business_map=None, accyear=None, n_jc=3, bom_path=None,
                              projection_rows=None) -> dict:
    from . import planning_settings
    s = settings or planning_settings.load()
    band = float(s.get("projection_band", 0.2))
    drop_zero = s.get("drop_all_zero_projection", True)

    mfg_keys = set()
    if bom_path:
        mfg_keys = set(load_bom_detailed(bom_path)["by_squash"].keys())

    # projection source: CRM rows (passed in) take precedence; else the Excel export.
    src = projection_rows if projection_rows is not None else load_projection_rows(bp_path, accyear)
    item_proj: dict[str, dict] = {}
    ci_proj: dict[tuple, dict] = {}
    for r in src:
        key = _squash(r["name"])
        a = item_proj.setdefault(key, {"name": r["name"], "segment2": r["segment2"],
                                       "segment3": r["segment3"], "current": 0.0, "next1": 0.0, "next2": 0.0})
        for k in ("current", "next1", "next2"):
            a[k] += r[k]
        ck = (r["collector"], key)
        b = ci_proj.setdefault(ck, {"collector": r["collector"], "name": r["name"],
                                    "segment2": r["segment2"], "segment3": r["segment3"],
                                    "current": 0.0, "next1": 0.0, "next2": 0.0})
        for k in ("current", "next1", "next2"):
            b[k] += r[k]

    dagg = aggregate_dispatch(dispatch_rows, n_jc)
    d_name, d_ci = dagg["by_name"], dagg["by_collector_item"]
    fg_stock = _fg_stock_by_name(stock_rows, s, business_map)

    def _row(p, avg, with_stock=False):
        cur = round(p["current"], 1)
        is_mfg = _squash(p["name"]) in mfg_keys
        row = {"name": p["name"], "segment2": p["segment2"], "segment3": p["segment3"],
               "current": cur, "next1": round(p["next1"], 1), "next2": round(p["next2"], 1),
               "avg_3jc_sales": round(avg, 1), "variance": round(cur - avg, 1),
               "variance_pct": round(100 * (cur - avg) / avg, 1) if avg > 0 else None,
               "flag": _proj_flag(cur, avg, band),
               "make_or_buy": "make" if is_mfg else "buy",
               "item_type": "Manufactured" if is_mfg else "Traded"}
        if with_stock:
            st = fg_stock.get(_squash(p["name"]), {})
            row["warehouse"] = round(st.get("warehouse", 0.0), 1)
            row["branch"] = round(st.get("branch", 0.0), 1)
            row["stock_total"] = round(row["warehouse"] + row["branch"], 1)
        return row

    items = []
    for key, p in item_proj.items():
        if drop_zero and (p["current"] + p["next1"] + p["next2"]) <= 0:
            continue
        items.append(_row(p, d_name.get(key, {}).get("avg", 0.0), with_stock=True))
    items.sort(key=lambda x: -abs(x["variance"]))

    collector_items = []
    for (coll, key), p in ci_proj.items():
        if drop_zero and (p["current"] + p["next1"] + p["next2"]) <= 0:
            continue
        row = _row(p, d_ci.get((coll, key), {}).get("avg", 0.0))
        row["collector"] = coll
        collector_items.append(row)
    collector_items.sort(key=lambda x: (x["collector"], -abs(x["variance"])))

    def _counts(rows):
        c = {"over": 0, "under": 0, "ontrack": 0, "new": 0, "none": 0}
        for r in rows:
            c[r["flag"]] = c.get(r["flag"], 0) + 1
        return c

    return {
        "n_jc": n_jc, "band_pct": round(band * 100),
        "items": items, "collector_items": collector_items,
        "segments": sorted({i["segment2"] for i in items if i["segment2"]}),
        "summary": {
            "items": len(items), "collector_items": len(collector_items),
            **{f"item_{k}": v for k, v in _counts(items).items()},
            "manufactured": sum(1 for i in items if i["make_or_buy"] == "make"),
            "traded": sum(1 for i in items if i["make_or_buy"] == "buy"),
            "stock_source": "CRM SPBiStockDetails (BiStockDetail)",
            "sales_source": "CRM SP_DespatchDetailsReport (FnDespatchDetails)",
        },
    }


# --- 8. Adhoc planning (post-freeze SOC vs projection + pending SOC) -----------
def _adhoc_status(order_qty, proj_qty, pend_qty, in_proj):
    """Adhoc order validation rules:
      new     -> item not in projection.
      covered -> order within projection (a), or within projection + pending SOC.
      exceeds -> order beyond projection + pending SOC (b); excess is the adhoc qty.
    Returns (status, adhoc_qty)."""
    if not in_proj:
        return "new", round(order_qty, 1)
    if order_qty <= proj_qty:
        return "covered", 0.0
    if order_qty > proj_qty + pend_qty:
        return "exceeds", round(order_qty - (proj_qty + pend_qty), 1)
    return "covered", 0.0


def build_adhoc_planning(soc_rows, bom_path, stock_rows, projected_names=None,
                         settings=None, business_map=None, projection=None,
                         pending=None, allocations=None, freeze_info=None) -> dict:
    from . import planning_settings
    s = settings or planning_settings.load()
    consider_subs = s.get("consider_substitutes", True)
    decode = s.get("decode_encoded_names", True)
    projected = projected_names or set()
    projection = projection or {}
    pending = pending or {}

    bom_idx = load_bom_detailed(bom_path) if bom_path else {"by_desc": {}, "by_squash": {}}
    bom_by_desc, bom_by_squash = bom_idx["by_desc"], bom_idx["by_squash"]

    # Available RM consolidated by item description (req 11), MINUS the RM already
    # allocated to the selected JC plan — so adhoc plans from the remaining stock
    # only (no duplicate allocation between the JC plan and the adhoc plan).
    _excl = {x.lower() for x in s.get("excluded_subinv", EXCLUDE_SUBINV)}
    rm_by_desc: dict[str, float] = {}
    for r in (stock_rows or []):
        if _norm(r.get("SubInv")).lower() in _excl:
            continue
        k = _squash(_norm(r.get("ItemDesc")))
        if k:
            rm_by_desc[k] = rm_by_desc.get(k, 0.0) + _num(r.get("Qty"))
    alloc_by_desc: dict[str, float] = {}
    for a in (allocations or []):
        k = _squash(a.get("rm_desc") or "")
        if k:
            alloc_by_desc[k] = alloc_by_desc.get(k, 0.0) + _num(a.get("allocated_qty"))

    def _avail(desc):
        k = _squash(desc)
        return max(0.0, rm_by_desc.get(k, 0.0) - alloc_by_desc.get(k, 0.0)), k

    products, cons = [], {}
    for r in soc_rows or []:
        name = _norm(r.get("ItemName"))
        order_qty = _num(r.get("SocQty"))
        if not name or order_qty <= 0:
            continue
        key = _squash(name)
        proj_qty = round(projection.get(key, 0.0), 1)
        pend_qty = round(pending.get(key, 0.0), 1)
        in_proj = (key in projection) or (key in projected)
        status, adhoc_qty = _adhoc_status(order_qty, proj_qty, pend_qty, in_proj)
        is_adhoc = adhoc_qty > 0

        variants = bom_by_desc.get(name.upper()) or bom_by_squash.get(key, [])
        mfg = [v for v in variants if not v["is_packing"]]
        sel, _alts = select_bom(mfg or variants, s)
        comps = []
        if sel and is_adhoc:
            for c in sel["components"]:
                gross = round(adhoc_qty * c["qty"], 1)
                subs = [{"code": su["code"], "desc": su["desc"], "stock": round(_avail(su["desc"])[0], 1)}
                        for su in c["substitutes"]]
                main_avail, mkey = _avail(c["comp_desc"])
                seen, sub_avail = {mkey}, 0.0
                for su in c["substitutes"]:
                    sa, sk = _avail(su["desc"])
                    if consider_subs and sk not in seen:
                        sub_avail += sa
                        seen.add(sk)
                avail = round(main_avail + sub_avail, 1)
                comps.append({"rm_code": c["comp_code"], "rm_desc": c["comp_desc"], "qty_per_unit": c["qty"],
                              "gross": gross, "main_stock": round(main_avail, 1),
                              "substitute_stock": round(sub_avail, 1), "available": avail,
                              "net_to_buy": round(max(0.0, gross - avail), 1), "substitutes": subs})
                a = cons.setdefault(mkey, {"rm_code": c["comp_code"], "rm_desc": c["comp_desc"],
                                           "gross": 0.0, "available": avail, "items": set()})
                a["gross"] += gross
                a["available"] = max(a["available"], avail)
                a["items"].add(name)
        products.append({"name": name, "soc_qty": round(order_qty, 1), "soc_count": int(_num(r.get("SocCount"))),
                         "segment2": _norm(r.get("Segment2")), "segment3": _norm(r.get("Segment3")),
                         "projected_qty": proj_qty, "pending_soc_qty": pend_qty, "adhoc_qty": adhoc_qty,
                         "status": status, "is_adhoc": is_adhoc, "has_bom": sel is not None,
                         "components": comps, "net_total": round(sum(c["net_to_buy"] for c in comps), 1)})

    consolidated = []
    for _k, a in cons.items():
        gross, avail = round(a["gross"], 1), round(a["available"], 1)
        consolidated.append({"rm_code": a["rm_code"], "rm_desc": a["rm_desc"], "gross": gross,
                             "main_stock": avail, "substitute_stock": 0.0, "available": avail,
                             "net_to_buy": round(max(0.0, gross - avail), 1),
                             "item_count": len(a["items"]), "items": sorted(a["items"])[:20]})
    consolidated.sort(key=lambda x: -x["net_to_buy"])
    products.sort(key=lambda x: ({"exceeds": 0, "new": 1, "covered": 2}.get(x["status"], 3), -x["adhoc_qty"]))
    adhoc = [p for p in products if p["is_adhoc"]]
    counts = {"exceeds": 0, "new": 0, "covered": 0}
    for p in products:
        counts[p["status"]] = counts.get(p["status"], 0) + 1
    return {
        "products": products, "consolidated_rm": consolidated, "decode_names": decode,
        "freeze": freeze_info or {},
        "summary": {
            "soc_items": len(products), "adhoc_items": len(adhoc),
            "adhoc_with_bom": sum(1 for p in adhoc if p["has_bom"]),
            "consolidated_rms": len(consolidated),
            "rms_to_buy": sum(1 for x in consolidated if x["net_to_buy"] > 0),
            "total_buy_qty": round(sum(x["net_to_buy"] for x in consolidated), 1),
            "adhoc_soc_qty": round(sum(p["adhoc_qty"] for p in adhoc), 1),
            "covered": counts["covered"], "exceeds": counts["exceeds"], "new": counts["new"],
            "rm_allocation_deducted": bool(allocations),
            "source": "CRM SP_SOCDetailReport (open SOC)",
        },
    }


# --- 9. Purchase Price Variance (PPV) -----------------------------------------
def build_ppv(paths, std_fy="2025-26", eval_fy=None, settings=None, rows=None) -> dict:
    from . import jc_calendar
    if eval_fy is None:
        eval_fy = jc_calendar.fiscal_label(date.today())   # current FY (from April)
    recs = []
    std_px: dict[str, list] = {}
    seg_by_code: dict[str, dict] = {}     # Division/Product/Category -> Segment 1/2/3
    for row in (rows if rows is not None else read_po_rows(paths)):
        code = _norm(row.get("Item Code"))
        rcd = _parse_date(row.get("Receipt Date"))
        qty = _num(row.get("Receipt Qty"))
        price = _num(row.get("Po Unit Price"))
        if not code or not rcd or qty <= 0 or price <= 0:
            continue
        sg = seg_by_code.setdefault(code, {"s1": "", "s2": "", "s3": ""})
        if not sg["s1"]:
            sg["s1"] = _norm(row.get("Division"))
        if not sg["s2"]:
            sg["s2"] = _norm(row.get("Product"))
        if not sg["s3"]:
            sg["s3"] = _norm(row.get("Category"))
        price *= (_num(row.get("Currency Rate")) or 1.0)
        fy = jc_calendar.fiscal_label(rcd)
        recs.append((code, _norm(row.get("Item Description")), fy, jc_calendar.fiscal_jc(rcd), price, qty))
        if fy == std_fy:
            a = std_px.setdefault(code, [0.0, 0.0])
            a[0] += price * qty
            a[1] += qty
    std = {c: (v[0] / v[1]) for c, v in std_px.items() if v[1] > 0}

    ij: dict[tuple, list] = {}
    names: dict[str, str] = {}
    jc_agg = {n: {"qty": 0.0, "spend": 0.0, "std_spend": 0.0, "ppv": 0.0} for n in range(1, 14)}
    for code, name, fy, jc, price, qty in recs:
        if fy != eval_fy or code not in std:     # evaluate the current FY vs the std-year WAP
            continue
        names[code] = name
        a = ij.setdefault((code, jc), [0.0, 0.0])
        a[0] += price * qty
        a[1] += qty
        j = jc_agg[jc]
        j["qty"] += qty
        j["spend"] += price * qty
        j["std_spend"] += std[code] * qty
        j["ppv"] += (price - std[code]) * qty

    item: dict[str, dict] = {}
    for (code, jc), (pxq, q) in ij.items():
        wap = pxq / q
        sp = std[code]
        jc_ppv = (wap - sp) * q
        it = item.setdefault(code, {"code": code, "name": names[code], "std": sp, "spend": 0.0,
                                    "qty": 0.0, "waps": [], "above": 0, "below": 0,
                                    "unfav": 0.0, "worst_jc": None, "worst": 0.0})
        it["spend"] += pxq
        it["qty"] += q
        it["waps"].append(wap)
        if wap > sp * 1.001:
            it["above"] += 1
        elif wap < sp * 0.999:
            it["below"] += 1
        if jc_ppv > 0:
            it["unfav"] += jc_ppv
        if jc_ppv > it["worst"]:
            it["worst"], it["worst_jc"] = jc_ppv, jc

    items = []
    for it in item.values():
        sp, waps = it["std"], it["waps"]
        sg = seg_by_code.get(it["code"], {})
        items.append({"code": it["code"], "name": it["name"], "std_price": round(sp, 2),
                      "segment1": sg.get("s1", ""), "segment2": sg.get("s2", ""), "segment3": sg.get("s3", ""),
                      "spend": round(it["spend"], 0), "qty": round(it["qty"], 1),
                      "min_price": round(min(waps), 2), "max_price": round(max(waps), 2),
                      "volatility_pct": round((max(waps) - min(waps)) / sp * 100, 1) if sp else 0.0,
                      "jcs_above": it["above"], "jcs_below": it["below"],
                      "timing_overspend": round(it["unfav"], 0), "worst_jc": it["worst_jc"]})
    items.sort(key=lambda x: -x["timing_overspend"])

    jc_perf = []
    for n in range(1, 14):
        j = jc_agg[n]
        if j["qty"] <= 0:
            continue
        jc_perf.append({"jc": n, "qty": round(j["qty"], 1), "spend": round(j["spend"], 0),
                       "ppv": round(j["ppv"], 0),
                       "ppv_pct": round(100 * j["ppv"] / j["std_spend"], 2) if j["std_spend"] else 0.0,
                       "status": "favourable" if j["ppv"] < 0 else "unfavourable"})

    total_spend = round(sum(jc_agg[n]["spend"] for n in range(1, 14)), 0)
    timing_overspend = round(sum(it["unfav"] for it in item.values()), 0)
    fys = sorted({fy for _, _, fy, _, _, _ in recs})
    eval_from = f"April {eval_fy.split('-')[0]}"
    empty_note = (f"No purchases recorded for FY{eval_fy} (from {eval_from}) yet — the PO "
                  f"receipts currently run up to FY{fys[-1] if fys else '—'}. JC-wise variance "
                  f"will populate once FY{eval_fy} receipts are loaded.") if not jc_perf else None
    return {
        "std_fy": std_fy, "eval_fy": eval_fy, "eval_from": eval_from,
        "fiscal_years": fys,
        "jc_performance": jc_perf, "items": items[:1000],
        "note": empty_note,
        "summary": {
            "std_items": len(std), "evaluated_items": len(item),
            "total_spend": total_spend, "timing_overspend": timing_overspend,
            "timing_overspend_pct": round(100 * timing_overspend / total_spend, 2) if total_spend else 0.0,
            "favourable_jcs": sum(1 for j in jc_perf if j["ppv"] < 0),
            "unfavourable_jcs": sum(1 for j in jc_perf if j["ppv"] > 0),
            "best_jc": min(jc_perf, key=lambda x: x["ppv_pct"])["jc"] if jc_perf else None,
            "worst_jc": max(jc_perf, key=lambda x: x["ppv_pct"])["jc"] if jc_perf else None,
            "note": (f"Standard = weighted-avg price over FY{std_fy}; evaluating FY{eval_fy} "
                     f"purchases (from {eval_from}) JC-wise. PPV = bought above (+, unfavourable) "
                     f"/ below (−, favourable) the FY{std_fy} average."),
        },
    }


# --- 10. Vooki planning (user-input FG quantities) ----------------------------
def load_vooki_master(path) -> dict[str, dict]:
    """Vooki SKU master -> unpack a stocked SKU quantity into consumer units and
    volume (ml). Keyed by item_code. The stock SP reports Vooki FG in packaged
    SKUs; this master converts each SKU to base units / KG-Lit."""
    out: dict[str, dict] = {}
    if not path:
        return out
    with open(path, encoding="utf-8-sig", errors="replace", newline="") as fh:
        for r in csv.DictReader(fh):
            code = _norm(r.get("item_code"))
            if not code:
                continue
            uom = _norm(r.get("uom")).upper()
            pack_type = _norm(r.get("pack_type")).upper()
            unit_ml = _num(r.get("unit_size_ml"))
            pack_qty = _num(r.get("pack_qty")) or 1
            upb = _num(r.get("units_per_box"))
            # consumer units contained in one stocked UOM — counted from CASE/BOX,
            # multipacks (pack) and combos too, not just single units.
            if pack_type in ("CASE", "BOX") or uom in ("BOX", "CASE"):
                units_each = upb or pack_qty or 1
            elif pack_type in ("MULTIPACK_SAME_ITEM", "COMBO_DIFFERENT_ITEMS"):
                units_each = pack_qty or upb or 1
            else:
                units_each = 1
            out[code] = {
                # item description is the planning base identity for Vooki FG
                "name": _norm(r.get("item_desc_original")) or _norm(r.get("item_name_clean")),
                "group": _norm(r.get("item_group_clean")),
                "uom": uom, "unit_ml": unit_ml, "units_each": units_each,
                "pack_type": pack_type, "pack_qty": pack_qty,
                "vol_ml_per_uom": unit_ml * units_each if unit_ml else 0.0,
                "division": _norm(r.get("division")),
            }
    return out


def _classify_vooki_stock(rows, settings=None, business_map=None, rm_codes=None) -> dict:
    """FG = Business 'Vooki Division'; RM = Business 'Raw Material', OR sourced from
    an RM organisation (General Chemicals / intermediates), OR any code used as a
    BOM component (rm_codes) — so every RM in a Vooki BOM shows its real stock.
    Rework/scrap/expired sub-inventories are excluded from both."""
    s = settings or {}
    excluded = {x.lower() for x in s.get("excluded_subinv", EXCLUDE_SUBINV)}
    rm_label = s.get("raw_material_business", "Raw Material")
    fg_business = s.get("vooki_business", "Vooki Division").lower()
    dm_codes = {_norm(c).upper() for c in s.get("dm_water_codes", []) if c}
    business_map = business_map or {}
    rm_codes = rm_codes or set()
    fg: dict[str, dict] = {}
    rm: dict[str, dict] = {}
    rm_business: set = set()
    for row in rows or []:
        if _norm(row.get("SubInv")).lower() in excluded:
            continue
        code = _norm(row.get("ItemCode"))
        if not code:
            continue
        qty = _num(row.get("Qty"))
        if qty == 0:
            continue
        org = _norm(row.get("Organization"))
        name = _norm(row.get("ItemDesc"))
        business = business_map.get(code, "")
        biz_rm = _is_raw_material(business, rm_label)
        # RM planning: only stock physically at MFG organisations counts; DM water dropped.
        is_mfg_org = "mfg" in org.lower()
        is_dm = code.upper() in dm_codes or _squash(name) == "DMWATER"
        if is_mfg_org and not is_dm and (biz_rm or code in rm_codes):
            rm.setdefault(code, {"name": name, "qty": 0.0})["qty"] += qty
            if biz_rm:
                rm_business.add(code)
        if business.strip().lower() == fg_business:
            fg.setdefault(code, {"name": name, "qty": 0.0})["qty"] += qty
    return {"fg": fg, "rm": rm, "rm_business": rm_business}


def build_vooki_planning(bom_path, master_path, stock_rows=None, po_path=None,
                         po_intel=None, settings=None, business_map=None, fg_map=None,
                         extra_fg=None) -> dict:
    """Vooki finished-goods planning. Demand is user-entered per product (applied
    client-side); this returns each Vooki product's selected BOM (+ alternates +
    packing) with per-unit RM requirement, live RM stock, PO received/in-transit,
    and unpacked FG stock so the RM plan recomputes instantly as quantities change.
    """
    from . import planning_settings
    s = settings or planning_settings.load()
    consider_subs = s.get("consider_substitutes", True)
    decode = s.get("decode_encoded_names", True)
    dm_codes = {_norm(c).upper() for c in s.get("dm_water_codes", []) if c}

    def _is_dm(code, desc):
        # DM water is excluded from RM planning (admin code list, or a 'DM WATER' name)
        return _norm(code).upper() in dm_codes or _squash(desc) == "DMWATER"

    master = load_vooki_master(master_path)
    fg_token = str(s.get("vooki_fg_code_token", "BULK")).upper()
    # admin-added FG SKUs (from CRM Vooki Division) extend the scope, regardless
    # of the bulk-token filter.
    extra_codes: dict[str, str] = {}
    for e in (extra_fg or []):
        code = _norm(e.get("code") or e.get("sku_code"))
        if code:
            extra_codes[code] = _norm(e.get("desc") or e.get("item_desc"))

    bom_idx = load_bom_detailed(bom_path) if bom_path else {"by_desc": {}, "by_squash": {}}
    bom_by_desc, bom_by_squash = bom_idx["by_desc"], bom_idx["by_squash"]

    # every RM (+ substitute) used in any BOM — so RM stock is picked up even when
    # an intermediate isn't tagged 'Raw Material' or held in an RM-source org.
    bom_rm_codes = set()
    for variants in bom_by_desc.values():
        for v in variants:
            for comp in v["components"]:
                bom_rm_codes.add(comp["comp_code"])
                for su in comp["substitutes"]:
                    bom_rm_codes.add(su["code"])

    stock = (_classify_vooki_stock(stock_rows, s, business_map, bom_rm_codes)
             if stock_rows is not None else {"fg": {}, "rm": {}})
    fg_stock, rm_stock = stock["fg"], stock["rm"]
    rm_business = stock.get("rm_business", set())
    # Only bulk finished-good SKUs are the planning FG (the manufactured bulk that
    # the BOM produces); packaged variants (250ml/5L/…) are excluded — except any
    # admin-added SKUs, which are always kept.
    if fg_token:
        fg_stock = {c: v for c, v in fg_stock.items()
                    if fg_token in c.upper() or c in extra_codes}
    po = load_po_pending(po_path) if po_path else {}

    # unpack FG stock (packaged SKUs) into units + volume, aggregated by product.
    # An admin mapping (fg_map: sku_code -> product name, from the MySQL store)
    # attributes a SKU to its product; otherwise fall back to name matching.
    fg_map = fg_map or {}
    fg_by_key: dict[str, dict] = {}
    fg_stock_rows: list = []          # per-SKU available FG stock (bulk + units)
    fg_units_total = fg_vol_total = fg_mapped = 0.0
    for code, fsr in fg_stock.items():
        m = master.get(code)
        qty = fsr["qty"]
        name = m["name"] if m else fsr["name"]
        if code in fg_map:
            key = _squash(fg_map[code]); prod_name = fg_map[code]
            fg_mapped += 1
        elif m:
            key = _squash(m["name"]); prod_name = m["name"]
        elif code in extra_codes and extra_codes[code]:
            key = _squash(extra_codes[code]); prod_name = extra_codes[code]
        else:
            key = _squash(name); prod_name = name
        units_each = m["units_each"] if m else 1
        units = qty * units_each
        vol_l = qty * (m["vol_ml_per_uom"] / 1000.0) if (m and m["vol_ml_per_uom"]) else 0.0
        a = fg_by_key.setdefault(key, {"units": 0.0, "volume_l": 0.0, "skus": 0})
        a["units"] += units
        a["volume_l"] += vol_l
        a["skus"] += 1
        fg_units_total += units
        fg_vol_total += vol_l
        fg_stock_rows.append({
            "product": prod_name, "sku_code": code, "sku_name": name,
            "pack_type": (m["pack_type"] if m else ""), "group": (m["group"] if m else ""),
            "uom": (m["uom"] if m else _norm(fsr.get("uom", ""))),
            "unit_ml": round(m["unit_ml"], 1) if m else 0.0,
            "units_each": units_each, "qty": round(qty, 1),
            "units": round(units, 1), "bulk_l": round(vol_l, 1),
        })
    fg_stock_rows.sort(key=lambda x: (x["product"], -x["qty"]))

    # (req 11) consolidate available stock by item DESCRIPTION across ALL item
    # codes and every business classification — the same RM often exists under
    # many codes, some tagged General Chemicals / intermediate rather than 'Raw
    # Material', so pool the full physical stock (excluding rework/scrap subinvs).
    _excl = {x.lower() for x in s.get("excluded_subinv", EXCLUDE_SUBINV)}
    rm_by_desc: dict[str, float] = {}
    for _r in (stock_rows or []):
        if _norm(_r.get("SubInv")).lower() in _excl:
            continue
        if "mfg" not in _norm(_r.get("Organization")).lower():   # RM planning: MFG orgs only
            continue
        _desc = _norm(_r.get("ItemDesc"))
        if _is_dm(_norm(_r.get("ItemCode")), _desc):             # exclude DM water
            continue
        _k = _squash(_desc)
        if _k:
            rm_by_desc[_k] = rm_by_desc.get(_k, 0.0) + _num(_r.get("Qty"))

    def _avail(code, desc):
        """Available stock for an RM, consolidated across all item codes sharing
        its description. Returns (qty, desc_key)."""
        name = rm_stock.get(code, {}).get("name") or desc
        k = _squash(name)
        return rm_by_desc.get(k, rm_stock.get(code, {}).get("qty", 0.0)), k

    def _components(variant):
        rows = []
        for comp in variant["components"]:
            if _is_dm(comp["comp_code"], comp["comp_desc"]):   # DM water excluded from planning
                continue
            main_stock, main_key = _avail(comp["comp_code"], comp["comp_desc"])
            main_stock = round(main_stock, 1)
            # (req 12) substitutes: pool eligible substitute availability, deduped
            # by description so a shared code isn't counted twice.
            seen = {main_key}
            subs, sub_stock = [], 0.0
            for su in comp["substitutes"]:
                st, sk = _avail(su["code"], su["desc"])
                subs.append({"code": su["code"], "desc": su["desc"], "stock": round(st, 1)})
                if consider_subs and sk not in seen:
                    sub_stock += st
                    seen.add(sk)
            sub_stock = round(sub_stock, 1)
            po_item = po.get(comp["comp_code"], {})
            in_transit = round(po_item.get("in_transit", 0.0), 1)
            available = round(main_stock + sub_stock + in_transit, 1)
            # FG units this RM can support = available RM / required consumption per unit
            producible = round(available / comp["qty"], 1) if comp["qty"] else 0.0
            intel = (po_intel or {}).get(comp["comp_code"], {})
            rows.append({
                "seq": comp["seq"], "rm_code": comp["comp_code"], "rm_desc": comp["comp_desc"],
                "qty_per_unit": comp["qty"], "main_stock": main_stock, "substitute_stock": sub_stock,
                "in_transit": in_transit, "received": round(po_item.get("received", 0.0), 1),
                "available": available, "producible": producible,
                "lead_time": intel.get("avg_lead_time_days"), "substitutes": subs,
            })
        return rows

    def _bom_entry(variant, preferred):
        return {"assembly_item": variant["assembly_item"], "org_code": variant["org_code"],
                "designator": variant["designator"], "bom_type": variant.get("bom_type", ""),
                "created": variant.get("created_iso"), "preferred": preferred,
                "components": _components(variant)}

    # BOM lookup by assembly-item code (fallback when the description doesn't match)
    bom_by_ai: dict[str, list] = {}
    for _vs in bom_by_desc.values():
        for v in _vs:
            bom_by_ai.setdefault(v["assembly_item"], []).append(v)

    # ── Real-RM explosion ─────────────────────────────────────────────────────────
    # Some Vooki BOM components are themselves INTERMEDIATES with their own recipe
    # (e.g. PURE INFINITY). For the true purchasing list we recursively explode every
    # such component down to its leaf (purchased) raw materials, tagging the
    # intermediate(s) each leaf was reached through. A component is an intermediate iff
    # its code (or description) is a known MANUFACTURING/REPACK assembly, not a 1:1 alias.
    _REAL_BOM = {"manufacturing", "repack_relabel"}
    _rec_cache: dict = {}

    def _is_alias_bom(v):
        comps = v.get("components", [])
        return (len(comps) == 1 and not _is_packing_comp(comps[0])
                and abs((comps[0].get("qty") or 0.0) - 1.0) < 1e-9)

    def _is_real_recipe(v):
        # a real recipe worth exploding = a MANUFACTURING/REPACK BOM that isn't a 1:1
        # alias. NOTE: the Vooki `is_packing` flag is derived only from the assembly
        # code (non-BULK/HDLK), so an intermediate manufactured under a non-bulk code
        # (e.g. PUREWET on a TD10HDCB… code) is still a real recipe — do not exclude it.
        return v.get("bom_class") in _REAL_BOM and not _is_alias_bom(v)

    def _sel_recipe(code, desc):
        # match an intermediate to explode — by code: any real recipe; by description
        # (fallback): only a manufacturing recipe (never a repack, which shares its
        # description with its own bulk base and would loop back on itself).
        by_code = [v for v in bom_by_ai.get(code, []) if _is_real_recipe(v)]
        vlist = by_code or [v for v in bom_by_squash.get(_squash(desc), [])
                            if v.get("bom_class") == "manufacturing" and not _is_alias_bom(v)]
        if not vlist:
            return None                       # leaf RM (no real recipe)
        ck = code or _squash(desc)
        if ck not in _rec_cache:
            mfg = [v for v in vlist if v.get("bom_class") == "manufacturing"]
            sel, _ = select_bom(mfg or vlist, s)
            _rec_cache[ck] = sel
        return _rec_cache[ck]

    def _looks_encoded(name):
        return bool(_ENC_NAME_RE.match(_norm(name).upper()))

    def _decode_name(code, desc, depth=0, seen=frozenset()):
        if not _looks_encoded(desc):
            return desc
        node = _squash(desc) or code
        if depth > 8 or node in seen:
            return desc
        vlist = bom_by_ai.get(code) or bom_by_squash.get(_squash(desc)) or []
        aliases = [v for v in vlist if _is_alias_bom(v) or v.get("bom_class") == "internal"]
        if not aliases:
            return desc
        nonpack = [c for c in aliases[0]["components"] if not _is_packing_comp(c)]
        if not nonpack:
            return desc
        return _decode_name(nonpack[0]["comp_code"], nonpack[0]["comp_desc"], depth + 1, seen | {node})

    def _dec(code, desc):
        return _decode_name(code, desc) if decode else desc

    def _explode_unit(code, desc, mult, depth=0, seen=frozenset(), trail=()):
        """Yield (leaf_code, leaf_desc, per_unit, via_intermediates, unresolved) for a
        component: any intermediate is exploded to its leaf RMs. per_unit = qty of the
        leaf per 1 unit of the finished good; via = the intermediate names traversed."""
        recipe = _sel_recipe(code, desc)
        node = code or _squash(desc)
        if recipe is None:
            yield (code, desc, mult, trail, False)          # genuine purchased leaf RM
            return
        if depth > 12 or node in seen:                      # has a recipe but can't resolve
            yield (code, desc, mult, trail, True)           # unresolved intermediate
            return
        nxt = trail + (recipe["assembly_desc"],)
        for comp in recipe["components"]:
            if _is_packing_comp(comp):
                continue
            yield from _explode_unit(comp["comp_code"], comp["comp_desc"],
                                     mult * (comp.get("qty") or 0.0), depth + 1,
                                     seen | {node}, nxt)

    def _real_rm_for(variant):
        """Per-unit leaf-RM requirement for a selected BOM — every intermediate
        exploded. Consolidated by leaf description; per_unit scales with the FG qty."""
        agg: dict = {}
        for comp in variant["components"]:
            if _is_packing_comp(comp):
                continue
            for lc, ld, lq, trail, unresolved in _explode_unit(comp["comp_code"], comp["comp_desc"], comp.get("qty") or 0.0):
                ldesc = _dec(lc, ld)
                if _is_dm(lc, ldesc):        # DM water excluded from planning
                    continue
                key = _squash(ldesc) or lc
                avail, _ = _avail(lc, ldesc)
                e = agg.setdefault(key, {"code": lc, "desc": ldesc, "per_unit": 0.0,
                                         "available": round(avail, 1), "via": set(), "unresolved": False})
                e["per_unit"] += lq
                e["via"].update(trail)
                if unresolved:
                    e["unresolved"] = True
        return [{"code": v["code"], "desc": v["desc"], "per_unit": round(v["per_unit"], 6),
                 "available": v["available"], "via": sorted(v["via"])[:6], "unresolved": v["unresolved"]}
                for v in sorted(agg.values(), key=lambda x: -x["per_unit"])]

    # Base planning items = the Vooki FG item descriptions (the admin FG SKU list:
    # bulk master SKUs + admin-added). Item description is the base identity; each
    # base is matched to a BOM by description / squash / assembly-item code.
    base_items: dict[str, dict] = {}

    def _add_base(desc, code):
        if not desc:
            return
        b = base_items.setdefault(_squash(desc), {"desc": desc, "codes": set()})
        b["codes"].add(code)

    for code, m in master.items():
        if fg_token and fg_token not in code.upper() and code not in extra_codes:
            continue
        _add_base(m["name"], code)
    for code, desc in extra_codes.items():
        _add_base(desc or (master.get(code) or {}).get("name") or code, code)

    products = []
    for k, b in base_items.items():
        desc = b["desc"]
        variants = list(bom_by_desc.get(desc.upper()) or bom_by_squash.get(k) or [])
        if not variants:
            for c in b["codes"]:
                variants.extend(bom_by_ai.get(c, []))
        mfg = [v for v in variants if not v["is_packing"]]
        packing = [v for v in variants if v["is_packing"]]
        if mfg:
            selected, alts = select_bom(mfg, s)
        elif packing:
            selected, alts = select_bom(packing, s)
            packing = []
        else:
            selected, alts = None, []
        fga = fg_by_key.get(k) or {}
        # per-unit real (leaf) RM requirement — intermediates exploded to purchased RMs
        real_rm = _real_rm_for(selected) if selected else []
        # producible NOW with current on-hand MFG RM stock: the bottleneck leaf RM
        # (available / per-unit). Intermediates are already exploded, so an out-of-stock
        # intermediate no longer blocks it if its raw materials are available.
        prod_now, limiting, lim_avail = None, "", None
        for e in real_rm:
            if e["per_unit"] > 0:
                cap = e["available"] / e["per_unit"]
                if prod_now is None or cap < prod_now:
                    prod_now, limiting, lim_avail = cap, e["desc"], e["available"]
        products.append({
            "name": desc, "has_bom": selected is not None, "alternatives": len(alts),
            "fg_units": round(fga.get("units", 0.0), 1),
            "fg_volume_l": round(fga.get("volume_l", 0.0), 1),
            "boms": ([_bom_entry(v, i == 0) for i, v in enumerate([selected] + alts)]
                     if selected else []),
            "packing_boms": ([_bom_entry(v, False) for v in sorted(
                packing, key=lambda x: (x["org_code"] != "PMO", x["assembly_item"]))]
                if selected else []),
            "real_rm": real_rm,
            "producible_now": round(prod_now, 1) if prod_now is not None else 0.0,
            "limiting_rm": limiting,
            "limiting_rm_available": round(lim_avail, 1) if lim_avail is not None else None,
        })
    products.sort(key=lambda p: (not p["has_bom"], p["name"]))

    # candidate Vooki FG SKUs for the admin name->SKU mapping (master ∪ in-stock),
    # restricted to bulk item codes only.
    fg_skus, seen_sku = [], set()
    for code, m in master.items():
        if fg_token and fg_token not in code.upper() and code not in extra_codes:
            continue
        fg_skus.append({"code": code, "name": m["name"], "group": m["group"],
                        "in_stock": round(fg_stock.get(code, {}).get("qty", 0.0), 1),
                        "added": code in extra_codes})
        seen_sku.add(code)
    for code, desc in extra_codes.items():        # admin-added SKUs not in the master
        if code not in seen_sku:
            fg_skus.append({"code": code, "name": desc or code, "group": "",
                            "in_stock": round(fg_stock.get(code, {}).get("qty", 0.0), 1), "added": True})
            seen_sku.add(code)
    for code, fsr in fg_stock.items():
        if code not in seen_sku:
            fg_skus.append({"code": code, "name": fsr["name"], "group": "",
                            "in_stock": round(fsr["qty"], 1), "added": False})
            seen_sku.add(code)
    fg_skus.sort(key=lambda x: x["name"])

    return {
        "products": products, "decode_names": decode,
        "fg_stock": {"total_units": round(fg_units_total, 1),
                     "total_volume_l": round(fg_vol_total, 1)},
        "fg_stock_rows": fg_stock_rows,
        "fg_skus": fg_skus,
        "summary": {
            "products": len(products),
            "with_packing_bom": sum(1 for p in products if p["packing_boms"]),
            "rm_items_in_stock": len(rm_business), "rm_items_in_stock_all": len(rm_stock),
            "fg_skus_in_stock": len(fg_stock),
            "po_pending_items": sum(1 for k, v in po.items() if (v.get("in_transit") or 0) > 0 and not _pack_code(k)),
            "po_items_total": len(po), "master_skus": len(master),
            "fg_stock_units": round(fg_units_total, 1),
            "fg_stock_volume_l": round(fg_vol_total, 1),
            "fg_skus_mapped": int(fg_mapped),
            "stock_source": "CRM SPBiStockDetails (BiStockDetail)",
        },
        "rules": {
            "fg_business": s.get("vooki_business", "Vooki Division"),
            "rm_business": s.get("raw_material_business", "Raw Material"),
            "rm_source_orgs": sorted(s.get("rm_source_orgs", RM_SOURCE_ORGS)),
            "excluded_subinv": sorted(s.get("excluded_subinv", EXCLUDE_SUBINV)),
            "bom_preference": " -> ".join(
                (["PMO"] if s.get("bom_prefer_pmo", True) else [])
                + (["BULK/HDLK"] if s.get("bom_prefer_bulk_hdlk", True) else [])
                + (["newest BOM"] if s.get("bom_prefer_creation_date", True) else [])
                + (["Primary"] if s.get("bom_prefer_primary", True) else [])) or "(no preference)",
            "consider_substitutes": consider_subs,
        },
    }

