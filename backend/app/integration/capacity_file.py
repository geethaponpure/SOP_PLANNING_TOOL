"""Capacity / cycle-time loader (PPS Products Cycle Time workbook).

Provides, per product (matched to CRM items BY NAME): the equipment it runs on,
min/max batch size (KG) and the batch cycle time (hours incl. cleaning). From
these we derive a run rate (KG/hr) and an equipment-based rough-cut capacity
check (Section 10.3) and real batch lot-sizing (Section 10.4).

The file does NOT carry available hours per equipment, so that is supplied via
EQUIPMENT_HOURS_PER_MONTH (default below) and can be overridden in .env.
"""
from __future__ import annotations

import os

COLS = {
    "division": 1, "category": 2, "product": 3, "equipment": 4,
    "min_batch": 6, "max_batch": 7, "cycle_time": 17,  # cycle time incl. cleaning
}


def _num(v) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _find_header(ws) -> int:
    for i, r in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True)):
        labels = [str(x).strip().upper() if x else "" for x in r]
        if "PRODUCT NAME" in labels and "EQUIPMENT ID" in labels:
            return i
    return 2


def load_cycle_time(path: str) -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    hi = _find_header(ws)
    by_product: dict[str, dict] = {}
    equipment: set[str] = set()
    for r in ws.iter_rows(min_row=hi + 2, values_only=True):
        if not r or len(r) <= COLS["cycle_time"]:
            continue
        prod = r[COLS["product"]]
        equip = r[COLS["equipment"]]
        if not prod or not equip:
            continue
        max_batch = _num(r[COLS["max_batch"]]) or _num(r[COLS["min_batch"]])
        cycle = _num(r[COLS["cycle_time"]])
        rate = (max_batch / cycle) if (max_batch and cycle) else 0.0  # KG/hr
        by_product[str(prod).strip().upper()] = {
            "equipment": str(equip).strip(),
            "min_batch": _num(r[COLS["min_batch"]]),
            "max_batch": max_batch,
            "cycle_time_hrs": cycle,
            "run_rate_kg_hr": round(rate, 2),
            "division": r[COLS["division"]],
        }
        equipment.add(str(equip).strip())

    hours = float(os.getenv("EQUIPMENT_HOURS_PER_MONTH", "416"))  # ~16h x 26d
    assets = {e: {"name": e, "hours": hours} for e in sorted(equipment)}
    return {"by_product": by_product, "assets": assets}
